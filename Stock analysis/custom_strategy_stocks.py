import pandas as pd
import numpy as np
import itertools
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
import os

def create_excel_with_top5_sheets(base_name, input_file, df, open_arr, close_arr, ma20_arr, ma60_arr, eom_arr, eomsig_arr, mfi_arr, macd_arr, strategy_name, param_names, param_grid, entry_condition_func, exit_condition_func, formula_generator):
    output_file = f"E:\\G-AI-1\\Stock analysis\\{base_name}_{strategy_name}_最佳化與回測.xlsx"
    print(f"=============================================")
    print(f"正在處理: {base_name} - {strategy_name} ...")
    
    shares = 1000
    len_df = len(df)
    results = []
    
    keys, values = zip(*param_grid.items())
    combinations = [dict(zip(keys, v)) for v in itertools.product(*values)]
    
    print("  計算最佳化參數組合...")
    for params in combinations:
        hold, buy_price = False, 0.0
        pnls, entry_prices = [], []
        
        for i in range(1, len_df - 1):
            if not hold:
                if entry_condition_func(i, close_arr, ma20_arr, ma60_arr, eom_arr, eomsig_arr, mfi_arr, macd_arr, params):
                    hold = True
                    buy_price = open_arr[i+1]
                    entry_prices.append(buy_price)
            else:
                n_open = open_arr[i+1]
                pnl_ratio = (n_open - buy_price) / buy_price if buy_price > 0 else 0
                
                # Dynamic Exit checking
                if exit_condition_func(i, pnl_ratio, close_arr, ma20_arr, ma60_arr, eom_arr, eomsig_arr, mfi_arr, macd_arr, params):
                    hold = False
                    sell_price = n_open
                    pnls.append((sell_price - buy_price) * shares)
                    buy_price = 0.0
                    
        if len(pnls) > 0:
            tc = len(pnls)
            cp = sum(pnls)
            mc = max(entry_prices) * shares if entry_prices else 0
            
            res = params.copy()
            res.update({
                '累積淨損益總額': cp, '最大所需本金': mc,
                '真實投報率(ROI)': cp / mc if mc > 0 else 0,
                '交易次數': tc, '勝率': len([p for p in pnls if p > 0]) / tc,
                '單筆最大獲利': max(pnls), '單筆最大虧損': min(pnls),
                '平均每筆損益': cp / tc,
                '期望報酬率(每筆)': (cp / tc) / ((sum(entry_prices)/len(entry_prices))*shares) if entry_prices else 0
            })
            results.append(res)
            
    res_df = pd.DataFrame(results)
    res_df = res_df[res_df['交易次數'] >= 2].sort_values(by='累積淨損益總額', ascending=False).reset_index(drop=True)
    
    def generate_comment(row, rank):
        if rank == 1: return "🏆 最高報酬：適合捕捉大趨勢。"
        elif rank <= 5:
            if row['勝率'] > 0.6: return "⭐ 高勝率：穩定度佳。"
            elif row['單筆最大虧損'] > -30000: return "🛡️ 低風險：最大回檔小。"
            elif row['真實投報率(ROI)'] > 0.15: return "🚀 高資金效率：投報率優越。"
            else: return "💡 綜合優異：平衡度高。"
        return ""
    
    res_df['最佳參數註解'] = [generate_comment(row, i+1) for i, row in res_df.iterrows()]
    
    wb = Workbook()
    ws_opt = wb.active
    ws_opt.title = "參數最佳化結果"
    
    header_names = param_names + [
        '累積淨損益總額', '最大所需本金', '真實投報率(ROI)',
        '交易次數', '勝率', '單筆最大獲利', '單筆最大虧損',
        '平均每筆損益', '期望報酬率(每筆)', '🤖 AI 策略評估與註解'
    ]
    ws_opt.append(header_names)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E40AF")
    top5_fill = PatternFill("solid", fgColor="FEF08A")
    for cell in ws_opt[1]:
        cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="center")
        
    for r_idx, row in res_df.iterrows():
        ws_opt.append([row[col] for col in res_df.columns])
        cr = ws_opt.max_row
        for col_idx, col_name in enumerate(param_names, 1):
            if 'TP' in col_name or 'SL' in col_name or '%' in col_name: ws_opt.cell(row=cr, column=col_idx).number_format = '0.0%'
            
        offset = len(param_names)
        for col in [1, 2, 6, 7, 8]: ws_opt.cell(row=cr, column=offset+col).number_format = '#,##0'
        for col in [3, 5, 9]: ws_opt.cell(row=cr, column=offset+col).number_format = '0.00%'
        
        if r_idx < 5:
            for col_num in range(1, len(header_names) + 1):
                ws_opt.cell(row=cr, column=col_num).fill = top5_fill
                ws_opt.cell(row=cr, column=col_num).font = Font(bold=True)
                
    for col in ws_opt.columns:
        ml = max([len(str(cell.value)) for cell in col if cell.value] + [0])
        ws_opt.column_dimensions[col[0].column_letter].width = min(ml + 2, 40)
        
    print("  建立 Top 5 回測分頁...")
    headers = list(df.columns) + ["進場訊號(P)", "出場訊號(Q)", "持倉狀態(R)", "進場成本(S)", "動作(T)", "交易價格(U)", "單筆損益(V)", "累積損益(W)"]
    
    for top_idx in range(min(5, len(res_df))):
        ws_data = wb.create_sheet(f"Top{top_idx+1}_回測")
        best_params = res_df.iloc[top_idx]
        
        ws_data['A1'] = f"【{strategy_name} - Top{top_idx+1} 參數】"
        ws_data['B1'] = "數值 (可直接修改)"
        for idx, p_name in enumerate(param_names, 2):
            ws_data[f'A{idx}'] = p_name
            ws_data[f'B{idx}'] = best_params[p_name]
        
        var_count = len(param_names)
        shares_row = var_count + 2
        ws_data[f'A{shares_row}'] = "交易張數 Shares"
        ws_data[f'B{shares_row}'] = 1000
        
        for i, title in enumerate(headers, 1): ws_data.cell(row=8, column=i, value=title)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 9):
            for c_idx, value in enumerate(row, 1): ws_data.cell(row=r_idx, column=c_idx, value=value)
            
        formula_generator(ws_data, 9, ws_data.max_row, param_names, shares_row)
        
        for col in ['A', 'B']:
            for row in range(1, shares_row + 1):
                ws_data[f'{col}{row}'].font = Font(bold=True)
                ws_data[f'{col}{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    wb.save(output_file)
    print(f"  > 輸出完成: {output_file}")


def process_3443():
    # 創意 3443 專屬策略：高波動趨勢動能爆發
    # 進場：Close > MA60 (趨勢向上) 且 EOM > EOM_Sig (動能爆發) 且 MACD > MACD_Th
    # 出場：停利 或 停損 或 (EOM < EOM_Sig 動能衰退)
    input_file = "E:\\G-AI-1\\Stock analysis\\3443_創意_EOM_20260429.xlsx"
    df = pd.read_excel(input_file)
    df.columns = [str(c) for c in df.columns]
    
    open_col, close_col = df.columns[1], df.columns[4]
    ma20_col, ma60_col = df.columns[7], df.columns[8]
    eom_col, eomsig_col = df.columns[10], df.columns[11]
    mfi_col, macd_col = df.columns[12], df.columns[14]
    
    for col in [open_col, close_col, ma20_col, ma60_col, eom_col, eomsig_col, mfi_col, macd_col]:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    open_arr = df[open_col].values; close_arr = df[close_col].values
    ma20_arr = df[ma20_col].values; ma60_arr = df[ma60_col].values
    eom_arr = df[eom_col].values; eomsig_arr = df[eomsig_col].values
    mfi_arr = df[mfi_col].values; macd_arr = df[macd_col].values

    param_names = ['TP(停利)', 'SL(停損)', 'MACD_Th(MACD門檻)']
    param_grid = {
        'TP(停利)': [0.1, 0.15, 0.20],
        'SL(停損)': [0.04, 0.06, 0.08],
        'MACD_Th(MACD門檻)': [-1.0, 0.0, 1.0]
    }
    
    def entry_cond(i, c, m20, m60, e, esig, mfi, macd, p):
        return c[i] > m60[i] and e[i] > esig[i] and macd[i] > p['MACD_Th(MACD門檻)']
        
    def exit_cond(i, pnl, c, m20, m60, e, esig, mfi, macd, p):
        if pnl >= p['TP(停利)'] or pnl <= -p['SL(停損)']: return True
        if e[i] < esig[i]: return True # 動能衰退
        return False
        
    def formulas(ws, start_row, max_row, p_names, shares_row):
        # B=Open, E=Close, I=MA60, K=EOM, L=EOM_Sig, O=MACD
        for r in range(start_row, max_row + 1):
            prev_R, prev_P, prev_S, prev_W = ("FALSE", "FALSE", "0", "0") if r == start_row else (f"R{r-1}", f"P{r-1}", f"S{r-1}", f"W{r-1}")
            # P: AND(E>I, K>L, O>$B$4)
            ws[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}>I{r}, K{r}>L{r}, O{r}>$B$4, {prev_R}=FALSE))'
            # Q: TP/SL or K<L.
            ws[f'Q{r}'] = f'=IF({prev_R}=TRUE, OR((B{r}-{prev_S})/{prev_S} >= $B$2, (B{r}-{prev_S})/{prev_S} <= -$B$3, K{r}<L{r}), FALSE)'
            ws[f'R{r}'] = f'=IF({prev_R}=FALSE, {prev_P}=TRUE, IF(Q{r}=TRUE, FALSE, TRUE))'
            ws[f'S{r}'] = f'=IF(AND({prev_R}=FALSE, {prev_P}=TRUE), B{r}, IF(AND({prev_R}=TRUE, Q{r}=FALSE), {prev_S}, 0))'
            ws[f'T{r}'] = f'=IF(AND({prev_R}=FALSE, {prev_P}=TRUE), "買進", IF(Q{r}=TRUE, "賣出", ""))'
            ws[f'U{r}'] = f'=IF(T{r}="買進", B{r}, IF(T{r}="賣出", B{r}, 0))'
            ws[f'V{r}'] = f'=IF(T{r}="賣出", (U{r} - {prev_S}) * $B${shares_row}, 0)'
            ws[f'W{r}'] = f'={prev_W} + V{r}'

    create_excel_with_top5_sheets("3443_創意", input_file, df, open_arr, close_arr, ma20_arr, ma60_arr, eom_arr, eomsig_arr, mfi_arr, macd_arr, "高波動趨勢動能爆發", param_names, param_grid, entry_cond, exit_cond, formulas)

def process_6269():
    # 台郡 6269 專屬策略：波段反轉低接策略
    # 進場：Close > MA20 (短趨勢轉強) 且 MFI < MFI_Oversold (超賣區反彈) 且 MACD > MACD_Th
    # 出場：停利 或 停損 或 MFI > MFI_Overbought (超買區獲利了結)
    input_file = "E:\\G-AI-1\\Stock analysis\\6269_台郡_EOM_20260429.xlsx"
    df = pd.read_excel(input_file)
    df.columns = [str(c) for c in df.columns]
    
    open_col, close_col = df.columns[1], df.columns[4]
    ma20_col, ma60_col = df.columns[7], df.columns[8]
    eom_col, eomsig_col = df.columns[10], df.columns[11]
    mfi_col, macd_col = df.columns[12], df.columns[14]
    
    for col in [open_col, close_col, ma20_col, ma60_col, eom_col, eomsig_col, mfi_col, macd_col]:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    open_arr = df[open_col].values; close_arr = df[close_col].values
    ma20_arr = df[ma20_col].values; ma60_arr = df[ma60_col].values
    eom_arr = df[eom_col].values; eomsig_arr = df[eomsig_col].values
    mfi_arr = df[mfi_col].values; macd_arr = df[macd_col].values

    param_names = ['TP(停利)', 'SL(停損)', 'MFI_超賣進場', 'MFI_超買出場', 'MACD_Th']
    param_grid = {
        'TP(停利)': [0.05, 0.08, 0.12],
        'SL(停損)': [0.03, 0.05, 0.07],
        'MFI_超賣進場': [35, 40, 45],
        'MFI_超買出場': [70, 75, 80],
        'MACD_Th': [-0.5, 0.0]
    }
    
    def entry_cond(i, c, m20, m60, e, esig, mfi, macd, p):
        return c[i] > m20[i] and mfi[i] < p['MFI_超賣進場'] and macd[i] > p['MACD_Th']
        
    def exit_cond(i, pnl, c, m20, m60, e, esig, mfi, macd, p):
        if pnl >= p['TP(停利)'] or pnl <= -p['SL(停損)']: return True
        if mfi[i] > p['MFI_超買出場']: return True 
        return False
        
    def formulas(ws, start_row, max_row, p_names, shares_row):
        # B=Open, E=Close, H=MA20, M=MFI, O=MACD
        for r in range(start_row, max_row + 1):
            prev_R, prev_P, prev_S, prev_W = ("FALSE", "FALSE", "0", "0") if r == start_row else (f"R{r-1}", f"P{r-1}", f"S{r-1}", f"W{r-1}")
            # P: AND(E>H, M<$B$4, O>$B$6)  -- Check row mappings: B2=TP, B3=SL, B4=MFI_OverS, B5=MFI_OverB, B6=MACD
            ws[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}>H{r}, M{r}<$B$4, O{r}>$B$6, {prev_R}=FALSE))'
            # Q: TP/SL or M>$B$5.
            ws[f'Q{r}'] = f'=IF({prev_R}=TRUE, OR((B{r}-{prev_S})/{prev_S} >= $B$2, (B{r}-{prev_S})/{prev_S} <= -$B$3, M{r}>$B$5), FALSE)'
            ws[f'R{r}'] = f'=IF({prev_R}=FALSE, {prev_P}=TRUE, IF(Q{r}=TRUE, FALSE, TRUE))'
            ws[f'S{r}'] = f'=IF(AND({prev_R}=FALSE, {prev_P}=TRUE), B{r}, IF(AND({prev_R}=TRUE, Q{r}=FALSE), {prev_S}, 0))'
            ws[f'T{r}'] = f'=IF(AND({prev_R}=FALSE, {prev_P}=TRUE), "買進", IF(Q{r}=TRUE, "賣出", ""))'
            ws[f'U{r}'] = f'=IF(T{r}="買進", B{r}, IF(T{r}="賣出", B{r}, 0))'
            ws[f'V{r}'] = f'=IF(T{r}="賣出", (U{r} - {prev_S}) * $B${shares_row}, 0)'
            ws[f'W{r}'] = f'={prev_W} + V{r}'

    create_excel_with_top5_sheets("6269_台郡", input_file, df, open_arr, close_arr, ma20_arr, ma60_arr, eom_arr, eomsig_arr, mfi_arr, macd_arr, "波段反轉低接", param_names, param_grid, entry_cond, exit_cond, formulas)

if __name__ == "__main__":
    process_3443()
    process_6269()
    print("全部獨立策略分析完畢！")
