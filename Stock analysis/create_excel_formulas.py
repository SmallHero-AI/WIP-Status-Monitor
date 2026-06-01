import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = "E:\\G-AI-1\\Stock analysis\\2330_台積電_EOM_20260430.xlsx"
output_file = "E:\\G-AI-1\\Stock analysis\\2330_台積電_公式回測_可修改參數.xlsx"

print(f"讀取原始資料: {input_file}")
df = pd.read_excel(input_file)

wb = Workbook()
ws = wb.active
ws.title = "Backtest Strategy"

# 1. 寫入參數設定區 (Rows 1 to 6)
ws['A1'] = "【策略參數設定區】"
ws['B1'] = "數值 (可直接修改)"

ws['A2'] = "停利 Take Profit (%)"
ws['B2'] = 0.09

ws['A3'] = "停損 Stop Loss (%)"
ws['B3'] = 0.046

ws['A4'] = "MACD 閾值"
ws['B4'] = 0.0

ws['A5'] = "MFI 閾值"
ws['B5'] = 40.0

ws['A6'] = "交易張數 Shares"
ws['B6'] = 1000

# 2. 寫入資料表頭 (Row 8)
columns = list(df.columns)
headers = columns + ["進場訊號(P)", "出場訊號(Q)", "持倉狀態(R)", "進場成本(S)", "動作(T)", "交易價格(U)", "單筆損益(V)", "累積損益(W)"]
for col_idx, title in enumerate(headers, 1):
    ws.cell(row=8, column=col_idx, value=title)

# 3. 寫入歷史資料 (Starting from Row 9)
print("正在寫入資料與公式...")
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 9):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# 4. 寫入 Excel 公式
max_row = ws.max_row

# 欄位對應:
# B=Open, E=Close, H=MA20, K=EOM, L=EOM_Sig, M=MFI, O=MACD
for r in range(9, max_row + 1):
    if r == 9:
        prev_R = "FALSE"
        prev_P = "FALSE"
        prev_S = "0"
        prev_W = "0"
    else:
        prev_r = r - 1
        prev_R = f"R{prev_r}"
        prev_P = f"P{prev_r}"
        prev_S = f"S{prev_r}"
        prev_W = f"W{prev_r}"
    
    # P: 進場訊號 (Buy Signal)
    ws[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}>H{r}, K{r}>L{r}, O{r}>$B$4, M{r}>$B$5, {prev_R}=FALSE))'
    
    # Q: 出場訊號 (Sell Signal)
    ws[f'Q{r}'] = f'=IF({prev_R}=TRUE, OR((B{r}-{prev_S})/{prev_S} >= $B$2, (B{r}-{prev_S})/{prev_S} <= -$B$3), FALSE)'
    
    # R: 持倉狀態 (Hold Status)
    ws[f'R{r}'] = f'=IF({prev_R}=FALSE, {prev_P}=TRUE, IF(Q{r}=TRUE, FALSE, TRUE))'
    
    # S: 進場成本 (Entry Price)
    ws[f'S{r}'] = f'=IF(AND({prev_R}=FALSE, {prev_P}=TRUE), B{r}, IF(AND({prev_R}=TRUE, Q{r}=FALSE), {prev_S}, 0))'
    
    # T: 動作 (Action)
    ws[f'T{r}'] = f'=IF(AND({prev_R}=FALSE, {prev_P}=TRUE), "買進", IF(Q{r}=TRUE, "賣出", ""))'
    
    # U: 交易價格 (Trade Price)
    ws[f'U{r}'] = f'=IF(T{r}="買進", B{r}, IF(T{r}="賣出", B{r}, 0))'
    
    # V: 單筆損益 (PnL)
    ws[f'V{r}'] = f'=IF(T{r}="賣出", (U{r} - {prev_S}) * $B$6, 0)'
    
    # W: 累積損益 (Cumulative PnL)
    ws[f'W{r}'] = f'={prev_W} + V{r}'

# 設定一些簡單的樣式
from openpyxl.styles import Font, PatternFill
for col in ['A', 'B']:
    for row in range(1, 7):
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

wb.save(output_file)
print(f"完成！包含動態公式的回測檔案已儲存至: {output_file}")
