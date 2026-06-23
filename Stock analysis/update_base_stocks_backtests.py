# -*- coding: utf-8 -*-
import os
import glob
import re
import sys
import pandas as pd
import openpyxl

# 解決 Windows 主機 Console 輸出 Unicode Emojis 編碼錯誤
if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except AttributeError:
        pass
if sys.stderr.encoding != 'utf-8':
    try:
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        pass

SCRIPT_DIR = r"E:\G-AI-1\Stock analysis"
AUTO_EXPORT_DIR = os.path.join(SCRIPT_DIR, "Stock original", "auto_export")
DATA_DIR = os.path.join(SCRIPT_DIR, "修正版_V6_Server", "data")

BASE_STOCKS = {
    '2330': '2330_台積電_V4_高勝率回測.xlsx',
    '2360': '2360_致茂_V4_高勝率回測.xlsx',
    '6205': '6205_詮欣_V4_高勝率回測.xlsx',
    '6274': '6274_台耀_V4_高勝率回測.xlsx',
    '6669': '6669_緯穎_V_Rebound_高勝率回測.xlsx',
    '3189': '3189_景碩_V_Rebound_高勝率回測.xlsx',
    '3455': '3455_由田_V_Rebound_高勝率回測.xlsx',
    '3535': '3535_晶彩科_V_Rebound_高勝率回測.xlsx',
    '4908': '4908_前鼎_V_Rebound_高勝率回測.xlsx',
    '6269': '6269_台郡_V_Rebound_高勝率回測.xlsx',
    '3443': '3443_創意_V_Rebound_高勝率回測.xlsx',
    '6261': '6261_久元_V_Dip_Buy_高勝率回測.xlsx'
}

def clean_value(val):
    if pd.isna(val):
        return 0
    if isinstance(val, (int, float)):
        return val
    # Clean string symbols
    s = str(val).replace('▲', '').replace('▼', '').replace('↑', '').replace('↓', '')
    s = s.replace('△', '').replace('▽', '').replace(',', '').replace('%', '').strip()
    try:
        if '.' in s:
            return float(s)
        return int(s)
    except ValueError:
        return s

def main():
    print("=" * 60)
    print("  [Start] 12 檔預設個股 Excel 歷史數據更新工具...")
    print("=" * 60)

    for code, filename in BASE_STOCKS.items():
        print(f"🔄 正在處理股號 {code} ({filename})...")
        
        # 1. 尋找最新的 CSV
        csv_pattern = os.path.join(AUTO_EXPORT_DIR, f"{code}_*.csv")
        csv_files = glob.glob(csv_pattern)
        if not csv_files:
            print(f"  ⚠️  找不到股號 {code} 的 CSV 行情檔案！跳過...")
            continue
            
        # 按檔名日期排序，挑選最新的
        csv_files.sort()
        latest_csv = csv_files[-1]
        print(f"  📌  讀取 CSV 行情: {os.path.basename(latest_csv)}")
        
        try:
            # 2. 讀取與清理數據
            df = pd.read_csv(latest_csv, encoding='cp950')
            
            # 取前 15 欄
            df_subset = df.iloc[:, :15].copy()
            
            # 清理所有值
            for col in df_subset.columns:
                df_subset[col] = df_subset[col].apply(clean_value)
                
            # 確保日期是整數，其他是數值
            df_subset.iloc[:, 0] = df_subset.iloc[:, 0].astype(int)
            for col_idx in range(1, 15):
                df_subset.iloc[:, col_idx] = pd.to_numeric(df_subset.iloc[:, col_idx], errors='coerce').fillna(0)
            
            # 3. 寫入目標 Excel
            excel_path = os.path.join(DATA_DIR, filename)
            if not os.path.exists(excel_path):
                print(f"  ⚠️  目標 Excel 檔案不存在: {excel_path}，將跳過...")
                continue
                
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            # 刪除第 2 行以下的所有舊資料
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
                
            # 寫入新資料 (從 row 2 開始)
            for r_idx, row_values in enumerate(df_subset.values, 2):
                for c_idx, val in enumerate(row_values, 1):
                    # 如果是日期，寫入整數型別；其他如果浮點數，寫入浮點數
                    if c_idx == 1:
                        ws.cell(row=r_idx, column=c_idx, value=int(val))
                    else:
                        ws.cell(row=r_idx, column=c_idx, value=float(val))
            
            # 存檔
            wb.save(excel_path)
            wb.close()
            print(f"  ✅  股號 {code} 的 Excel 檔案已成功更新！共寫入 {len(df_subset)} 筆數據。")
            
        except Exception as e:
            print(f"  ❌  更新股號 {code} 時發生錯誤: {e}")

    print("=" * 60)
    print("  [Done] 12 檔預設個股數據更新完成！")
    print("=" * 60)

if __name__ == "__main__":
    main()
