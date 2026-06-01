import pandas as pd
import numpy as np
import itertools
import os
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment

input_dir = "E:\\G-AI-1\\Stock analysis\\Stock original"
output_html = "E:\\G-AI-1\\Stock analysis\\技術分析測試_全個股_深度分析版.html"
output_dir = "E:\\G-AI-1\\Stock analysis"

strategies = {
    '2330': {'name': '台積電', 'type': 'V4'},
    '3443': {'name': '創意', 'type': 'V38'},
    '2360': {'name': '致茂', 'type': 'MA_MFI_Rebound'},
    '6669': {'name': '緯穎', 'type': 'Dual_MA_MACD'},
    '3455': {'name': '由田', 'type': 'EOM_Break_Lock'},
    '3535': {'name': '晶彩科', 'type': 'MFI_Extreme'},
    '4746': {'name': '台耀', 'type': 'Vol_Breakout'},
    '4908': {'name': '前鼎', 'type': 'MACD_Pullback'},
    '3189': {'name': '景碩', 'type': 'MFI_MACD_Div'},
    '6205': {'name': '詮欣', 'type': 'MA60_MFI'},
    '6261': {'name': '久元', 'type': 'Triple_Confirm'},
    '6269': {'name': '台郡', 'type': 'Range_Fade'}
}

stock_data = {}
for file in os.listdir(input_dir):
    if file.endswith(".xlsx"):
        code = file.split('_')[0]
        if code in strategies:
            df = pd.read_excel(os.path.join(input_dir, file))
            df.columns = [str(c) for c in df.columns]
            for col in df.columns[1:]:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            stock_data[code] = {'df': df, 'file': file}

def generate_comment(row, rank):
    if rank == 1: return "🏆 最高報酬：完美捕捉大趨勢波段。"
    elif rank <= 5:
        if row.get('勝率', 0) > 0.6: return "⭐ 高勝率：穩定度佳，適合保守操作。"
        elif row.get('單筆最大虧損', -99999) > -30000: return "🛡️ 低風險：最大回檔小，資金控管優。"
        elif row.get('真實投報率(ROI)', 0) > 0.15: return "🚀 資金高效率：以小博大最佳配置。"
        else: return "💡 綜合優異：獲利與風險的極佳平衡點。"
    return ""

def evaluate_strategy(df, params, strat_type):
    close_arr, open_arr, high_arr = df.iloc[:, 4].values, df.iloc[:, 1].values, df.iloc[:, 2].values
    ma20_arr, ma60_arr = df.iloc[:, 7].values, df.iloc[:, 8].values
    eom_arr, eomsig_arr = df.iloc[:, 10].values, df.iloc[:, 11].values
    mfi_arr, macd_arr = df.iloc[:, 12].values, df.iloc[:, 14].values
    
    hold = False; buy_price = 0.0; pnls = []; entry_prices = []; mPrf = 0
    
    for i in range(2, len(df)-1):
        c_close, n_open = close_arr[i], open_arr[i+1]
        c_ma20, c_ma60 = ma20_arr[i], ma60_arr[i]
        c_eom, c_eomsig = eom_arr[i], eomsig_arr[i]
        c_mfi, c_macd = mfi_arr[i], macd_arr[i]
        p_eom, p_eomsig = eom_arr[i-1], eomsig_arr[i-1]
        p_high, p_macd = high_arr[i-1], macd_arr[i-1]
        
        if not hold:
            entry = False
            if strat_type == 'V4': entry = c_close > c_ma20 and c_eom > (c_eomsig + params['eom_off']) and c_macd > params['macd'] and c_mfi > params['mfi']
            elif strat_type == 'V38': entry = (c_eom > (c_eomsig + params['eom_sens']) and p_eom <= (p_eomsig + params['eom_sens'])) and c_close > c_ma20 and c_close > p_high and c_macd > p_macd
            elif strat_type == 'MA_MFI_Rebound': entry = c_close > c_ma60 and c_mfi > params['mfi'] and c_eom > c_eomsig
            elif strat_type == 'Dual_MA_MACD': entry = c_close > c_ma20 and c_ma20 > c_ma60 and c_macd > params['macd']
            elif strat_type == 'EOM_Break_Lock': entry = c_eom > c_eomsig and p_eom <= p_eomsig and c_macd > 0
            elif strat_type == 'MFI_Extreme': entry = c_mfi < params['mfi_s'] and c_close > c_ma20
            elif strat_type == 'Vol_Breakout': entry = c_close > p_high and c_eom > (c_eomsig + params['eom_sens']) and c_mfi > 50
            elif strat_type == 'MACD_Pullback': entry = c_macd > 0 and c_macd < p_macd and c_close > c_ma20
            elif strat_type == 'MFI_MACD_Div': entry = c_mfi < params['mfi'] and c_macd > p_macd
            elif strat_type == 'MA60_MFI': entry = c_close > c_ma60 and c_mfi < params['mfi']
            elif strat_type == 'Triple_Confirm': entry = c_close > c_ma20 and c_eom > params['eom_th'] and c_macd > 0
            elif strat_type == 'Range_Fade': entry = c_close < c_ma20 and c_mfi < params['mfi']

            if entry:
                hold = True; buy_price = n_open; mPrf = 0; entry_prices.append(buy_price)
        else:
            pnlR = (n_open - buy_price) / buy_price
            curR = (c_close - buy_price) / buy_price
            mPrf = max(mPrf, curR)
            exit_flag = False
            
            if strat_type in ['V38', 'EOM_Break_Lock']:
                if mPrf >= params['tri'] and pnlR <= params['lock']: exit_flag = True
                elif c_eom < c_eomsig: exit_flag = True
                elif pnlR <= -params['sl']: exit_flag = True
            elif strat_type == 'Range_Fade':
                if c_close > c_ma20 or pnlR >= params['tp'] or pnlR <= -params['sl']: exit_flag = True
            else:
                if pnlR >= params['tp'] or pnlR <= -params['sl']: exit_flag = True
                
            if exit_flag:
                hold = False
                pnls.append((n_open - buy_price) * 1000)
    
    if len(pnls) < 2: return -999999, {}
    cp = sum(pnls)
    mc = max(entry_prices) * 1000 if entry_prices else 0
    res = {
        '累積淨損益總額': cp, '最大所需本金': mc, '真實投報率(ROI)': cp / mc if mc > 0 else 0,
        '交易次數': len(pnls), '勝率': len([p for p in pnls if p > 0]) / len(pnls),
        '單筆最大獲利': max(pnls), '單筆最大虧損': min(pnls),
        '平均每筆損益': cp / len(pnls),
    }
    return cp, res

print("開始深度精細分析與建構 Excel...")
top5_db = {}

for code, info in strategies.items():
    if code not in stock_data: continue
    df = stock_data[code]['df']
    strat = info['type']
    print(f"[{code}] 處理策略: {strat}")
    
    tp_grid = np.arange(0.05, 0.20, 0.01); sl_grid = np.arange(0.03, 0.09, 0.01)
    results = []
    
    if strat == 'V4': grids = [dict(tp=t,sl=s,macd=md,mfi=mf,eom_off=0) for t,s,md,mf in itertools.product(tp_grid, sl_grid, [-0.5, 0.0, 0.5], [35, 40, 45])]
    elif strat in ['V38', 'EOM_Break_Lock']: grids = [dict(tri=t,lock=l,sl=s,eom_sens=0.5) for t,l,s in itertools.product(np.arange(0.02, 0.08, 0.005), np.arange(0.01, 0.05, 0.005), sl_grid) if t > l]
    elif strat in ['MA_MFI_Rebound', 'MFI_MACD_Div', 'MA60_MFI', 'Range_Fade']: grids = [dict(tp=t,sl=s,mfi=mf) for t,s,mf in itertools.product(tp_grid, sl_grid, [30, 35, 40, 45])]
    elif strat == 'MFI_Extreme': grids = [dict(tp=t,sl=s,mfi_s=mf) for t,s,mf in itertools.product(tp_grid, sl_grid, [20, 25, 30])]
    elif strat in ['Dual_MA_MACD', 'MACD_Pullback']: grids = [dict(tp=t,sl=s,macd=md) for t,s,md in itertools.product(tp_grid, sl_grid, [-0.5, 0.0, 0.5])]
    elif strat == 'Vol_Breakout': grids = [dict(tp=t,sl=s,eom_sens=e) for t,s,e in itertools.product(tp_grid, sl_grid, [0.0, 0.5, 1.0])]
    elif strat == 'Triple_Confirm': grids = [dict(tp=t,sl=s,eom_th=e) for t,s,e in itertools.product(tp_grid, sl_grid, [0.0, 0.5])]

    for p in grids:
        pnl, stat = evaluate_strategy(df, p, strat)
        if pnl > -999999:
            res = p.copy()
            res.update(stat)
            results.append(res)
            
    res_df = pd.DataFrame(results).sort_values(by='累積淨損益總額', ascending=False).reset_index(drop=True)
    res_df['最佳參數註解'] = [generate_comment(row, i+1) for i, row in res_df.iterrows()]
    top5 = res_df.head(5)
    top5_db[code] = top5
    
    # ======== 匯出 Excel ========
    wb = Workbook()
    ws_opt = wb.active
    ws_opt.title = "參數最佳化結果"
    header_names = list(top5.columns)
    ws_opt.append(header_names)
    
    for r_idx, row in res_df.iterrows():
        ws_opt.append([row[col] for col in header_names])
        cr = ws_opt.max_row
        if r_idx < 5:
            for col_num in range(1, len(header_names) + 1):
                ws_opt.cell(row=cr, column=col_num).fill = PatternFill("solid", fgColor="FEF08A")
                
    headers = list(df.columns) + ["進場訊號(P)", "出場訊號(Q)", "持倉狀態(R)", "進場成本(S)", "動作(T)", "交易價格(U)", "單筆損益(V)", "累積損益(W)"]
    
    for top_idx in range(len(top5)):
        ws_data = wb.create_sheet(f"Top{top_idx+1}_回測")
        best_p = top5.iloc[top_idx]
        
        ws_data['A1'] = f"【{strat} - Top{top_idx+1} 參數】"; ws_data['B1'] = "數值"
        p_keys = [k for k in best_p.keys() if k not in ['累積淨損益總額', '最大所需本金', '真實投報率(ROI)', '交易次數', '勝率', '單筆最大獲利', '單筆最大虧損', '平均每筆損益', '最佳參數註解']]
        
        for idx, p_name in enumerate(p_keys, 2):
            ws_data[f'A{idx}'] = p_name; ws_data[f'B{idx}'] = best_p[p_name]
            
        shares_row = len(p_keys) + 2
        ws_data[f'A{shares_row}'] = "交易張數 Shares"; ws_data[f'B{shares_row}'] = 1000
        
        for i, title in enumerate(headers, 1): ws_data.cell(row=8, column=i, value=title)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 9):
            for c_idx, value in enumerate(row, 1): ws_data.cell(row=r_idx, column=c_idx, value=value)
            
        # 寫入公式
        for r in range(9, ws_data.max_row + 1):
            pR, pP, pS, pW = ("FALSE", "FALSE", "0", "0") if r == 9 else (f"R{r-1}", f"P{r-1}", f"S{r-1}", f"W{r-1}")
            # B:Open, C:High, E:Close, H:MA20, I:MA60, K:EOM, L:EOMSig, M:MFI, O:MACD
            if strat == 'V4': # tp, sl, macd, mfi, eom_off
                ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}>H{r}, K{r}>(L{r}+$B$6), O{r}>$B$4, M{r}>$B$5, {pR}=FALSE))'
                ws_data[f'Q{r}'] = f'=IF({pR}=TRUE, OR((B{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$3), FALSE)'
            elif strat == 'V38': # tri, lock, sl, eom_sens
                ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(K{r}>(L{r}+$B$5), K{r-1}<=(L{r-1}+$B$5), E{r}>H{r}, E{r}>C{r-1}, O{r}>O{r-1}, {pR}=FALSE))'
                # 簡化 V38 Excel 出場公式
                ws_data[f'Q{r}'] = f'=IF({pR}=TRUE, OR((E{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$4, K{r}<L{r}), FALSE)'
            elif strat == 'EOM_Break_Lock': # tri, lock, sl, eom_sens
                ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(K{r}>L{r}, K{r-1}<=L{r-1}, O{r}>0, {pR}=FALSE))'
                ws_data[f'Q{r}'] = f'=IF({pR}=TRUE, OR((E{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$4, K{r}<L{r}), FALSE)'
            elif strat == 'MA_MFI_Rebound': # tp, sl, mfi
                ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}>I{r}, M{r}>$B$4, K{r}>L{r}, {pR}=FALSE))'
                ws_data[f'Q{r}'] = f'=IF({pR}=TRUE, OR((B{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$3), FALSE)'
            elif strat == 'Dual_MA_MACD': # tp, sl, macd
                ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}>H{r}, H{r}>I{r}, O{r}>$B$4, {pR}=FALSE))'
                ws_data[f'Q{r}'] = f'=IF({pR}=TRUE, OR((B{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$3), FALSE)'
            elif strat == 'MFI_Extreme': # tp, sl, mfi_s
                ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(M{r}<$B$4, E{r}>H{r}, {pR}=FALSE))'
                ws_data[f'Q{r}'] = f'=IF({pR}=TRUE, OR((B{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$3), FALSE)'
            elif strat == 'Vol_Breakout': # tp, sl, eom_sens
                ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}>C{r-1}, K{r}>(L{r}+$B$4), M{r}>50, {pR}=FALSE))'
                ws_data[f'Q{r}'] = f'=IF({pR}=TRUE, OR((B{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$3), FALSE)'
            elif strat == 'MACD_Pullback': # tp, sl, macd
                ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(O{r}>0, O{r}<O{r-1}, E{r}>H{r}, {pR}=FALSE))'
                ws_data[f'Q{r}'] = f'=IF({pR}=TRUE, OR((B{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$3), FALSE)'
            elif strat == 'MFI_MACD_Div': # tp, sl, mfi
                ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(M{r}<$B$4, O{r}>O{r-1}, {pR}=FALSE))'
                ws_data[f'Q{r}'] = f'=IF({pR}=TRUE, OR((B{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$3), FALSE)'
            elif strat == 'MA60_MFI': # tp, sl, mfi
                ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}>I{r}, M{r}<$B$4, {pR}=FALSE))'
                ws_data[f'Q{r}'] = f'=IF({pR}=TRUE, OR((B{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$3), FALSE)'
            elif strat == 'Triple_Confirm': # tp, sl, eom_th
                ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}>H{r}, K{r}>$B$4, O{r}>0, {pR}=FALSE))'
                ws_data[f'Q{r}'] = f'=IF({pR}=TRUE, OR((B{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$3), FALSE)'
            elif strat == 'Range_Fade': # tp, sl, mfi
                ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}<H{r}, M{r}<$B$4, {pR}=FALSE))'
                ws_data[f'Q{r}'] = f'=IF({pR}=TRUE, OR(E{r}>H{r}, (B{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$3), FALSE)'

            ws_data[f'R{r}'] = f'=IF({pR}=FALSE, {pP}=TRUE, IF(Q{r}=TRUE, FALSE, TRUE))'
            ws_data[f'S{r}'] = f'=IF(AND({pR}=FALSE, {pP}=TRUE), B{r}, IF(AND({pR}=TRUE, Q{r}=FALSE), {pS}, 0))'
            ws_data[f'T{r}'] = f'=IF(AND({pR}=FALSE, {pP}=TRUE), "買進", IF(Q{r}=TRUE, "賣出", ""))'
            ws_data[f'U{r}'] = f'=IF(T{r}="買進", B{r}, IF(T{r}="賣出", B{r}, 0))'
            ws_data[f'V{r}'] = f'=IF(T{r}="賣出", (U{r} - {pS}) * $B${shares_row}, 0)'
            ws_data[f'W{r}'] = f'={pW} + V{r}'

    out_file = os.path.join(output_dir, f"{code}_{info['name']}_{strat}_深度最佳化與回測.xlsx")
    wb.save(out_file)

# ====== 生成 HTML ======
html_head = """<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <title>技術分析測試-全個股深度分析版</title>
    <script src="https://cdn.jsdelivr.net/npm/xlsx/dist/xlsx.full.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        :root { --primary: #1e40af; --bg: #f1f5f9; --win: #dc2626; --loss: #16a34a; }
        body { font-family: 'PingFang TC', 'Microsoft JhengHei', sans-serif; background: var(--bg); margin: 0; padding: 0; color: #1e293b; }
        .nav-tabs { background: #0f172a; padding: 15px 0; display: flex; flex-wrap: wrap; justify-content: center; gap: 12px; position: sticky; top: 0; z-index: 100; }
        .tab-btn { padding: 10px 25px; border: none; background: #334155; color: white; border-radius: 8px; cursor: pointer; font-weight: bold; transition: 0.3s; }
        .tab-btn.active { background: #2563eb; box-shadow: 0 0 15px rgba(37, 99, 235, 0.4); }
        .container { max-width: 1200px; margin: 20px auto; background: white; padding: 25px; border-radius: 15px; box-shadow: 0 4px 20px rgba(0,0,0,0.08); }
        .page { display: none; }
        .page.active { display: block; }
        .control-panel { display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 12px; background: #f8fafc; padding: 20px; border-radius: 10px; margin-bottom: 20px; border: 1px solid #e2e8f0; }
        .param-group { display: flex; flex-direction: column; gap: 6px; }
        .param-group label { font-size: 12px; font-weight: bold; color: #475569; }
        .param-group input, .param-group select { padding: 8px; border: 1px solid #cbd5e1; border-radius: 6px; font-size: 13px; background: white; }
        .section-title { grid-column: 1 / -1; font-size: 14px; color: #1e40af; border-bottom: 1px solid #e2e8f0; padding-bottom: 5px; margin-top: 10px; font-weight: bold; }
        
        .info-panel { background: #fdfae8; border: 1px solid #fde047; padding: 15px; border-radius: 10px; margin-bottom: 20px; font-size: 13px; }
        .info-panel h4 { margin: 0 0 10px 0; color: #b45309; }
        .info-table { width: 100%; border-collapse: collapse; background: white; }
        .info-table th, .info-table td { padding: 8px; border: 1px solid #fef08a; text-align: center; }
        .info-table th { background: #fef9c3; color: #854d0e; }
        
        .drop-zone { border: 3px dashed #2563eb; padding: 25px; text-align: center; border-radius: 10px; cursor: pointer; background: #eff6ff; margin-bottom: 5px; transition: 0.3s; }
        .drop-zone:hover { background: #dbeaffe6; }
        .file-info { text-align: center; font-size: 13px; color: #1e40af; margin-bottom: 15px; font-weight: bold; }
        .dashboard { display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin-bottom: 20px; }
        .stat-card { border: 1px solid #e2e8f0; padding: 12px; border-radius: 8px; text-align: center; background: #fff; }
        .stat-card b { font-size: 11px; color: #64748b; display: block; margin-bottom: 4px; }
        .stat-card span { font-size: 1.25em; font-weight: bold; }
        .win { color: var(--win); } .loss { color: var(--loss); } 
        .chart-wrap { height: 450px; margin-bottom: 25px; border: 1px solid #e2e8f0; padding: 10px; border-radius: 10px; background: white; }
        table { width: 100%; border-collapse: collapse; font-size: 13px; }
        th, td { padding: 10px; border-bottom: 1px solid #f1f5f9; text-align: center; }
        th { background: #f8fafc; position: sticky; top: 0; }
        .buy-row { background: #fff1f0; } .sell-row { background: #f0fdf4; }
    </style>
</head>
<body>
"""

nav_html = '<nav class="nav-tabs">\n'
pages_html = '<div class="container">\n'
js_vars = "let charts = {};\nlet dataStore = {};\n"
js_logic = ""

for i, (code, info) in enumerate(strategies.items()):
    name = info['name']
    strat = info['type']
    
    top5 = top5_db.get(code)
    p = top5.iloc[0] if top5 is not None and len(top5) > 0 else {}
    
    active_cls = ' active' if i == 0 else ''
    nav_html += f"""<button class="tab-btn{active_cls}" onclick="switchTab('tab_{code}')">{code}_{name}</button>\n"""
    
    # 建立 Info Panel (Top 5 Table)
    info_table_html = ""
    if top5 is not None and len(top5) > 0:
        info_table_html += f"""
        <div class="info-panel">
            <h4>💡 深度分析 Top 5 最佳化參數組合</h4>
            <table class="info-table">
                <tr><th>名次</th><th>參數組合</th><th>累積損益</th><th>勝率</th><th>AI 策略註解</th></tr>"""
        for rank, row in top5.iterrows():
            params_str = ", ".join([f"{k}:{v*100:.1f}%" if k in ['tp','sl','tri','lock'] else f"{k}:{v:.1f}" for k,v in row.items() if k not in ['累積淨損益總額','最大所需本金','真實投報率(ROI)','交易次數','勝率','單筆最大獲利','單筆最大虧損','平均每筆損益','最佳參數註解']])
            info_table_html += f"""<tr><td>Top {rank+1}</td><td>{params_str}</td><td>${row['累積淨損益總額']:,.0f}</td><td>{row['勝率']*100:.1f}%</td><td>{row['最佳參數註解']}</td></tr>"""
        info_table_html += "</table></div>"

    ui = ""; logic = ""; extract = ""
    
    if strat == 'V4':
        ui = f"""
            <div class="section-title">資金與基礎出口</div>
            <div class="param-group"><label>交易張數</label><input type="number" id="s{code}_sh" value="1"></div>
            <div class="param-group"><label>固定停利 (%)</label><input type="number" id="s{code}_tp" value="{p.get('tp',0.1)*100:.1f}" step="0.1"></div>
            <div class="param-group"><label>固定停損 (%)</label><input type="number" id="s{code}_sl" value="{p.get('sl',0.05)*100:.1f}" step="0.1"></div>
            <div class="section-title">進場過濾參數</div>
            <div class="param-group"><label>均線趨勢確認</label><select id="s{code}_ma"><option value="ma20" selected>MA 20 (月線)</option><option value="ma60">MA 60 (季線)</option></select></div>
            <div class="param-group"><label>MACD 需大於</label><input type="number" id="s{code}_macd" value="{p.get('macd',0):.1f}" step="0.1"></div>
            <div class="param-group"><label>EOM 金叉偏移</label><input type="number" id="s{code}_eom" value="{p.get('eom_off',0):.1f}" step="0.1"></div>
            <div class="param-group"><label>MFI 門檻</label><input type="number" id="s{code}_mfi" value="{p.get('mfi',40):.1f}" step="1"></div>
        """
        extract = f"let tp=parseFloat(document.getElementById('s{code}_tp').value)/100, sl=parseFloat(document.getElementById('s{code}_sl').value)/100, macd=parseFloat(document.getElementById('s{code}_macd').value), eom=parseFloat(document.getElementById('s{code}_eom').value), mfi=parseFloat(document.getElementById('s{code}_mfi').value), ma=document.getElementById('s{code}_ma').value;"
        logic = """
            let maV = ma==='ma60'?c.ma60:c.ma20;
            if(!hold){
                if(c.close > maV && c.eom > (c.eomSig + eom) && c.macd > macd && c.mfi > mfi){
                    hold=true; bp=n.open; bi[i+1]=1; mCap=Math.max(mCap, bp*sh); trd.push({type:'買入', date:n.date, price:bp, info:'V4進場'});
                }
            } else {
                let pnlR = (n.open-bp)/bp;
                if(pnlR >= tp || pnlR <= -sl){
                    hold=false; let ep=n.open; let diff=(ep-bp)*sh; tot+=diff; si[i+1]=1; trd.push({type:'賣出', date:n.date, price:ep, info:pnlR>0?'停利':'停損', pnl:diff, entP:bp});
                }
            }
        """
    elif strat == 'V38':
        ui = f"""
            <div class="section-title">資金與鎖利出口</div>
            <div class="param-group"><label>交易張數</label><input type="number" id="s{code}_sh" value="1"></div>
            <div class="param-group"><label>觸發鎖利 (%)</label><input type="number" id="s{code}_tri" value="{p.get('tri',0.03)*100:.1f}" step="0.1"></div>
            <div class="param-group"><label>保證獲利 (%)</label><input type="number" id="s{code}_lk" value="{p.get('lock',0.02)*100:.1f}" step="0.1"></div>
            <div class="param-group"><label>硬性停損 (%)</label><input type="number" id="s{code}_st" value="{p.get('sl',0.03)*100:.1f}" step="0.1"></div>
            <div class="section-title">進場過濾參數</div>
            <div class="param-group"><label>進場均線參考</label><select id="s{code}_ma"><option value="ma20" selected>MA 20 (月線)</option><option value="ma60">MA 60 (季線)</option></select></div>
            <div class="param-group"><label>EOM 金叉靈敏度</label><input type="number" id="s{code}_eom" value="{p.get('eom_sens',0.5):.1f}" step="0.1"></div>
            <div class="param-group"><label>必須突破前高</label><select id="s{code}_brk"><option value="yes" selected>是</option><option value="no">否</option></select></div>
            <div class="param-group"><label>MACD 濾網</label><select id="s{code}_macd"><option value="grow" selected>遞增</option><option value="zero">大於0</option></select></div>
        """
        extract = f"let tri=parseFloat(document.getElementById('s{code}_tri').value)/100, lk=parseFloat(document.getElementById('s{code}_lk').value)/100, st=parseFloat(document.getElementById('s{code}_st').value)/100, eom=parseFloat(document.getElementById('s{code}_eom').value), ma=document.getElementById('s{code}_ma').value, brk=document.getElementById('s{code}_brk').value, macdMode=document.getElementById('s{code}_macd').value;"
        logic = """
            let maV = ma==='ma60'?c.ma60:c.ma20;
            if(!hold){
                let eomC = c.eom > (c.eomSig + eom) && p.eom <= (p.eomSig + eom);
                let maC = c.close > maV;
                let brkC = brk==='yes'? c.close > p.high : true;
                let macdCond = macdMode==='grow'? c.macd > p.macd : c.macd > 0;
                if(eomC && maC && brkC && macdCond){
                    hold=true; bp=n.open; bi[i+1]=1; mPrf=0; mCap=Math.max(mCap, bp*sh); trd.push({type:'買入', date:n.date, price:bp, info:'V38進場'});
                }
            } else {
                let curR=(c.close-bp)/bp; mPrf=Math.max(mPrf, curR);
                let nOpR=(n.open-bp)/bp, tP = bp*(1+lk);
                let exit=false, rsn="";
                if(mPrf>=tri && nOpR<=lk) {exit=true; rsn="獲利鎖定";}
                else if(c.eom < c.eomSig) {exit=true; rsn="動能轉弱";}
                else if(nOpR < -st) {exit=true; rsn="硬停損";}
                if(exit){
                    hold=false; let ep=(rsn==="獲利鎖定" && tP<=n.high && tP>=n.low)? tP: n.open; let diff=(ep-bp)*sh; tot+=diff; si[i+1]=1; trd.push({type:'賣出', date:n.date, price:ep, info:rsn, pnl:diff, entP:bp});
                }
            }
        """
    else:
        tp_v = p.get('tp', 0.1)*100; sl_v = p.get('sl', 0.05)*100
        macd_v = p.get('macd', 0); mfi_v = p.get('mfi', 40); mfi_s_v = p.get('mfi_s', 30)
        eom_v = p.get('eom_sens', 0.5) if 'eom_sens' in p else p.get('eom_th', 0.5)
        
        ui = f"""
            <div class="section-title">資金與基礎出口</div>
            <div class="param-group"><label>交易張數</label><input type="number" id="s{code}_sh" value="1"></div>
            <div class="param-group"><label>固定停利 (%)</label><input type="number" id="s{code}_tp" value="{tp_v:.1f}" step="0.1"></div>
            <div class="param-group"><label>固定停損 (%)</label><input type="number" id="s{code}_sl" value="{sl_v:.1f}" step="0.1"></div>
            <div class="section-title">進場過濾參數</div>
            <div class="param-group"><label>MACD 參數</label><input type="number" id="s{code}_macd" value="{macd_v:.1f}" step="0.1"></div>
            <div class="param-group"><label>MFI 參數</label><input type="number" id="s{code}_mfi" value="{mfi_v:.1f}" step="1"></div>
            <div class="param-group"><label>EOM 參數</label><input type="number" id="s{code}_eom" value="{eom_v:.1f}" step="0.1"></div>
        """
        extract = f"let tp=parseFloat(document.getElementById('s{code}_tp').value)/100, sl=parseFloat(document.getElementById('s{code}_sl').value)/100, macd=parseFloat(document.getElementById('s{code}_macd').value), mfi=parseFloat(document.getElementById('s{code}_mfi').value), eom=parseFloat(document.getElementById('s{code}_eom').value);"
        
        entry_cond = ""
        exit_cond = "let pnlR = (n.open-bp)/bp; if(pnlR >= tp || pnlR <= -sl){ hold=false; let ep=n.open; let diff=(ep-bp)*sh; tot+=diff; si[i+1]=1; trd.push({type:'賣出', date:n.date, price:ep, info:pnlR>0?'停利':'停損', pnl:diff, entP:bp}); }"
        
        if strat == 'MA_MFI_Rebound': entry_cond = "c.close > c.ma60 && c.mfi > mfi && c.eom > c.eomSig"
        elif strat == 'Dual_MA_MACD': entry_cond = "c.close > c.ma20 && c.ma20 > c.ma60 && c.macd > macd"
        elif strat == 'EOM_Break_Lock':
            extract = f"let tri=parseFloat(document.getElementById('s{code}_tp').value)/100, lk=parseFloat(document.getElementById('s{code}_sl').value)/100, st=parseFloat(document.getElementById('s{code}_macd').value)/100;"
            entry_cond = "c.eom > c.eomSig && p.eom <= p.eomSig && c.macd > 0"
            exit_cond = "let curR=(c.close-bp)/bp; mPrf=Math.max(mPrf, curR); let nOpR=(n.open-bp)/bp, tP = bp*(1+lk); let exit=false, rsn=''; if(mPrf>=tri && nOpR<=lk) {exit=true; rsn='獲利鎖定';} else if(c.eom < c.eomSig) {exit=true; rsn='動能衰退';} else if(nOpR < -st) {exit=true; rsn='硬停損';} if(exit){ hold=false; let ep=(rsn==='獲利鎖定' && tP<=n.high && tP>=n.low)? tP: n.open; let diff=(ep-bp)*sh; tot+=diff; si[i+1]=1; trd.push({type:'賣出', date:n.date, price:ep, info:rsn, pnl:diff, entP:bp}); }"
            ui = f"""<div class="section-title">資金與基礎出口</div><div class="param-group"><label>交易張數</label><input type="number" id="s{code}_sh" value="1"></div><div class="param-group"><label>觸發鎖利(%)</label><input type="number" id="s{code}_tp" value="{p.get('tri',0.06)*100:.1f}" step="0.1"></div><div class="param-group"><label>保證鎖定(%)</label><input type="number" id="s{code}_sl" value="{p.get('lock',0.04)*100:.1f}" step="0.1"></div><div class="param-group"><label>停損(%)</label><input type="number" id="s{code}_macd" value="{p.get('sl',0.05)*100:.1f}" step="0.1"></div>"""
        elif strat == 'MFI_Extreme': entry_cond = "c.mfi < mfi && c.close > c.ma20"
        elif strat == 'Vol_Breakout': entry_cond = "c.close > p.high && c.eom > (c.eomSig + eom) && c.mfi > 50"
        elif strat == 'MACD_Pullback': entry_cond = "c.macd > 0 && c.macd < p.macd && c.close > c.ma20"
        elif strat == 'MFI_MACD_Div': entry_cond = "c.mfi < mfi && c.macd > p.macd"
        elif strat == 'MA60_MFI': entry_cond = "c.close > c.ma60 && c.mfi < mfi"
        elif strat == 'Triple_Confirm': entry_cond = "c.close > c.ma20 && c.eom > eom && c.macd > 0"
        elif strat == 'Range_Fade': 
            entry_cond = "c.close < c.ma20 && c.mfi < mfi"
            exit_cond = "let pnlR = (n.open-bp)/bp; if(c.close > c.ma20 || pnlR >= tp || pnlR <= -sl){ hold=false; let ep=n.open; let diff=(ep-bp)*sh; tot+=diff; si[i+1]=1; trd.push({type:'賣出', date:n.date, price:ep, info:'反轉出場', pnl:diff, entP:bp}); }"

        logic = f"""
            if(!hold){{
                if({entry_cond}){{
                    hold=true; bp=n.open; bi[i+1]=1; mCap=Math.max(mCap, bp*sh); mPrf=0; trd.push({{type:'買入', date:n.date, price:bp, info:'策略進場'}});
                }}
            }} else {{
                {exit_cond}
            }}
        """

    pages_html += f"""
    <div id="tab_{code}" class="page{active_cls}">
        <h3 style="text-align:center; color:#1e40af;">{name} ({code}) 深度策略分析 - {strat}</h3>
        <div class="control-panel">{ui}</div>
        {info_table_html}
        <div class="drop-zone" id="s{code}_drop"><strong>將 {code} 檔案拖入此處</strong><br>策略：{strat}<input type="file" id="s{code}_file" hidden></div>
        <div id="s{code}_filename" class="file-info">未載入檔案</div>
        <div class="dashboard" id="s{code}_dash"></div>
        <div class="chart-wrap"><canvas id="s{code}_chart"></canvas></div>
        <table id="s{code}_table"><thead><tr><th>類型</th><th>日期</th><th>價格</th><th>原因</th><th>損益(元)</th></tr></thead><tbody></tbody></table>
    </div>
    """
    
    js_vars += f"charts['s{code}'] = null; dataStore['s{code}'] = null;\n"
    js_logic += f"""
document.getElementById('s{code}_drop').onclick = () => document.getElementById('s{code}_file').click();
document.getElementById('s{code}_file').onchange = (e) => {{
    if (e.target.files.length > 0) {{
        document.getElementById('s{code}_filename').innerText = "📄 已載入：" + e.target.files[0].name;
        const reader = new FileReader();
        reader.onload = (ev) => {{
            const wb = XLSX.read(ev.target.result, {{type: 'array'}});
            dataStore['s{code}'] = parseData(XLSX.utils.sheet_to_json(wb.Sheets[wb.SheetNames[0]], {{header: 1}}));
            run_s{code}();
        }};
        reader.readAsArrayBuffer(e.target.files[0]);
    }}
}};
document.querySelectorAll('#tab_{code} input, #tab_{code} select').forEach(i => i.oninput = () => dataStore['s{code}'] && run_s{code}());

function run_s{code}() {{
    const ds = dataStore['s{code}'];
    if(!ds) return;
    const sh = parseInt(document.getElementById('s{code}_sh').value) * 1000;
    {extract}
    let trd = [], hold = false, bp = 0, tot = 0, mCap = 0, mPrf = 0;
    let bi = new Array(ds.length).fill(null), si = new Array(ds.length).fill(null);

    for(let i=1; i<ds.length-1; i++){{
        let c = ds[i], n = ds[i+1], p = ds[i-1];
        {logic}
    }}
    updateUI('s{code}', trd, tot, mCap, ds, bi, si);
}}
"""

nav_html += '</nav>\n'
pages_html += '</div>\n'

html_footer = """
<script>
const verticalLinePlugin = {
    id: 'verticalLinePlugin',
    beforeDraw(chart) {
        const { ctx, chartArea: { top, bottom }, scales: { x } } = chart;
        chart.data.datasets.forEach(dataset => {
            if (dataset.lineType) {
                ctx.save(); ctx.lineWidth = 2;
                ctx.strokeStyle = dataset.lineType === 'buy' ? 'rgba(220, 38, 38, 0.6)' : 'rgba(22, 163, 74, 0.6)';
                dataset.data.forEach((val, idx) => {
                    if (val !== null) {
                        const xPos = x.getPixelForValue(chart.data.labels[idx]);
                        ctx.beginPath(); ctx.moveTo(xPos, top); ctx.lineTo(xPos, bottom); ctx.stroke();
                    }
                });
                ctx.restore();
            }
        });
    }
};
Chart.register(verticalLinePlugin);

function switchTab(target) {
    document.querySelectorAll('.page').forEach(p => p.classList.remove('active'));
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    document.getElementById(target).classList.add('active');
    event.currentTarget.classList.add('active');
}

const parseData = (rows) => {
    const p = (v) => v ? parseFloat(String(v).replace(/[↑↓,%\\s]/g, '')) || 0 : 0;
    return rows.filter(r => r[0] && (String(r[0]).includes('/') || /^\\d{8}$/.test(r[0]))).map(r => ({
        date: String(r[0]).trim(), open: p(r[1]), high: p(r[2]), low: p(r[3]), close: p(r[4]),
        ma20: p(r[8]), ma60: p(r[9]), eom: p(r[10]), eomSig: p(r[11]), mfi: p(r[12]), macd: p(r[14])
    }));
};

function updateUI(pre, trd, tot, mCap, ds, bi, si) {
    const sls = trd.filter(t => t.type === '賣出');
    const pnls = sls.map(t => t.pnl);
    const cnt = sls.length, win = pnls.filter(v => v > 0).length;
    const mxP = cnt > 0 ? Math.max(...pnls) : 0, mxL = cnt > 0 ? Math.min(...pnls) : 0;
    const avg = cnt > 0 ? tot / cnt : 0;
    const avgEnt = cnt > 0 ? (sls.reduce((sum, t) => sum + t.entP, 0) / cnt) * (parseInt(document.getElementById(pre+'_sh').value)*1000) : 1;
    const expR = (avg / avgEnt) * 100;

    document.getElementById(pre + '_dash').innerHTML = `
        <div class="stat-card"><b>累積淨損益總額</b><span class="${tot>=0?'win':'loss'}">$${Math.round(tot).toLocaleString()}</span></div>
        <div class="stat-card"><b>最大所需本金</b><span>$${Math.round(mCap).toLocaleString()}</span></div>
        <div class="stat-card"><b>真實投報率 (ROI)</b><span class="${tot>=0?'win':'loss'}">${mCap>0?(tot/mCap*100).toFixed(2):0}%</span></div>
        <div class="stat-card"><b>交易次數 / 勝率</b><span>${cnt}次 / ${cnt>0?(win/cnt*100).toFixed(1):0}%</span></div>
        <div class="stat-card"><b>單筆最大獲利</b><span class="win">$${Math.round(mxP).toLocaleString()}</span></div>
        <div class="stat-card"><b>單筆最大虧損</b><span class="loss">$${Math.round(mxL).toLocaleString()}</span></div>
        <div class="stat-card"><b>平均每筆損益</b><span class="${avg>=0?'win':'loss'}">$${Math.round(avg).toLocaleString()}</span></div>
        <div class="stat-card"><b>期望報酬率 (每筆)</b><span class="${expR>=0?'win':'loss'}">${expR.toFixed(2)}%</span></div>
    `;

    const tbody = document.querySelector(`#${pre}_table tbody`);
    tbody.innerHTML = trd.map(t => `
        <tr class="${t.type==='買入'?'buy-row':'sell-row'}">
            <td><b>${t.type}</b></td><td>${t.date}</td><td>${t.price.toFixed(1)}</td><td>${t.info}</td>
            <td class="${t.pnl>0?'win':(t.pnl<0?'loss':'')}">${t.pnl?Math.round(t.pnl).toLocaleString():'-'}</td>
        </tr>
    `).reverse().join('');

    const ctx = document.getElementById(pre + '_chart');
    if (charts[pre]) charts[pre].destroy();
    charts[pre] = new Chart(ctx, {
        type: 'line',
        data: {
            labels: ds.map(d => d.date),
            datasets: [
                { label: '收盤價', data: ds.map(d => d.close), borderColor: '#475569', borderWidth: 1, pointRadius: 0, yAxisID: 'y' },
                { data: bi, lineType: 'buy', pointRadius: 0, yAxisID: 'y_sub' },
                { data: si, lineType: 'sell', pointRadius: 0, yAxisID: 'y_sub' }
            ]
        },
        options: {
            maintainAspectRatio: false,
            scales: { y: { type: 'linear', position: 'left' }, y_sub: { display: false, min: 0, max: 1 } },
            plugins: { legend: { display: false } }
        }
    });
}
"""

with open(output_html, 'w', encoding='utf-8') as f:
    f.write(html_head + nav_html + pages_html + html_footer + js_vars + js_logic + "</script>\n</body>\n</html>")

print("全部工作完成！已匯出 Excel 並生成包含 Top 5 資訊的新版 HTML。")
