import os, shutil, pandas as pd, numpy as np, itertools
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

src_html = r"E:\G-AI-1\Stock analysis\修正版_最終\技術分析測試_全個股_高勝率修正版.html"
out_dir = r"E:\G-AI-1\Stock analysis\修正版_V2"
os.makedirs(out_dir, exist_ok=True)

# ========== PART 1: Patch HTML ==========
with open(src_html, 'r', encoding='utf-8') as f:
    html = f.read()

# 1a. parseData: add ma5, ma120
html = html.replace(
    'ma20: p(r[7]), ma60: p(r[8]), eom: p(r[10]),',
    'ma5: p(r[6]), ma20: p(r[7]), ma60: p(r[8]), ma120: p(r[9]), eom: p(r[10]),'
)

# 1b. Add MA dropdown to all control panels
def ma_dd(code):
    return (f'<div class="param-group"><label>進場均線參考</label>'
            f'<select id="s{code}_ma">'
            f'<option value="none">不參考</option>'
            f'<option value="ma5">MA 5</option>'
            f'<option value="ma20" selected>MA 20</option>'
            f'<option value="ma60">MA 60</option>'
            f'<option value="ma120">MA 120</option>'
            f'</select></div>\n            ')

for code in ['2330', '2360', '6205']:
    old = f'<div class="param-group"><label>MACD 需大於</label><input type="number" id="s{code}_macd"'
    html = html.replace(old, ma_dd(code) + old)

for code in ['6669', '3443', '3189', '3455', '3535', '4908', '6269', '4746', '6261']:
    old = f'<div class="param-group"><label>超賣區 (MFI <)</label><input type="number" id="s{code}_mfi_s"'
    html = html.replace(old, ma_dd(code) + old)

# 1c. Update run functions - add maType reading + use in conditions
# Helper to get MA value JS snippet
def ma_js(code):
    return (f"let maType=document.getElementById('s{code}_ma').value;"
            f"let _gm=function(c){{return maType==='ma5'?c.ma5:maType==='ma60'?c.ma60:maType==='ma120'?c.ma120:maType==='ma20'?c.ma20:null;}};")

# V4 stocks: add maType after params, change c.close > c.ma20 to use selector
for code in ['2330', '2360', '6205']:
    # Add maType reading
    old_params = f"mfi=parseFloat(document.getElementById('s{code}_mfi').value);"
    html = html.replace(old_params, old_params + '\n    ' + ma_js(code))
    # Change entry condition
    html = html.replace(
        f"        let c = ds[i], n = ds[i+1], p = ds[i-1];\r\n        \r\n            if(!hold){{\r\n                if(c.close > c.ma20 && c.eom > (c.eomSig + eom) && c.macd > macd && c.mfi > mfi){{\r\n                    hold=true; bp=n.open; bi[i+1]=1; mCap=Math.max(mCap, bp*sh); trd.push({{type:'買入', date:n.date, price:bp, info:'V4進場'}});".replace('{{','{').replace('}}','}'),
        f"        let c = ds[i], n = ds[i+1], p = ds[i-1];\r\n            let _mv=_gm(c);\r\n            if(!hold){{\r\n                if((_mv===null||c.close>_mv) && c.eom > (c.eomSig + eom) && c.macd > macd && c.mfi > mfi){{\r\n                    hold=true; bp=n.open; bi[i+1]=1; mCap=Math.max(mCap, bp*sh); trd.push({{type:'買入', date:n.date, price:bp, info:'V4進場'}});".replace('{{','{').replace('}}','}')
    )

# V_Rebound stocks: add maType after params, change c.close < c.ma20
for code in ['6669', '3443', '3189', '3455', '3535', '4908', '6269']:
    old_params = f"mfi_b=parseFloat(document.getElementById('s{code}_mfi_b').value);"
    html = html.replace(old_params, old_params + '\n    ' + ma_js(code))
    html = html.replace(
        f"        let c = ds[i], n = ds[i+1], p = ds[i-1];\r\n        \r\n            if(!hold){{\r\n                if(c.close < c.ma20 && c.mfi < mfi_s && c.macd > p.macd){{\r\n                    hold=true; bp=n.open; bi[i+1]=1; mCap=Math.max(mCap, bp*sh); trd.push({{type:'買入', date:n.date, price:bp, info:'逆勢反轉進場'}});".replace('{{','{').replace('}}','}'),
        f"        let c = ds[i], n = ds[i+1], p = ds[i-1];\r\n            let _mv=_gm(c);\r\n            if(!hold){{\r\n                if((_mv===null||c.close<_mv) && c.mfi < mfi_s && c.macd > p.macd){{\r\n                    hold=true; bp=n.open; bi[i+1]=1; mCap=Math.max(mCap, bp*sh); trd.push({{type:'買入', date:n.date, price:bp, info:'逆勢反轉進場'}});".replace('{{','{').replace('}}','}')
    )

# V_Dip_Buy stocks: add maType after params, add MA condition
for code in ['4746', '6261']:
    old_params = f"mfi_b=parseFloat(document.getElementById('s{code}_mfi_b').value);"
    html = html.replace(old_params, old_params + '\n    ' + ma_js(code))
    html = html.replace(
        f"        let c = ds[i], n = ds[i+1], p = ds[i-1];\r\n        \r\n            if(!hold){{\r\n                if(c.mfi < mfi_s && c.eom > p.eom && c.macd > 0){{\r\n                    hold=true; bp=n.open; bi[i+1]=1; mCap=Math.max(mCap, bp*sh); trd.push({{type:'買入', date:n.date, price:bp, info:'動能低接進場'}});".replace('{{','{').replace('}}','}'),
        f"        let c = ds[i], n = ds[i+1], p = ds[i-1];\r\n            let _mv=_gm(c);\r\n            if(!hold){{\r\n                if((_mv===null||c.close>_mv) && c.mfi < mfi_s && c.eom > p.eom && c.macd > 0){{\r\n                    hold=true; bp=n.open; bi[i+1]=1; mCap=Math.max(mCap, bp*sh); trd.push({{type:'買入', date:n.date, price:bp, info:'動能低接進場'}});".replace('{{','{').replace('}}','}')
    )

with open(os.path.join(out_dir, '技術分析測試_全個股_高勝率修正版.html'), 'w', encoding='utf-8') as f:
    f.write(html)
print("HTML 修正完成")

# ========== PART 2: Regenerate Excel with proper headers ==========
input_dir = r"E:\G-AI-1\Stock analysis\Stock original"
strategies = {
    '2330': {'name': '台積電', 'type': 'V4', 'params': {'tp':0.11,'sl':0.09,'macd':2.1,'mfi':30,'eom_off':0.0}},
    '2360': {'name': '致茂', 'type': 'V4', 'params': {'tp':0.189,'sl':0.08,'macd':2.1,'mfi':30,'eom_off':0.0}},
    '6205': {'name': '詮欣', 'type': 'V4', 'params': {'tp':0.08,'sl':0.05,'macd':2.1,'mfi':30,'eom_off':0.0}},
    '6669': {'name': '緯穎', 'type': 'V_Rebound', 'params': {'tp':0.19,'sl':0.05,'mfi_s':25,'mfi_b':75}},
    '3443': {'name': '創意', 'type': 'V_Rebound', 'params': {'tp':0.17,'sl':0.07,'mfi_s':35,'mfi_b':75}},
    '3189': {'name': '景碩', 'type': 'V_Rebound', 'params': {'tp':0.19,'sl':0.09,'mfi_s':25,'mfi_b':80}},
    '3455': {'name': '由田', 'type': 'V_Rebound', 'params': {'tp':0.18,'sl':0.04,'mfi_s':25,'mfi_b':80}},
    '3535': {'name': '晶彩科', 'type': 'V_Rebound', 'params': {'tp':0.13,'sl':0.07,'mfi_s':25,'mfi_b':80}},
    '4908': {'name': '前鼎', 'type': 'V_Rebound', 'params': {'tp':0.14,'sl':0.04,'mfi_s':25,'mfi_b':80}},
    '6269': {'name': '台郡', 'type': 'V_Rebound', 'params': {'tp':0.09,'sl':0.07,'mfi_s':25,'mfi_b':75}},
    '4746': {'name': '台耀', 'type': 'V_Dip_Buy', 'params': {'tp':0.08,'sl':0.04,'mfi_s':30,'mfi_b':80}},
    '6261': {'name': '久元', 'type': 'V_Dip_Buy', 'params': {'tp':0.08,'sl':0.04,'mfi_s':30,'mfi_b':80}},
}

stock_data = {}
for file in os.listdir(input_dir):
    if file.endswith(".xlsx"):
        code = file.split('_')[0]
        if code in strategies:
            df = pd.read_excel(os.path.join(input_dir, file))
            df.columns = [str(c) for c in df.columns]
            for col in df.columns[1:]:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            stock_data[code] = df

for code, info in strategies.items():
    if code not in stock_data: continue
    df = stock_data[code]
    strat = info['type']
    best_p = info['params']
    
    wb = Workbook()
    ws_opt = wb.active; ws_opt.title = "參數最佳化結果"
    ws_opt.append(["策略", "參數", "說明"])
    ws_opt.append([strat, str(best_p), "Top 1 最佳參數"])
    
    ws_data = wb.create_sheet("Top1_回測")
    # 寫入參數到 A1:B 區
    ws_data['A1'] = f"【{strat} - Top1 參數】"; ws_data['B1'] = "數值"
    p_keys = list(best_p.keys())
    for idx, pk in enumerate(p_keys, 2):
        ws_data[f'A{idx}'] = pk; ws_data[f'B{idx}'] = best_p[pk]
    shares_row = len(p_keys) + 2
    ws_data[f'A{shares_row}'] = "交易張數 Shares"; ws_data[f'B{shares_row}'] = 1000
    
    # 正確的中文標題
    headers = list(df.columns) + ["進場訊號(P)", "出場訊號(Q)", "持倉狀態(R)", "進場成本(S)", "動作(T)", "交易價格(U)", "單筆損益(V)", "累積損益(W)"]
    for i, t in enumerate(headers, 1):
        ws_data.cell(row=8, column=i, value=t)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 9):
        for c_idx, v in enumerate(row, 1):
            ws_data.cell(row=r_idx, column=c_idx, value=v)
    
    # 寫入公式
    for r in range(9, ws_data.max_row + 1):
        pR = "FALSE" if r == 9 else f"R{r-1}"
        pP = "FALSE" if r == 9 else f"P{r-1}"
        pS = "0" if r == 9 else f"S{r-1}"
        pW = "0" if r == 9 else f"W{r-1}"
        
        if r == 9:
            ws_data[f'P{r}'] = False; ws_data[f'Q{r}'] = False
        else:
            if strat == 'V4':
                ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}>H{r}, K{r}>(L{r}+$B$6), O{r}>$B$4, M{r}>$B$5, R{r}=FALSE))'
                ws_data[f'Q{r}'] = f'=IF({pR}=TRUE, OR((B{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$3), FALSE)'
            elif strat == 'V_Rebound':
                ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}<H{r}, M{r}<$B$4, O{r}>O{r-1}, R{r}=FALSE))'
                ws_data[f'Q{r}'] = f'=IF({pR}=TRUE, OR((B{r}-{pS})/{pS}>=$B$2, (B{r}-{pS})/{pS}<=-$B$3, M{r-1}>$B$5), FALSE)'
            elif strat == 'V_Dip_Buy':
                ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(M{r}<$B$4, K{r}>K{r-1}, O{r}>0, R{r}=FALSE))'
                ws_data[f'Q{r}'] = f'=IF({pR}=TRUE, OR((B{r}-{pS})/{pS}>=$B$2, (B{r}-{pS})/{pS}<=-$B$3, M{r-1}>$B$5), FALSE)'
        
        ws_data[f'R{r}'] = f'=IF({pR}=FALSE, {pP}=TRUE, IF(Q{r}=TRUE, FALSE, TRUE))'
        ws_data[f'S{r}'] = f'=IF(AND({pR}=FALSE, {pP}=TRUE), B{r}, IF(AND({pR}=TRUE, Q{r}=FALSE), {pS}, 0))'
        ws_data[f'T{r}'] = f'=IF(AND({pR}=FALSE, {pP}=TRUE), "買進", IF(Q{r}=TRUE, "賣出", ""))'
        ws_data[f'U{r}'] = f'=IF(T{r}="買進", B{r}, IF(T{r}="賣出", B{r}, 0))'
        ws_data[f'V{r}'] = f'=IF(T{r}="賣出", (U{r} - {pS}) * $B${shares_row}, 0)'
        ws_data[f'W{r}'] = f'={pW} + V{r}'
    
    out_file = os.path.join(out_dir, f"{code}_{info['name']}_{strat}_高勝率回測.xlsx")
    wb.save(out_file)
    print(f"[{code}] Excel 已產出")

print("全部完成！")
