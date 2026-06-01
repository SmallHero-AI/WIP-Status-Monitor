import pandas as pd
import numpy as np
import itertools
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
import os
import glob

# 輸入與輸出路徑
input_dir = "E:\\G-AI-1\\Stock analysis\\Stock original"
output_dir = "E:\\G-AI-1\\Stock analysis"

def generate_comment(row, rank):
    if rank == 1: return "🏆 最高報酬：適合捕捉大趨勢。"
    elif rank <= 5:
        if row['勝率'] > 0.6: return "⭐ 高勝率：穩定度佳。"
        elif row['單筆最大虧損'] > -30000: return "🛡️ 低風險：最大回檔小。"
        elif row['真實投報率(ROI)'] > 0.15: return "🚀 高資金效率：投報率優越。"
        else: return "💡 綜合優異：平衡度高。"
    return ""

def create_excel_with_top5_sheets(base_name, input_file, df, open_arr, close_arr, ma20_arr, ma60_arr, eom_arr, eomsig_arr, mfi_arr, macd_arr, strategy_name, param_names, param_grid, entry_condition_func, exit_condition_func, formula_generator):
    output_file = os.path.join(output_dir, f"{base_name}_{strategy_name}_最佳化與回測.xlsx")
    print(f"[{base_name}] 策略: {strategy_name}")
    
    shares = 1000
    len_df = len(df)
    results = []
    
    keys, values = zip(*param_grid.items())
    combinations = [dict(zip(keys, v)) for v in itertools.product(*values)]
    
    for params in combinations:
        hold, buy_price = False, 0.0
        pnls, entry_prices = [], []
        for i in range(1, len_df - 1):
            if not hold:
                if entry_condition_func(i, close_arr, ma20_arr, ma60_arr, eom_arr, eomsig_arr, mfi_arr, macd_arr, params):
                    hold = True
                    buy_price = open_arr[i+1]
                    entry_prices.append(buy_price)
            else:
                n_open = open_arr[i+1]
                pnl_ratio = (n_open - buy_price) / buy_price if buy_price > 0 else 0
                if exit_condition_func(i, pnl_ratio, close_arr, ma20_arr, ma60_arr, eom_arr, eomsig_arr, mfi_arr, macd_arr, params):
                    hold = False
                    sell_price = n_open
                    pnls.append((sell_price - buy_price) * shares)
                    buy_price = 0.0
                    
        if len(pnls) > 0:
            tc = len(pnls)
            cp = sum(pnls)
            mc = max(entry_prices) * shares if entry_prices else 0
            
            res = params.copy()
            res.update({
                '累積淨損益總額': cp, '最大所需本金': mc,
                '真實投報率(ROI)': cp / mc if mc > 0 else 0,
                '交易次數': tc, '勝率': len([p for p in pnls if p > 0]) / tc,
                '單筆最大獲利': max(pnls), '單筆最大虧損': min(pnls),
                '平均每筆損益': cp / tc,
                '期望報酬率(每筆)': (cp / tc) / ((sum(entry_prices)/len(entry_prices))*shares) if entry_prices else 0
            })
            results.append(res)
            
    res_df = pd.DataFrame(results)
    if not res_df.empty:
        res_df = res_df[res_df['交易次數'] >= 2].sort_values(by='累積淨損益總額', ascending=False).reset_index(drop=True)
        res_df['最佳參數註解'] = [generate_comment(row, i+1) for i, row in res_df.iterrows()]
    
    wb = Workbook()
    ws_opt = wb.active
    ws_opt.title = "參數最佳化結果"
    
    header_names = param_names + [
        '累積淨損益總額', '最大所需本金', '真實投報率(ROI)',
        '交易次數', '勝率', '單筆最大獲利', '單筆最大虧損',
        '平均每筆損益', '期望報酬率(每筆)', '🤖 AI 策略評估與註解'
    ]
    ws_opt.append(header_names)
    
    header_font, header_fill, top5_fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1E40AF"), PatternFill("solid", fgColor="FEF08A")
    for cell in ws_opt[1]:
        cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="center")
        
    if not res_df.empty:
        for r_idx, row in res_df.iterrows():
            ws_opt.append([row[col] for col in res_df.columns])
            cr = ws_opt.max_row
            for col_idx, col_name in enumerate(param_names, 1):
                if 'TP' in col_name or 'SL' in col_name or '%' in col_name: ws_opt.cell(row=cr, column=col_idx).number_format = '0.0%'
            offset = len(param_names)
            for col in [1, 2, 6, 7, 8]: ws_opt.cell(row=cr, column=offset+col).number_format = '#,##0'
            for col in [3, 5, 9]: ws_opt.cell(row=cr, column=offset+col).number_format = '0.00%'
            if r_idx < 5:
                for col_num in range(1, len(header_names) + 1):
                    ws_opt.cell(row=cr, column=col_num).fill = top5_fill; ws_opt.cell(row=cr, column=col_num).font = Font(bold=True)
                    
    for col in ws_opt.columns:
        ml = max([len(str(cell.value)) for cell in col if cell.value] + [0])
        ws_opt.column_dimensions[col[0].column_letter].width = min(ml + 2, 40)
        
    headers = list(df.columns) + ["進場訊號(P)", "出場訊號(Q)", "持倉狀態(R)", "進場成本(S)", "動作(T)", "交易價格(U)", "單筆損益(V)", "累積損益(W)"]
    
    loop_count = min(5, len(res_df)) if not res_df.empty else 1
    for top_idx in range(loop_count):
        ws_data = wb.create_sheet(f"Top{top_idx+1}_回測")
        if res_df.empty:
            best_params = {k: v[0] for k, v in param_grid.items()} # fallback
        else:
            best_params = res_df.iloc[top_idx]
            
        ws_data['A1'] = f"【{strategy_name} - Top{top_idx+1} 參數】"
        ws_data['B1'] = "數值 (可直接修改)"
        for idx, p_name in enumerate(param_names, 2):
            ws_data[f'A{idx}'] = p_name
            ws_data[f'B{idx}'] = best_params[p_name]
        
        var_count = len(param_names)
        shares_row = var_count + 2
        ws_data[f'A{shares_row}'] = "交易張數 Shares"
        ws_data[f'B{shares_row}'] = 1000
        
        for i, title in enumerate(headers, 1): ws_data.cell(row=8, column=i, value=title)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 9):
            for c_idx, value in enumerate(row, 1): ws_data.cell(row=r_idx, column=c_idx, value=value)
            
        formula_generator(ws_data, 9, ws_data.max_row, param_names, shares_row)
        
        for col in ['A', 'B']:
            for row in range(1, shares_row + 1):
                ws_data[f'{col}{row}'].font = Font(bold=True); ws_data[f'{col}{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    wb.save(output_file)

# 策略 1：穩健均線順勢策略 (適合大型權值股、穩定成長股)
def strat_1_trend(base_name, input_file, df, *arrs):
    param_names = ['TP(停利)', 'SL(停損)', 'MACD_Th']
    param_grid = {'TP(停利)': [0.08, 0.12, 0.15], 'SL(停損)': [0.04, 0.06, 0.08], 'MACD_Th': [-0.5, 0.0, 0.5]}
    def entry_cond(i, c, m20, m60, e, esig, mfi, macd, p): return c[i] > m20[i] and macd[i] > p['MACD_Th']
    def exit_cond(i, pnl, c, m20, m60, e, esig, mfi, macd, p): return pnl >= p['TP(停利)'] or pnl <= -p['SL(停損)']
    def formulas(ws, s_row, m_row, p_names, sh_row):
        for r in range(s_row, m_row + 1):
            pR, pP, pS, pW = ("FALSE", "FALSE", "0", "0") if r == s_row else (f"R{r-1}", f"P{r-1}", f"S{r-1}", f"W{r-1}")
            ws[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}>H{r}, O{r}>$B$4, {pR}=FALSE))'
            ws[f'Q{r}'] = f'=IF({pR}=TRUE, OR((B{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$3), FALSE)'
            ws[f'R{r}'] = f'=IF({pR}=FALSE, {pP}=TRUE, IF(Q{r}=TRUE, FALSE, TRUE))'
            ws[f'S{r}'] = f'=IF(AND({pR}=FALSE, {pP}=TRUE), B{r}, IF(AND({pR}=TRUE, Q{r}=FALSE), {pS}, 0))'
            ws[f'T{r}'] = f'=IF(AND({pR}=FALSE, {pP}=TRUE), "買進", IF(Q{r}=TRUE, "賣出", ""))'
            ws[f'U{r}'] = f'=IF(T{r}="買進", B{r}, IF(T{r}="賣出", B{r}, 0))'
            ws[f'V{r}'] = f'=IF(T{r}="賣出", (U{r} - {pS}) * $B${sh_row}, 0)'
            ws[f'W{r}'] = f'={pW} + V{r}'
    create_excel_with_top5_sheets(base_name, input_file, df, *arrs, "穩健均線順勢", param_names, param_grid, entry_cond, exit_cond, formulas)

# 策略 2：高波動趨勢動能爆發 (適合 IC 設計、生技等高波動股)
def strat_2_momentum(base_name, input_file, df, *arrs):
    param_names = ['TP(停利)', 'SL(停損)', 'MACD_Th']
    param_grid = {'TP(停利)': [0.15, 0.20, 0.25], 'SL(停損)': [0.05, 0.08, 0.10], 'MACD_Th': [-1.0, 0.0]}
    def entry_cond(i, c, m20, m60, e, esig, mfi, macd, p): return c[i] > m60[i] and e[i] > esig[i] and macd[i] > p['MACD_Th']
    def exit_cond(i, pnl, c, m20, m60, e, esig, mfi, macd, p): return pnl >= p['TP(停利)'] or pnl <= -p['SL(停損)'] or e[i] < esig[i]
    def formulas(ws, s_row, m_row, p_names, sh_row):
        for r in range(s_row, m_row + 1):
            pR, pP, pS, pW = ("FALSE", "FALSE", "0", "0") if r == s_row else (f"R{r-1}", f"P{r-1}", f"S{r-1}", f"W{r-1}")
            ws[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}>I{r}, K{r}>L{r}, O{r}>$B$4, {pR}=FALSE))'
            ws[f'Q{r}'] = f'=IF({pR}=TRUE, OR((B{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$3, K{r}<L{r}), FALSE)'
            ws[f'R{r}'] = f'=IF({pR}=FALSE, {pP}=TRUE, IF(Q{r}=TRUE, FALSE, TRUE))'
            ws[f'S{r}'] = f'=IF(AND({pR}=FALSE, {pP}=TRUE), B{r}, IF(AND({pR}=TRUE, Q{r}=FALSE), {pS}, 0))'
            ws[f'T{r}'] = f'=IF(AND({pR}=FALSE, {pP}=TRUE), "買進", IF(Q{r}=TRUE, "賣出", ""))'
            ws[f'U{r}'] = f'=IF(T{r}="買進", B{r}, IF(T{r}="賣出", B{r}, 0))'
            ws[f'V{r}'] = f'=IF(T{r}="賣出", (U{r} - {pS}) * $B${sh_row}, 0)'
            ws[f'W{r}'] = f'={pW} + V{r}'
    create_excel_with_top5_sheets(base_name, input_file, df, *arrs, "高波動動能爆發", param_names, param_grid, entry_cond, exit_cond, formulas)

# 策略 3：波段反轉低接 (適合區間震盪、景氣循環股如 PCB)
def strat_3_swing(base_name, input_file, df, *arrs):
    param_names = ['TP(停利)', 'SL(停損)', 'MFI_超賣區', 'MFI_超買區']
    param_grid = {'TP(停利)': [0.08, 0.12, 0.16], 'SL(停損)': [0.05, 0.07], 'MFI_超賣區': [35, 45], 'MFI_超買區': [75, 85]}
    def entry_cond(i, c, m20, m60, e, esig, mfi, macd, p): return c[i] > m20[i] and mfi[i] < p['MFI_超賣區']
    def exit_cond(i, pnl, c, m20, m60, e, esig, mfi, macd, p): return pnl >= p['TP(停利)'] or pnl <= -p['SL(停損)'] or mfi[i] > p['MFI_超買區']
    def formulas(ws, s_row, m_row, p_names, sh_row):
        for r in range(s_row, m_row + 1):
            pR, pP, pS, pW = ("FALSE", "FALSE", "0", "0") if r == s_row else (f"R{r-1}", f"P{r-1}", f"S{r-1}", f"W{r-1}")
            ws[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}>H{r}, M{r}<$B$4, {pR}=FALSE))'
            ws[f'Q{r}'] = f'=IF({pR}=TRUE, OR((B{r}-{pS})/{pS} >= $B$2, (B{r}-{pS})/{pS} <= -$B$3, M{r}>$B$5), FALSE)'
            ws[f'R{r}'] = f'=IF({pR}=FALSE, {pP}=TRUE, IF(Q{r}=TRUE, FALSE, TRUE))'
            ws[f'S{r}'] = f'=IF(AND({pR}=FALSE, {pP}=TRUE), B{r}, IF(AND({pR}=TRUE, Q{r}=FALSE), {pS}, 0))'
            ws[f'T{r}'] = f'=IF(AND({pR}=FALSE, {pP}=TRUE), "買進", IF(Q{r}=TRUE, "賣出", ""))'
            ws[f'U{r}'] = f'=IF(T{r}="買進", B{r}, IF(T{r}="賣出", B{r}, 0))'
            ws[f'V{r}'] = f'=IF(T{r}="賣出", (U{r} - {pS}) * $B${sh_row}, 0)'
            ws[f'W{r}'] = f'={pW} + V{r}'
    create_excel_with_top5_sheets(base_name, input_file, df, *arrs, "波段反轉低接", param_names, param_grid, entry_cond, exit_cond, formulas)

# 依據股票代號或名稱指派策略
strategy_map = {
    '2330': strat_1_trend, '6669': strat_1_trend, '2360': strat_1_trend,
    '3443': strat_2_momentum, '3455': strat_2_momentum, '3535': strat_2_momentum, '4908': strat_2_momentum, '4746': strat_2_momentum,
    '3189': strat_3_swing, '6269': strat_3_swing, '6205': strat_3_swing, '6261': strat_3_swing
}

all_files = glob.glob(os.path.join(input_dir, "*.xlsx"))
print(f"找到 {len(all_files)} 個原始檔案，開始批次處理...")

for file_path in all_files:
    filename = os.path.basename(file_path)
    base_name = filename.split('_EOM')[0]
    stock_code = base_name.split('_')[0]
    
    # 指派策略，若未命中則預設使用 strat_1
    strat_func = strategy_map.get(stock_code, strat_1_trend)
    
    df = pd.read_excel(file_path)
    df.columns = [str(c) for c in df.columns]
    
    open_col, close_col, ma20_col = df.columns[1], df.columns[4], df.columns[7]
    ma60_col, eom_col, eomsig_col = df.columns[8], df.columns[10], df.columns[11]
    mfi_col, macd_col = df.columns[12], df.columns[14]
    
    for col in [open_col, close_col, ma20_col, ma60_col, eom_col, eomsig_col, mfi_col, macd_col]:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    open_arr = df[open_col].values; close_arr = df[close_col].values
    ma20_arr = df[ma20_col].values; ma60_arr = df[ma60_col].values
    eom_arr = df[eom_col].values; eomsig_arr = df[eomsig_col].values
    mfi_arr = df[mfi_col].values; macd_arr = df[macd_col].values
    
    strat_func(base_name, file_path, df, open_arr, close_arr, ma20_arr, ma60_arr, eom_arr, eomsig_arr, mfi_arr, macd_arr)

print("✅ 所有個股皆已完成專屬策略處理與匯出！")
