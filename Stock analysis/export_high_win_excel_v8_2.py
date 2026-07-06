# -*- coding: utf-8 -*-
import os
import glob
import json
import shutil
import re
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── 路徑設定 ──
SCRIPT_DIR = r"E:\G-AI-1\Stock analysis"
STOCK_ORIGINAL_DIR = os.path.join(SCRIPT_DIR, "Stock original")
AUTO_EXPORT_DIR = os.path.join(STOCK_ORIGINAL_DIR, "auto_export")
LEADERBOARD_PATH = os.path.join(SCRIPT_DIR, "修正版_V6_Server", "public", "leaderboard_v8_2.json")
SERVER_DATA_DIR = os.path.join(SCRIPT_DIR, "修正版_V6_Server", "data")

def get_file_date(filepath):
    filename = os.path.basename(filepath)
    match = re.search(r'\d{8}', filename)
    if match:
        return int(match.group(0))
    return 0

def find_column(df, keywords):
    for col in df.columns:
        for kw in keywords:
            if kw.lower() in str(col).lower():
                return col
    return None

def clean_and_parse_series(df, col_name):
    if col_name is not None and col_name in df.columns:
        s = df[col_name].astype(str).str.replace(r'[▲▼↑↓△▽,%\s]', '', regex=True)
        return pd.to_numeric(s, errors='coerce').ffill().bfill().fillna(0).values
    return np.zeros(len(df))

def get_strategy_signals(df, ent_name, ext_name, is_short=False):
    """
    計算對應的進出場訊號 Boolean 陣列
    """
    n_rows = len(df)
    
    # 匹配技術指標欄位
    c_open = find_column(df, ['開盤價', '開盤', 'Open', 'open'])
    c_high = find_column(df, ['最高價', '最高', 'High', 'high'])
    c_low = find_column(df, ['最低價', '最低', 'Low', 'low'])
    c_close = find_column(df, ['收盤價', '收盤', 'Close', 'close'])
    c_ma5 = find_column(df, ['均價[5]', 'MA5', 'ma5'])
    c_ma20 = find_column(df, ['均價[20]', 'MA20', 'ma20'])
    c_ma60 = find_column(df, ['均價[60]', 'MA60', 'ma60'])
    c_ma120 = find_column(df, ['均價[120]', 'MA120', 'ma120'])
    c_eom = find_column(df, ['EOM[60]', 'EOM'])
    c_eom_sig = find_column(df, ['Signal[20]', 'EOM_Signal'])
    c_mfi = find_column(df, ['MFI[14]', 'MFI'])
    c_macd = find_column(df, ['MACD', 'MACD柱', 'MACD 柱', 'MACD 柱狀體'])
    c_k = find_column(df, ['%KS', 'K(9,3,3)', 'K值'])
    c_d = find_column(df, ['%DS', 'D(9,3,3)', 'D值'])
    c_rsi = find_column(df, ['RSI[14]', 'RSI'])
    c_bias20 = find_column(df, ['BIAS[20]', 'BIAS20'])
    c_cci = find_column(df, ['CCI[20]', 'CCI'])
    c_bb_upper = find_column(df, ['布林加通道 上軌', 'BB_Upper'])
    c_bb_lower = find_column(df, ['布林加通道 下軌', 'BB_Lower'])
    c_pivot = find_column(df, ['Pivot', 'pivot'])
    c_r1 = find_column(df, ['第1道壓力', 'R1'])
    c_s1 = find_column(df, ['第1道支撐', 'S1'])

    # 轉換陣列
    opens = clean_and_parse_series(df, c_open)
    highs = clean_and_parse_series(df, c_high)
    lows = clean_and_parse_series(df, c_low)
    closes = clean_and_parse_series(df, c_close)
    ma5 = clean_and_parse_series(df, c_ma5)
    ma20 = clean_and_parse_series(df, c_ma20)
    ma60 = clean_and_parse_series(df, c_ma60)
    ma120 = clean_and_parse_series(df, c_ma120)
    eom = clean_and_parse_series(df, c_eom)
    eom_sig = clean_and_parse_series(df, c_eom_sig)
    mfi = clean_and_parse_series(df, c_mfi)
    macd = clean_and_parse_series(df, c_macd)
    k = clean_and_parse_series(df, c_k)
    d = clean_and_parse_series(df, c_d)
    rsi = clean_and_parse_series(df, c_rsi)
    bias20 = clean_and_parse_series(df, c_bias20)
    cci = clean_and_parse_series(df, c_cci)
    bb_upper = clean_and_parse_series(df, c_bb_upper)
    bb_lower = clean_and_parse_series(df, c_bb_lower)
    pivot = clean_and_parse_series(df, c_pivot)
    r1 = clean_and_parse_series(df, c_r1)
    s1 = clean_and_parse_series(df, c_s1)

    c_volume = find_column(df, ['成交量', '量', 'Volume', 'volume'])
    vols = clean_and_parse_series(df, c_volume)

    # 計算 ATR[14] - 向量化 (取代 Python for 迴圈，速度提升 ~100x)
    prev_closes_e = np.roll(closes, 1); prev_closes_e[0] = closes[0]
    tr = np.maximum(highs - lows, np.maximum(np.abs(highs - prev_closes_e), np.abs(lows - prev_closes_e)))
    atr = pd.Series(tr).rolling(14).mean().values
    atr = np.nan_to_num(atr, nan=tr[0] if n_rows > 0 else 0)

    # 成交量 MA20 - 向量化
    vol_ma20 = pd.Series(vols).rolling(20).mean().values
    vol_ma20 = np.nan_to_num(vol_ma20, nan=vols[0] if len(vols) > 0 else 1)

    # 真突破判定指標
    f_volume_confirm = (vols > 1.5 * vol_ma20) & (vols > 0)
    f_volume_confirm_short = (vols > 1.2 * vol_ma20) & (vols > 0)
    f_candle_strength = (closes - lows) > 0.6 * (highs - lows)
    f_candle_strength_short = (highs - closes) > 0.6 * (highs - lows)
    f_trend_align = closes > ma60
    f_trend_align_short = closes < ma60

    prev_k = np.roll(k, 1); prev_k[0] = k[0]
    prev_d = np.roll(d, 1); prev_d[0] = d[0]
    prev_macd = np.roll(macd, 1); prev_macd[0] = macd[0]

    c_low_col = find_column(df, ['最低價', '最低', 'Low', 'low'])
    low_vals = df[c_low_col].values if c_low_col else lows
    c_high_col = find_column(df, ['最高價', '最高', 'High', 'high'])
    high_vals = df[c_high_col].values if c_high_col else highs

    # ── 自適應多尺度「型態」運算 (V8.2 新增) ──
    f_w_bottom = np.zeros(n_rows, dtype=bool)
    f_hs_bottom = np.zeros(n_rows, dtype=bool)
    f_v_rebound = np.zeros(n_rows, dtype=bool)
    f_triangle_breakout = np.zeros(n_rows, dtype=bool)
    f_box_breakout = np.zeros(n_rows, dtype=bool)

    f_m_top = np.zeros(n_rows, dtype=bool)
    f_hs_top = np.zeros(n_rows, dtype=bool)
    f_inverted_v = np.zeros(n_rows, dtype=bool)
    f_triangle_breakdown = np.zeros(n_rows, dtype=bool)
    f_box_breakdown = np.zeros(n_rows, dtype=bool)

    for S in [15, 30, 60]:
        if n_rows < S + 5:
            continue
        
        # W底 (雙重底)
        v1_idx_shift = S
        v1_len = max(5, int(S * 0.6))
        v2_idx_shift = 2
        v2_len = max(4, int(S * 0.4))
        peak_len = max(4, int(S * 0.4))
        
        min1 = pd.Series(lows).shift(v1_idx_shift).rolling(v1_len).min().values
        min2 = pd.Series(lows).shift(v2_idx_shift).rolling(v2_len).min().values
        peak = pd.Series(highs).shift(v2_len).rolling(peak_len).max().values
        
        c_w_base = (np.abs(min1 - min2) / (min1 + 1e-5) < 0.03) & (peak > min1)
        c_w_break = (closes > peak) & (np.roll(closes, 1) <= peak)
        f_w_bottom = f_w_bottom | (c_w_base & c_w_break)
        
        # 頭肩底
        s_div = max(3, int(S / 3))
        v1_hs = pd.Series(lows).shift(s_div * 2).rolling(s_div).min().values
        v2_hs = pd.Series(lows).shift(s_div).rolling(s_div).min().values
        v3_hs = pd.Series(lows).shift(2).rolling(s_div).min().values
        peak_hs = pd.Series(highs).shift(2).rolling(S).max().values
        
        c_hs_base = (v2_hs < v1_hs) & (v2_hs < v3_hs) & (np.abs(v1_hs - v3_hs) / (v1_hs + 1e-5) < 0.04)
        c_hs_break = (closes > peak_hs) & (np.roll(closes, 1) <= peak_hs)
        f_hs_bottom = f_hs_bottom | (c_hs_base & c_hs_break)

        # V型反轉 (短線急跌後爆量彈升)
        fall_len = max(4, int(S * 0.7))
        rise_len = max(3, int(S * 0.3))
        fall_steep = pd.Series(closes).shift(rise_len).pct_change(fall_len).values < -0.12
        rise_sharp = pd.Series(closes).pct_change(rise_len).values > 0.08
        f_v_rebound = f_v_rebound | (fall_steep & rise_sharp)

        # 三角收斂
        range_S = pd.Series(highs - lows).rolling(S).mean().values
        range_prev = pd.Series(highs - lows).shift(S).rolling(S).mean().values
        volatility_shrink = range_S < (range_prev * 0.7)
        bound_upper = pd.Series(highs).shift(1).rolling(S).max().values
        bound_lower = pd.Series(lows).shift(1).rolling(S).min().values
        f_triangle_breakout = f_triangle_breakout | (volatility_shrink & (closes > bound_upper))
        f_triangle_breakdown = f_triangle_breakdown | (volatility_shrink & (closes < bound_lower))

        # 箱型整理 (矩形)
        box_high = pd.Series(highs).shift(1).rolling(S).max().values
        box_low = pd.Series(lows).shift(1).rolling(S).min().values
        box_range_tight = (box_high - box_low) / (box_low + 1e-5) < 0.06
        f_box_breakout = f_box_breakout | (box_range_tight & (closes > box_high))
        f_box_breakdown = f_box_breakdown | (box_range_tight & (closes < box_low))

        # M頭 (雙重頂)
        max1 = pd.Series(highs).shift(v1_idx_shift).rolling(v1_len).max().values
        max2 = pd.Series(highs).shift(v2_idx_shift).rolling(v2_len).max().values
        valley = pd.Series(lows).shift(v2_len).rolling(peak_len).min().values
        c_m_base = (np.abs(max1 - max2) / (max1 + 1e-5) < 0.03) & (valley < max1)
        c_m_break = (closes < valley) & (np.roll(closes, 1) >= valley)
        f_m_top = f_m_top | (c_m_base & c_m_break)

        # 頭肩頂
        p1_hs = pd.Series(highs).shift(s_div * 2).rolling(s_div).max().values
        p2_hs = pd.Series(highs).shift(s_div).rolling(s_div).max().values
        p3_hs = pd.Series(highs).shift(2).rolling(s_div).max().values
        valley_hs = pd.Series(lows).shift(2).rolling(S).min().values
        c_hst_base = (p2_hs > p1_hs) & (p2_hs > p3_hs) & (np.abs(p1_hs - p3_hs) / (p1_hs + 1e-5) < 0.04)
        c_hst_break = (closes < valley_hs) & (np.roll(closes, 1) >= valley_hs)
        f_hs_top = f_hs_top | (c_hst_base & c_hst_break)

        # 倒V型反轉 (急漲後下殺)
        rise_steep = pd.Series(closes).shift(rise_len).pct_change(fall_len).values > 0.12
        fall_sharp = pd.Series(closes).pct_change(rise_len).values < -0.08
        f_inverted_v = f_inverted_v | (rise_steep & fall_sharp)

    # 應用真突破判定條件 (量能、K線實體強度、趨勢方向)
    f_w_bottom = f_w_bottom & f_volume_confirm & f_candle_strength & f_trend_align
    f_hs_bottom = f_hs_bottom & f_volume_confirm & f_candle_strength & f_trend_align
    f_v_rebound = f_v_rebound & f_volume_confirm & f_candle_strength & f_trend_align
    f_triangle_breakout = f_triangle_breakout & f_volume_confirm & f_candle_strength & f_trend_align
    f_box_breakout = f_box_breakout & f_volume_confirm & f_candle_strength & f_trend_align

    f_m_top = f_m_top & f_volume_confirm_short & f_candle_strength_short & f_trend_align_short
    f_hs_top = f_hs_top & f_volume_confirm_short & f_candle_strength_short & f_trend_align_short
    f_inverted_v = f_inverted_v & f_volume_confirm_short & f_candle_strength_short & f_trend_align_short
    f_triangle_breakdown = f_triangle_breakdown & f_volume_confirm_short & f_candle_strength_short & f_trend_align_short
    f_box_breakdown = f_box_breakdown & f_volume_confirm_short & f_candle_strength_short & f_trend_align_short

    ent_sig_raw = np.zeros(n_rows, dtype=bool)
    ext_sig_raw = np.zeros(n_rows, dtype=bool)
    tp, sl = None, None

    if not is_short:
        # 多單買入/賣出訊號
        f_kd_gold = (k > d) & (prev_k <= prev_d)
        f_rsi_oversold = rsi < 35
        f_rsi_bull = rsi > 50
        f_macd_grow = macd > prev_macd
        f_macd_positive = macd > 0
        f_bias_oversold = bias20 < -4
        f_cci_oversold = cci < -100
        f_eom_bullish = eom > eom_sig
        f_mfi_oversold = mfi < 30
        f_bb_oversold = (closes < bb_lower) & (bb_lower > 0)
        f_above_pivot = closes > pivot
        f_break_r1 = (closes > r1) & (r1 > 0)
        f_ma_bull = (ma5 > ma20) & (ma20 > ma60)
        f_above_ma20 = closes > ma20
        f_above_ma60 = closes > ma60
        f_above_ma120 = closes > ma120

        f_kd_dead = (k < d) & (prev_k >= prev_d)
        f_rsi_overbought = rsi > 65
        f_macd_shrink = macd < prev_macd
        f_below_ma20 = closes < ma20
        f_bb_overbought = (closes > bb_upper) & (bb_upper > 0)
        f_below_s1 = (closes < s1) & (s1 > 0)

        # 匹配進場
        if ent_name == "強勢均線多頭+KD金叉": ent_sig_raw = f_ma_bull & f_kd_gold & f_above_ma20
        elif ent_name == "中長期多頭+RSI低檔轉強": ent_sig_raw = f_above_ma60 & f_above_ma120 & (rsi < 45) & f_macd_grow
        elif ent_name == "雙重均線突破+MACD翻紅": ent_sig_raw = f_above_ma20 & f_above_ma60 & f_macd_positive & f_macd_grow
        elif ent_name == "EOM動能突破+均線多頭": ent_sig_raw = f_eom_bullish & f_ma_bull & f_above_ma20
        elif ent_name == "極度超跌共振 (BIAS+RSI+MFI)": ent_sig_raw = f_bias_oversold & f_rsi_oversold & f_mfi_oversold
        elif ent_name == "布林通道下軌+KD低檔金叉": ent_sig_raw = f_bb_oversold & f_kd_gold
        elif ent_name == "CCI超賣+MFI超賣+KD金叉": ent_sig_raw = f_cci_oversold & f_mfi_oversold & f_kd_gold
        elif ent_name == "雙重超賣 (RSI+CCI)+MACD轉強": ent_sig_raw = f_rsi_oversold & f_cci_oversold & f_macd_grow
        elif ent_name == "樞軸點(S1)支撐+KD金叉": ent_sig_raw = (closes > s1) & (low_vals <= s1) & f_kd_gold & (s1 > 0)
        elif ent_name == "突破阻力(R1)+MACD多頭": ent_sig_raw = f_break_r1 & f_macd_positive & f_above_ma20
        elif ent_name == "突破樞軸(Pivot)+RSI轉強": ent_sig_raw = f_above_pivot & f_rsi_bull & f_macd_grow
        elif ent_name == "S1支撐不破+MACD紅柱增長": ent_sig_raw = (closes > s1) & f_macd_grow & (s1 > 0)
        elif ent_name == "多頭拉回：MA60之上+KD金叉+RSI<45": ent_sig_raw = f_above_ma60 & f_kd_gold & (rsi < 45)
        elif ent_name == "布林中軌支撐+MACD紅柱": ent_sig_raw = (closes > bb_lower) & (closes < bb_upper) & f_above_ma20 & f_macd_grow
        elif ent_name == "動能共振：EOM突破+MACD>0+RSI>50": ent_sig_raw = f_eom_bullish & f_macd_positive & f_rsi_bull
        elif ent_name == "雙保險超跌：BIAS超賣+布林下軌觸及": ent_sig_raw = f_bias_oversold & f_bb_oversold
        elif ent_name == "主力資金流入：MFI超賣+EOM突破": ent_sig_raw = f_mfi_oversold & f_eom_bullish
        elif ent_name == "壓力突破：價格>R1+EOM強勢+MACD>0": ent_sig_raw = f_break_r1 & f_eom_bullish & f_macd_positive
        elif ent_name == "長期均線支撐：站上MA120+KD金叉+CCI超賣": ent_sig_raw = f_above_ma120 & f_kd_gold & f_cci_oversold
        elif ent_name == "CCI超賣反彈+MACD紅柱增長": ent_sig_raw = f_cci_oversold & f_macd_grow & f_above_ma20
        elif ent_name == "W底雙重底突破頸線 (多頭趨勢)": ent_sig_raw = f_w_bottom
        elif ent_name == "頭肩底突破頸線確立 (多頭趨勢)": ent_sig_raw = f_hs_bottom
        elif ent_name == "V型反轉爆量向上拉升 (多頭趨勢)": ent_sig_raw = f_v_rebound
        elif ent_name == "三角收斂突破上軌加速 (多頭整理)": ent_sig_raw = f_triangle_breakout
        elif ent_name == "箱型整理突破阻力上限 (多頭整理)": ent_sig_raw = f_box_breakout

        # 匹配出場
        if ext_name == "6%停利 / 6%停損 (穩健勝率)": tp, sl = 0.06, 0.06
        elif ext_name == "8%停利 / 8%停損 (均衡配置)": tp, sl = 0.08, 0.08
        elif ext_name == "12%停利 / 6%停損 (高盈虧比)": tp, sl = 0.12, 0.06
        elif ext_name == "KD死叉離場或6%停損": sl, ext_sig_raw = 0.06, f_kd_dead
        elif ext_name == "RSI超買離場或6%停損": sl, ext_sig_raw = 0.06, f_rsi_overbought
        elif ext_name == "MACD紅柱縮短離場或6%停損": sl, ext_sig_raw = 0.06, f_macd_shrink
        elif ext_name == "收盤跌破MA20或6%停損": sl, ext_sig_raw = 0.06, f_below_ma20
        elif ext_name == "觸碰布林上軌停利或6%停損": sl, ext_sig_raw = 0.06, f_bb_overbought
        elif ext_name == "跌破樞軸支撐(S1)或8%停利": tp, ext_sig_raw = 0.08, f_below_s1
        # 波段離場
        elif ext_name == "波段追蹤：關閉固定停利 + 跌破MA20下彎離場 (波段操作)": pass
        elif ext_name == "波段追蹤：ATR追蹤停損 (最高價拉回3*ATR離場)": pass
        elif ext_name == "波段追蹤：跌破前10日最低點停損 (波段操作)": pass
    else:
        # 空單放空/回補訊號
        f_ma_bear = (ma5 < ma20) & (ma20 < ma60)
        f_below_pivot = closes < pivot
        f_break_s1 = (closes < s1) & (s1 > 0)
        f_bias_overbought = bias20 > 4
        f_cci_overbought = cci > 100
        f_eom_bearish = eom < eom_sig
        f_mfi_overbought = mfi > 70
        f_bb_overbought_short = (closes > bb_upper) & (bb_upper > 0)

        f_kd_dead = (k < d) & (prev_k >= prev_d)
        f_rsi_overbought = rsi > 65
        f_macd_shrink = macd < prev_macd
        f_below_ma20 = closes < ma20
        f_below_ma60 = closes < ma60
        f_below_ma120 = closes < ma120

        f_kd_gold_cover = (k > d) & (prev_k <= prev_d)
        f_rsi_oversold_cover = rsi < 35
        f_macd_grow_cover = macd > prev_macd
        f_above_ma20_cover = closes > ma20
        f_bb_oversold_cover = (closes < bb_lower) & (bb_lower > 0)
        f_above_r1_cover = (closes > r1) & (r1 > 0)

        # 匹配進場 (空單)
        if ent_name == "弱勢均線空頭+KD死叉": ent_sig_raw = f_ma_bear & f_kd_dead & f_below_ma20
        elif ent_name == "中長期空頭+RSI高檔轉弱": ent_sig_raw = f_below_ma60 & f_below_ma120 & (rsi > 55) & f_macd_shrink
        elif ent_name == "雙重均線跌破+MACD翻綠": ent_sig_raw = f_below_ma20 & f_below_ma60 & (macd < 0) & f_macd_shrink
        elif ent_name == "EOM動能下跌+均線空頭": ent_sig_raw = f_eom_bearish & f_ma_bear & f_below_ma20
        elif ent_name == "極度超買共振 (BIAS+RSI+MFI)": ent_sig_raw = f_bias_overbought & f_rsi_overbought & f_mfi_overbought
        elif ent_name == "布林通道上軌+KD高檔死叉": ent_sig_raw = f_bb_overbought_short & f_kd_dead
        elif ent_name == "CCI超買+MFI超買+KD死叉": ent_sig_raw = f_cci_overbought & f_mfi_overbought & f_kd_dead
        elif ent_name == "雙重超買 (RSI+CCI)+MACD轉弱": ent_sig_raw = f_rsi_overbought & f_cci_overbought & f_macd_shrink
        elif ent_name == "樞軸點(R1)壓力+KD死叉": ent_sig_raw = (closes < r1) & (high_vals >= r1) & f_kd_dead & (r1 > 0)
        elif ent_name == "跌破支撐(S1)+MACD空頭": ent_sig_raw = f_break_s1 & (macd < 0) & f_below_ma20
        elif ent_name == "跌破樞軸(Pivot)+RSI轉弱": ent_sig_raw = f_below_pivot & (rsi < 50) & f_macd_shrink
        elif ent_name == "R1壓力不破+MACD綠柱增長": ent_sig_raw = (closes < r1) & f_macd_shrink & (r1 > 0)
        elif ent_name == "空頭拉回：MA60之下+KD死叉+RSI>55": ent_sig_raw = f_below_ma60 & f_kd_dead & (rsi > 55)
        elif ent_name == "布林中軌阻力+MACD綠柱": ent_sig_raw = (closes > bb_lower) & (closes < bb_upper) & f_below_ma20 & f_macd_shrink
        elif ent_name == "動能共振：EOM下跌+MACD<0+RSI<50": ent_sig_raw = f_eom_bearish & (macd < 0) & (rsi < 50)
        elif ent_name == "雙保險超買：BIAS超買+布林上軌觸及": ent_sig_raw = f_bias_overbought & f_bb_overbought_short
        elif ent_name == "主力資金流出：MFI超買+EOM下跌": ent_sig_raw = f_mfi_overbought & f_eom_bearish
        elif ent_name == "支撐跌破：價格<S1+EOM弱勢+MACD<0": ent_sig_raw = f_break_s1 & f_eom_bearish & (macd < 0)
        elif ent_name == "長期均線阻力：跌破MA120+KD死叉+CCI超買": ent_sig_raw = (closes < ma120) & f_kd_dead & f_cci_overbought
        elif ent_name == "CCI超買回檔+MACD綠柱增長": ent_sig_raw = f_cci_overbought & f_macd_shrink & f_below_ma20
        elif ent_name == "M頭雙重頂跌破頸線 (空頭趨勢)": ent_sig_raw = f_m_top
        elif ent_name == "頭肩頂跌破頸線反轉 (空頭趨勢)": ent_sig_raw = f_hs_top
        elif ent_name == "倒V型反轉急速暴跌 (空頭趨勢)": ent_sig_raw = f_inverted_v
        elif ent_name == "三角收斂跌破下軌加速 (空頭整理)": ent_sig_raw = f_triangle_breakdown
        elif ent_name == "箱型整理跌破支撐下限 (空頭整理)": ent_sig_raw = f_box_breakdown

        # 匹配出場 (空單)
        if ext_name == "6%停利 / 6%停損 (穩健勝率)": tp, sl = 0.06, 0.06
        elif ext_name == "8%停利 / 8%停損 (均衡配置)": tp, sl = 0.08, 0.08
        elif ext_name == "12%停利 / 6%停損 (高盈虧比)": tp, sl = 0.12, 0.06
        elif ext_name == "KD金叉離場或6%停損": sl, ext_sig_raw = 0.06, f_kd_gold_cover
        elif ext_name == "RSI超賣離場或6%停損": sl, ext_sig_raw = 0.06, f_rsi_oversold_cover
        elif ext_name == "MACD綠柱縮短離場或6%停損": sl, ext_sig_raw = 0.06, f_macd_grow_cover
        elif ext_name == "收盤站上MA20或6%停損": sl, ext_sig_raw = 0.06, f_above_ma20_cover
        elif ext_name == "觸碰布林下軌停利或6%停損": sl, ext_sig_raw = 0.06, f_bb_oversold_cover
        elif ext_name == "突破壓力(R1)或8%停利": tp, ext_sig_raw = 0.08, f_above_r1_cover
        # 波段離場 (空單)
        elif ext_name == "波段追蹤：關閉固定停利 + 站上MA20上彎離場 (波段操作)": pass
        elif ext_name == "波段追蹤：ATR追蹤回補 (最低價拉回3*ATR離場)": pass
        elif ext_name == "波段追蹤：突破前10日最高點回補 (波段操作)": pass

    # 執行回測狀態機以取得精確的進出場訊號
    ent_sig = np.zeros(n_rows, dtype=bool)
    ext_sig = np.zeros(n_rows, dtype=bool)
    
    hold = False
    buy_price = 0
    highest_close_since_entry = 0
    lowest_close_since_entry = 99999999

    for i in range(1, n_rows - 1):
        if not hold:
            if ent_sig_raw[i]:
                hold = True
                buy_price = opens[i+1]
                highest_close_since_entry = closes[i+1]
                lowest_close_since_entry = closes[i+1]
                ent_sig[i] = True
        else:
            next_open = opens[i+1]
            trade_roi = (next_open - buy_price) / buy_price if buy_price > 0 else 0
            if is_short:
                trade_roi = -trade_roi
                lowest_close_since_entry = min(lowest_close_since_entry, closes[i])
            else:
                highest_close_since_entry = max(highest_close_since_entry, closes[i])
                
            cond_tp = (tp is not None) and (trade_roi >= tp)
            cond_sl = (sl is not None) and (trade_roi <= -sl)
            cond_signal = ext_sig_raw[i]

            # 動態離場條件判定
            if not is_short:
                if ext_name == "波段追蹤：關閉固定停利 + 跌破MA20下彎離場 (波段操作)":
                    cond_signal = (closes[i] < ma20[i]) and (ma20[i] < ma20[i-1] if i > 0 else True)
                    cond_sl = trade_roi <= -0.10
                elif ext_name == "波段追蹤：ATR追蹤停損 (最高價拉回3*ATR離場)":
                    cond_signal = closes[i] < (highest_close_since_entry - 3 * atr[i])
                    cond_sl = trade_roi <= -0.10
                elif ext_name == "波段追蹤：跌破前10日最低點停損 (波段操作)":
                    recent_low = min(lows[max(0, i-10):i]) if i > 0 else lows[0]
                    cond_signal = closes[i] < recent_low
                    cond_sl = trade_roi <= -0.10
            else:
                if ext_name == "波段追蹤：關閉固定停利 + 站上MA20上彎離場 (波段操作)":
                    cond_signal = (closes[i] > ma20[i]) and (ma20[i] > ma20[i-1] if i > 0 else True)
                    cond_sl = trade_roi <= -0.10
                elif ext_name == "波段追蹤：ATR追蹤回補 (最低價拉回3*ATR離場)":
                    cond_signal = closes[i] > (lowest_close_since_entry + 3 * atr[i])
                    cond_sl = trade_roi <= -0.10
                elif ext_name == "波段追蹤：突破前10日最高點回補 (波段操作)":
                    recent_high = max(highs[max(0, i-10):i]) if i > 0 else highs[0]
                    cond_signal = closes[i] > recent_high
                    cond_sl = trade_roi <= -0.10

            if cond_tp or cond_sl or cond_signal:
                hold = False
                ext_sig[i] = True
                buy_price = 0

    return ent_sig, ext_sig, tp, sl

def main():
    print("=" * 60)
    print("  V8.2 高勝率回測 Excel 匯出工具啟動...")
    print("=" * 60)

    if not os.path.exists(LEADERBOARD_PATH):
        print(f"[錯誤] 排行榜檔案 {LEADERBOARD_PATH} 不存在，請先執行回測。")
        return

    with open(LEADERBOARD_PATH, 'r', encoding='utf-8') as f:
        leaderboard = json.load(f)

    # 篩選勝率 >= 75% 的股票
    high_win = [x for x in leaderboard if x['winRate'] >= 75]
    print(f"[資訊] 共有 {len(high_win)} 檔股票勝率 >= 75%")

    # 尋找所有可用個股資料
    candidate_files = []
    for folder in [AUTO_EXPORT_DIR, STOCK_ORIGINAL_DIR]:
        if os.path.exists(folder):
            candidate_files.extend(glob.glob(os.path.join(folder, "*.csv")))
            candidate_files.extend(glob.glob(os.path.join(folder, "*.xlsx")))

    # 按代號與日期排序尋找最新檔案 (自適應 EOM 與 20260701 的排序邏輯)
    file_map = {}
    for f in candidate_files:
        filename = os.path.basename(f)
        parts = filename.split('_')
        if len(parts) >= 2 and parts[0].isdigit():
            code = parts[0]
            if "AI" in filename or "V4" in filename or "V_Rebound" in filename or "V_Dip" in filename:
                continue
            date_val = get_file_date(f)
            is_xlsx = 1 if f.endswith('.xlsx') else 0
            if code not in file_map or (date_val, is_xlsx) > file_map[code][1]:
                file_map[code] = (f, (date_val, is_xlsx))

    os.makedirs(SERVER_DATA_DIR, exist_ok=True)

    for item in high_win:
        code = item['code']
        name = item['name']
        strategy_type = item.get('type', 'long')
        strategy_full = item['strategy']
        win_rate = item['winRate']
        trades = item['trades']
        profit = item['profit']
        roi = item['roi']

        if code not in file_map:
            print(f"[警告] 找不到個股 {code} {name} 的資料檔案，跳過 Excel 匯出")
            continue

        file_path = file_map[code][0]
        filename = os.path.basename(file_path)

        # 讀取資料
        try:
            if file_path.endswith('.csv'):
                try:
                    df = pd.read_csv(file_path, encoding='cp950', header=0)
                except Exception:
                    df = pd.read_csv(file_path, encoding='utf-8', header=0)
            else:
                df = pd.read_excel(file_path, header=0)
        except Exception as ex:
            print(f"[警告] 無法讀取檔案 {filename}: {ex}")
            continue

        df.columns = [str(c).strip() for c in df.columns]
        df = df.tail(900).reset_index(drop=True)

        parts_strat = strategy_full.split(' & ')
        if len(parts_strat) != 2:
            print(f"[警告] 策略格式異常: {strategy_full}，跳過")
            continue
        ent_name = parts_strat[0].strip()
        ext_name = parts_strat[1].strip()

        is_short = (strategy_type == "short")
        # 計算進出場訊號
        ent_sig, ext_sig, tp, sl = get_strategy_signals(df, ent_name, ext_name, is_short=is_short)

        # 準備 Excel
        wb = Workbook()

        # Sheet 1: 參數與績效總覽
        ws_summary = wb.active
        ws_summary.title = "回測績效總覽"

        # 設計標題樣式與填充
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        title_font = Font(name="Noto Sans TC", size=11, bold=True, color="FFFFFF")
        normal_font = Font(name="Noto Sans TC", size=10)
        bold_font = Font(name="Noto Sans TC", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        ws_summary['A1'] = "高勝率量化分析回測參數與績效總覽"
        ws_summary['A1'].font = Font(name="Noto Sans TC", size=14, bold=True, color="1F497D")
        ws_summary.merge_cells('A1:E1')

        headers = ["屬性項目", "設定數值 / 分析指標", "說明", "交易公式變數", "參考儲存格"]
        for idx, h in enumerate(headers, 1):
            cell = ws_summary.cell(row=3, column=idx, value=h)
            cell.fill = header_fill
            cell.font = title_font
            cell.alignment = center_align

        # 寫入靜態參數
        pos_type_text = "空單" if is_short else "多單"
        params_rows = [
            ["個股代碼", code, "回測分析目標股號", "code", f"='{ws_summary.title}'!B4"],
            ["個股名稱", name, "回測分析目標股票名稱", "name", f"='{ws_summary.title}'!B5"],
            ["單筆下單股數", 1000, "每次交易之固定股數", "shares", f"='{ws_summary.title}'!B6"],
            ["策略類型", pos_type_text, "交易方向 (多單/空單)", "type", f"='{ws_summary.title}'!B7"],
            ["進場策略名稱", ent_name, "選定的進場訊號公式", "entry_strategy", f"='{ws_summary.title}'!B8"],
            ["出場策略名稱", ext_name, "選定的出場條件與停損利條件", "exit_strategy", f"='{ws_summary.title}'!B9"],
            ["強制停利點 (%)", tp if tp is not None else "無", "滿足百分比時即刻出場", "take_profit", f"='{ws_summary.title}'!B10"],
            ["強制停損點 (%)", sl if sl is not None else "無", "滿足百分比時即刻止損", "stop_loss", f"='{ws_summary.title}'!B11"],
            ["勝率 (Win Rate)", f"{win_rate:.1f}%", "高於零損益的交易比例", "win_rate", f"='{ws_summary.title}'!B12"],
            ["總交易次數", trades, "歷史回測總交易次數", "trades", f"='{ws_summary.title}'!B13"],
            ["整體 ROI (%)", f"{roi:.1f}%", "總損益 / 單筆最大交易金額", "roi", f"='{ws_summary.title}'!B14"],
            ["累積淨利 (TWD)", profit, "歷史累計淨損益金額", "net_profit", f"='{ws_summary.title}'!B15"]
        ]

        for r_idx, row in enumerate(params_rows, 4):
            for c_idx, val in enumerate(row, 1):
                cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
                cell.font = normal_font
                if c_idx in [1, 3]:
                    cell.alignment = left_align
                elif c_idx == 2:
                    cell.font = bold_font
                    cell.alignment = left_align
                else:
                    cell.alignment = left_align

        # Sheet 2: 詳細交易數據與公式
        ws_data = wb.create_sheet(title="每日收盤與回測公式")

        # 寫入 DataFrame 基本資訊
        base_cols = ["日期", "收盤價", "開盤價", "最高價", "最低價"]
        df_cols_present = [find_column(df, [bc]) for bc in base_cols]
        df_cols_present = [c for c in df_cols_present if c is not None]

        # 如果找不到足夠欄位，使用前 5 欄
        if len(df_cols_present) < 2:
            df_cols_present = list(df.columns[:5])

        # 寫入欄位標頭
        headers_data = df_cols_present + ["進場訊號 (Entry)", "出場訊號 (Exit)", "動作訊號 (Action)", "交易價格 (Price)", "單筆損益 (TWD)", "累積損益 (TWD)"]
        for idx, h in enumerate(headers_data, 1):
            cell = ws_data.cell(row=8, column=idx, value=h)
            cell.fill = header_fill
            cell.font = title_font
            cell.alignment = center_align

        # 寫入每日收盤明細
        for r_idx, row in df[df_cols_present].iterrows():
            excel_row = r_idx + 9
            for c_idx, col_name in enumerate(df_cols_present, 1):
                ws_data.cell(row=excel_row, column=c_idx, value=row[col_name]).font = normal_font

        # 寫入動態公式 (Excel Column Letters)
        col_R = get_column_letter(len(df_cols_present) + 1) # Entry
        col_S = get_column_letter(len(df_cols_present) + 2) # Exit
        col_T = get_column_letter(len(df_cols_present) + 3) # Action
        col_U = get_column_letter(len(df_cols_present) + 4) # Price
        col_V = get_column_letter(len(df_cols_present) + 5) # Profit
        col_W = get_column_letter(len(df_cols_present) + 6) # CumProfit

        # 寫入全域變數引用 (放第一列/第二列作為參數)
        ws_data['A1'] = "個股交易回測明細"
        ws_data['A1'].font = Font(name="Noto Sans TC", size=14, bold=True, color="1F497D")

        ws_data['A3'] = "策略類型"
        ws_data['B3'] = pos_type_text
        ws_data['A4'] = "進場策略"
        ws_data['B4'] = ent_name
        ws_data['A5'] = "出場策略"
        ws_data['B5'] = ext_name

        ws_data['A6'] = "下單股數"
        ws_data['B6'] = 1000

        for cell_ref in ['A3', 'A4', 'A5', 'A6']:
            ws_data[cell_ref].font = bold_font
        for cell_ref in ['B3', 'B4', 'B5', 'B6']:
            ws_data[cell_ref].font = normal_font

        # 找出「開盤價」在 df_cols_present 中的 index
        c_open_col_idx = 3
        try:
            c_open_actual = find_column(df, ["開盤價", "開盤", "Open", "open"])
            if c_open_actual in df_cols_present:
                c_open_col_idx = df_cols_present.index(c_open_actual) + 1
        except Exception:
            pass
        col_C_open = get_column_letter(c_open_col_idx)

        # 寫入公式
        n_data = len(df)
        for idx in range(n_data):
            r = idx + 9
            # 寫入 Boolean 進出場訊號值
            ws_data[f'{col_R}{r}'] = bool(ent_sig[idx])
            ws_data[f'{col_S}{r}'] = bool(ext_sig[idx])

            if r == 9:
                # 第一行初始值 (無前一日狀態)
                ws_data[f'{col_T}{r}'] = ""
                ws_data[f'{col_U}{r}'] = 0.0
                ws_data[f'{col_V}{r}'] = 0.0
                ws_data[f'{col_W}{r}'] = 0.0
            else:
                pr = r - 1
                pW = f'{col_W}{pr}'
                
                if not is_short:
                    # 多單動作(T)：前一日 Entry 為 TRUE 則今日買進；前一日 Exit 為 TRUE 則今日賣出
                    ws_data[f'{col_T}{r}'] = f'=IF({col_R}{pr}=TRUE, "買進", IF({col_S}{pr}=TRUE, "賣出", ""))'
                    # 交易價格(U)：買進或賣出時為當日開盤價，否則若前一日非賣出且有持倉價，則延續前一日交易價格
                    ws_data[f'{col_U}{r}'] = f'=IF({col_T}{r}="買進", {col_C_open}{r}, IF({col_T}{r}="賣出", {col_C_open}{r}, IF({col_T}{pr}="賣出", 0, {col_U}{pr})))'
                    # 多單單筆損益(V): (賣出價格 - 買進成本) * 股數
                    ws_data[f'{col_V}{r}'] = f'=IF({col_T}{r}="賣出", ({col_U}{r} - {col_U}{pr}) * $B$6, 0)'
                else:
                    # 空單動作(T)：前一日 Entry 為 TRUE 則今日放空；前一日 Exit 為 TRUE 則今日回補
                    ws_data[f'{col_T}{r}'] = f'=IF({col_R}{pr}=TRUE, "放空", IF({col_S}{pr}=TRUE, "回補", ""))'
                    # 交易價格(U)：放空或回補時為當日開盤價，否則若前一日非回補且有持倉價，則延續前一日交易價格
                    ws_data[f'{col_U}{r}'] = f'=IF({col_T}{r}="放空", {col_C_open}{r}, IF({col_T}{r}="回補", {col_C_open}{r}, IF({col_T}{pr}="回補", 0, {col_U}{pr})))'
                    # 空單單筆損益(V): (放空成本 - 回補價格) * 股數
                    ws_data[f'{col_V}{r}'] = f'=IF({col_T}{r}="回補", ({col_U}{pr} - {col_U}{r}) * $B$6, 0)'

                # 累積損益(W)
                ws_data[f'{col_W}{r}'] = f'={pW} + {col_V}{r}'

            # 格式套用
            for col_letter in [col_R, col_S, col_T, col_U, col_V, col_W]:
                ws_data[f'{col_letter}{r}'].font = normal_font

        # 自動調整欄寬
        for ws in [ws_summary, ws_data]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # 儲存 Excel 到 2 個位置 (一個給使用者，一個給網頁伺服器)
        excel_name = f"{code}_AI_v8_2.xlsx"
        user_out_path = os.path.join(SCRIPT_DIR, excel_name)
        server_out_path = os.path.join(SERVER_DATA_DIR, excel_name)
        
        wb.save(user_out_path)
        shutil.copy2(user_out_path, server_out_path)
        print(f"  [成功] 產出並複製：{excel_name} (型態: {pos_type_text})")

    print("\n[完成] 所有勝率 75% 以上個股的 V8.2 回測 EXCEL 檔案已成功產出與複製！")

if __name__ == "__main__":
    main()
