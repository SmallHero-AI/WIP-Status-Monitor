# -*- coding: utf-8 -*-
import os
import glob
import json
import shutil
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── 路徑設定 ──
SCRIPT_DIR = r"E:\G-AI-1\Stock analysis"
STOCK_ORIGINAL_DIR = os.path.join(SCRIPT_DIR, "Stock original")
AUTO_EXPORT_DIR = os.path.join(STOCK_ORIGINAL_DIR, "auto_export")
LEADERBOARD_PATH = os.path.join(SCRIPT_DIR, "修正版_V6_Server", "public", "leaderboard_v8.json")
SERVER_DATA_DIR = os.path.join(SCRIPT_DIR, "修正版_V6_Server", "data")

def find_column(df, keywords):
    for col in df.columns:
        for kw in keywords:
            if kw.lower() in str(col).lower():
                return col
    return None

def clean_and_parse_series(df, col_name):
    if col_name is not None and col_name in df.columns:
        s = df[col_name].astype(str).str.replace(r'[▲▼↑↓△▽,%\s]', '', regex=True)
        return pd.to_numeric(s, errors='coerce').ffill().bfill().fillna(0).values
    return np.zeros(len(df))

def get_strategy_signals(df, ent_name, ext_name, is_short=False):
    """
    計算對應的進出場訊號 Boolean 陣列
    """
    n_rows = len(df)
    
    # 匹配技術指標欄位
    c_open = find_column(df, ['開盤價', '開盤', 'Open', 'open'])
    c_high = find_column(df, ['最高價', '最高', 'High', 'high'])
    c_low = find_column(df, ['最低價', '最低', 'Low', 'low'])
    c_close = find_column(df, ['收盤價', '收盤', 'Close', 'close'])
    c_ma5 = find_column(df, ['均價[5]', 'MA5', 'ma5'])
    c_ma20 = find_column(df, ['均價[20]', 'MA20', 'ma20'])
    c_ma60 = find_column(df, ['均價[60]', 'MA60', 'ma60'])
    c_ma120 = find_column(df, ['均價[120]', 'MA120', 'ma120'])
    c_eom = find_column(df, ['EOM[60]', 'EOM'])
    c_eom_sig = find_column(df, ['Signal[20]', 'EOM_Signal'])
    c_mfi = find_column(df, ['MFI[14]', 'MFI'])
    c_macd = find_column(df, ['MACD', 'MACD柱', 'MACD 柱', 'MACD 柱狀體'])
    c_k = find_column(df, ['%KS', 'K(9,3,3)', 'K值'])
    c_d = find_column(df, ['%DS', 'D(9,3,3)', 'D值'])
    c_rsi = find_column(df, ['RSI[14]', 'RSI'])
    c_bias20 = find_column(df, ['BIAS[20]', 'BIAS20'])
    c_cci = find_column(df, ['CCI[20]', 'CCI'])
    c_bb_upper = find_column(df, ['布林加通道 上軌', 'BB_Upper'])
    c_bb_lower = find_column(df, ['布林加通道 下軌', 'BB_Lower'])
    c_pivot = find_column(df, ['Pivot', 'pivot'])
    c_r1 = find_column(df, ['第1道壓力', 'R1'])
    c_s1 = find_column(df, ['第1道支撐', 'S1'])

    # 轉換陣列
    opens = clean_and_parse_series(df, c_open)
    highs = clean_and_parse_series(df, c_high)
    lows = clean_and_parse_series(df, c_low)
    closes = clean_and_parse_series(df, c_close)
    ma5 = clean_and_parse_series(df, c_ma5)
    ma20 = clean_and_parse_series(df, c_ma20)
    ma60 = clean_and_parse_series(df, c_ma60)
    ma120 = clean_and_parse_series(df, c_ma120)
    eom = clean_and_parse_series(df, c_eom)
    eom_sig = clean_and_parse_series(df, c_eom_sig)
    mfi = clean_and_parse_series(df, c_mfi)
    macd = clean_and_parse_series(df, c_macd)
    k = clean_and_parse_series(df, c_k)
    d = clean_and_parse_series(df, c_d)
    rsi = clean_and_parse_series(df, c_rsi)
    bias20 = clean_and_parse_series(df, c_bias20)
    cci = clean_and_parse_series(df, c_cci)
    bb_upper = clean_and_parse_series(df, c_bb_upper)
    bb_lower = clean_and_parse_series(df, c_bb_lower)
    pivot = clean_and_parse_series(df, c_pivot)
    r1 = clean_and_parse_series(df, c_r1)
    s1 = clean_and_parse_series(df, c_s1)

    prev_k = np.roll(k, 1); prev_k[0] = k[0]
    prev_d = np.roll(d, 1); prev_d[0] = d[0]
    prev_macd = np.roll(macd, 1); prev_macd[0] = macd[0]

    c_low_col = find_column(df, ['最低價', '最低', 'Low', 'low'])
    low_vals = df[c_low_col].values if c_low_col else lows
    c_high_col = find_column(df, ['最高價', '最高', 'High', 'high'])
    high_vals = df[c_high_col].values if c_high_col else highs

    if not is_short:
        # 多單買入/賣出訊號
        f_kd_gold = (k > d) & (prev_k <= prev_d)
        f_rsi_oversold = rsi < 35
        f_rsi_bull = rsi > 50
        f_macd_grow = macd > prev_macd
        f_macd_positive = macd > 0
        f_bias_oversold = bias20 < -4
        f_cci_oversold = cci < -100
        f_eom_bullish = eom > eom_sig
        f_mfi_oversold = mfi < 30
        f_bb_oversold = (closes < bb_lower) & (bb_lower > 0)
        f_above_pivot = closes > pivot
        f_break_r1 = (closes > r1) & (r1 > 0)
        f_ma_bull = (ma5 > ma20) & (ma20 > ma60)
        f_above_ma20 = closes > ma20
        f_above_ma60 = closes > ma60
        f_above_ma120 = closes > ma120

        f_kd_dead = (k < d) & (prev_k >= prev_d)
        f_rsi_overbought = rsi > 65
        f_macd_shrink = macd < prev_macd
        f_below_ma20 = closes < ma20
        f_bb_overbought = (closes > bb_upper) & (bb_upper > 0)
        f_below_s1 = (closes < s1) & (s1 > 0)

        # 匹配進場
        ent_sig = np.zeros(n_rows, dtype=bool)
        if ent_name == "強勢均線多頭+KD金叉": ent_sig = f_ma_bull & f_kd_gold & f_above_ma20
        elif ent_name == "中長期多頭+RSI低檔轉強": ent_sig = f_above_ma60 & f_above_ma120 & (rsi < 45) & f_macd_grow
        elif ent_name == "雙重均線突破+MACD翻紅": ent_sig = f_above_ma20 & f_above_ma60 & f_macd_positive & f_macd_grow
        elif ent_name == "EOM動能突破+均線多頭": ent_sig = f_eom_bullish & f_ma_bull & f_above_ma20
        elif ent_name == "極度超跌共振 (BIAS+RSI+MFI)": ent_sig = f_bias_oversold & f_rsi_oversold & f_mfi_oversold
        elif ent_name == "布林通道下軌+KD低檔金叉": ent_sig = f_bb_oversold & f_kd_gold
        elif ent_name == "CCI超賣+MFI超賣+KD金叉": ent_sig = f_cci_oversold & f_mfi_oversold & f_kd_gold
        elif ent_name == "雙重超賣 (RSI+CCI)+MACD轉強": ent_sig = f_rsi_oversold & f_cci_oversold & f_macd_grow
        elif ent_name == "樞軸點(S1)支撐+KD金叉": ent_sig = (closes > s1) & (low_vals <= s1) & f_kd_gold & (s1 > 0)
        elif ent_name == "突破阻力(R1)+MACD多頭": ent_sig = f_break_r1 & f_macd_positive & f_above_ma20
        elif ent_name == "突破樞軸(Pivot)+RSI轉強": ent_sig = f_above_pivot & f_rsi_bull & f_macd_grow
        elif ent_name == "S1支撐不破+MACD紅柱增長": ent_sig = (closes > s1) & f_macd_grow & (s1 > 0)
        elif ent_name == "多頭拉回：MA60之上+KD金叉+RSI<45": ent_sig = f_above_ma60 & f_kd_gold & (rsi < 45)
        elif ent_name == "布林中軌支撐+MACD紅柱": ent_sig = (closes > bb_lower) & (closes < bb_upper) & f_above_ma20 & f_macd_grow
        elif ent_name == "動能共振：EOM突破+MACD>0+RSI>50": ent_sig = f_eom_bullish & f_macd_positive & f_rsi_bull
        elif ent_name == "雙保險超跌：BIAS超賣+布林下軌觸及": ent_sig = f_bias_oversold & f_bb_oversold
        elif ent_name == "主力資金流入：MFI超賣+EOM突破": ent_sig = f_mfi_oversold & f_eom_bullish
        elif ent_name == "壓力突破：價格>R1+EOM強勢+MACD>0": ent_sig = f_break_r1 & f_eom_bullish & f_macd_positive
        elif ent_name == "長期均線支撐：站上MA120+KD金叉+CCI超賣": ent_sig = f_above_ma120 & f_kd_gold & f_cci_oversold
        elif ent_name == "CCI超賣反彈+MACD紅柱增長": ent_sig = f_cci_oversold & f_macd_grow & f_above_ma20

        # 匹配出場
        tp, sl = None, None
        ext_sig = np.zeros(n_rows, dtype=bool)
        if ext_name == "6%停利 / 6%停損 (穩健勝率)": tp, sl = 0.06, 0.06
        elif ext_name == "8%停利 / 8%停損 (均衡配置)": tp, sl = 0.08, 0.08
        elif ext_name == "12%停利 / 6%停損 (高盈虧比)": tp, sl = 0.12, 0.06
        elif ext_name == "KD死叉離場或6%停損": sl, ext_sig = 0.06, f_kd_dead
        elif ext_name == "RSI超買離場或6%停損": sl, ext_sig = 0.06, f_rsi_overbought
        elif ext_name == "MACD紅柱縮短離場或6%停損": sl, ext_sig = 0.06, f_macd_shrink
        elif ext_name == "收盤跌破MA20或6%停損": sl, ext_sig = 0.06, f_below_ma20
        elif ext_name == "觸碰布林上軌停利或6%停損": sl, ext_sig = 0.06, f_bb_overbought
        elif ext_name == "跌破樞軸支撐(S1)或8%停利": tp, ext_sig = 0.08, f_below_s1

        return ent_sig, ext_sig, tp, sl
    else:
        # 空單放空/回補訊號
        f_ma_bear = (ma5 < ma20) & (ma20 < ma60)
        f_below_pivot = closes < pivot
        f_break_s1 = (closes < s1) & (s1 > 0)
        f_bias_overbought = bias20 > 4
        f_cci_overbought = cci > 100
        f_eom_bearish = eom < eom_sig
        f_mfi_overbought = mfi > 70
        f_bb_overbought_short = (closes > bb_upper) & (bb_upper > 0)
        
        # 新增遺漏的空單技術指標定義
        f_kd_dead = (k < d) & (prev_k >= prev_d)
        f_rsi_overbought = rsi > 65
        f_macd_shrink = macd < prev_macd
        f_below_ma20 = closes < ma20
        f_below_ma60 = closes < ma60
        f_below_ma120 = closes < ma120

        f_kd_gold_cover = (k > d) & (prev_k <= prev_d)
        f_rsi_oversold_cover = rsi < 35
        f_macd_grow_cover = macd > prev_macd
        f_above_ma20_cover = closes > ma20
        f_bb_oversold_cover = (closes < bb_lower) & (bb_lower > 0)
        f_above_r1_cover = (closes > r1) & (r1 > 0)

        # 匹配進場 (空單)
        ent_sig = np.zeros(n_rows, dtype=bool)
        if ent_name == "弱勢均線空頭+KD死叉": ent_sig = f_ma_bear & f_kd_dead & f_below_ma20
        elif ent_name == "中長期空頭+RSI高檔轉弱": ent_sig = f_below_ma60 & f_below_ma120 & (rsi > 55) & f_macd_shrink
        elif ent_name == "雙重均線跌破+MACD翻綠": ent_sig = f_below_ma20 & f_below_ma60 & (macd < 0) & f_macd_shrink
        elif ent_name == "EOM動能下跌+均線空頭": ent_sig = f_eom_bearish & f_ma_bear & f_below_ma20
        elif ent_name == "極度超買共振 (BIAS+RSI+MFI)": ent_sig = f_bias_overbought & f_rsi_overbought & f_mfi_overbought
        elif ent_name == "布林通道上軌+KD高檔死叉": ent_sig = f_bb_overbought_short & f_kd_dead
        elif ent_name == "CCI超買+MFI超買+KD死叉": ent_sig = f_cci_overbought & f_mfi_overbought & f_kd_dead
        elif ent_name == "雙重超買 (RSI+CCI)+MACD轉弱": ent_sig = f_rsi_overbought & f_cci_overbought & f_macd_shrink
        elif ent_name == "樞軸點(R1)壓力+KD死叉": ent_sig = (closes < r1) & (high_vals >= r1) & f_kd_dead & (r1 > 0)
        elif ent_name == "跌破支撐(S1)+MACD空頭": ent_sig = f_break_s1 & (macd < 0) & f_below_ma20
        elif ent_name == "跌破樞軸(Pivot)+RSI轉弱": ent_sig = f_below_pivot & (rsi < 50) & f_macd_shrink
        elif ent_name == "R1壓力不破+MACD綠柱增長": ent_sig = (closes < r1) & f_macd_shrink & (r1 > 0)
        elif ent_name == "空頭拉回：MA60之下+KD死叉+RSI>55": ent_sig = f_below_ma60 & f_kd_dead & (rsi > 55)
        elif ent_name == "布林中軌阻力+MACD綠柱": ent_sig = (closes > bb_lower) & (closes < bb_upper) & f_below_ma20 & f_macd_shrink
        elif ent_name == "動能共振：EOM下跌+MACD<0+RSI<50": ent_sig = f_eom_bearish & (macd < 0) & (rsi < 50)
        elif ent_name == "雙保險超買：BIAS超買+布林上軌觸及": ent_sig = f_bias_overbought & f_bb_overbought_short
        elif ent_name == "主力資金流出：MFI超買+EOM下跌": ent_sig = f_mfi_overbought & f_eom_bearish
        elif ent_name == "支撐跌破：價格<S1+EOM弱勢+MACD<0": ent_sig = f_break_s1 & f_eom_bearish & (macd < 0)
        elif ent_name == "長期均線阻力：跌破MA120+KD死叉+CCI超買": ent_sig = (closes < ma120) & f_kd_dead & f_cci_overbought
        elif ent_name == "CCI超買回檔+MACD綠柱增長": ent_sig = f_cci_overbought & f_macd_shrink & f_below_ma20

        # 匹配出場 (空單)
        tp, sl = None, None
        ext_sig = np.zeros(n_rows, dtype=bool)
        if ext_name == "6%停利 / 6%停損 (穩健勝率)": tp, sl = 0.06, 0.06
        elif ext_name == "8%停利 / 8%停損 (均衡配置)": tp, sl = 0.08, 0.08
        elif ext_name == "12%停利 / 6%停損 (高盈虧比)": tp, sl = 0.12, 0.06
        elif ext_name == "KD金叉離場或6%停損": sl, ext_sig = 0.06, f_kd_gold_cover
        elif ext_name == "RSI超賣離場或6%停損": sl, ext_sig = 0.06, f_rsi_oversold_cover
        elif ext_name == "MACD綠柱縮短離場或6%停損": sl, ext_sig = 0.06, f_macd_grow_cover
        elif ext_name == "收盤站上MA20或6%停損": sl, ext_sig = 0.06, f_above_ma20_cover
        elif ext_name == "觸碰布林下軌停利或6%停損": sl, ext_sig = 0.06, f_bb_oversold_cover
        elif ext_name == "突破壓力(R1)或8%停利": tp, ext_sig = 0.08, f_above_r1_cover

        return ent_sig, ext_sig, tp, sl

def main():
    print("=" * 60)
    print("  V8 高勝率回測 Excel 匯出工具啟動...")
    print("=" * 60)

    if not os.path.exists(LEADERBOARD_PATH):
        print(f"[錯誤] 排行榜檔案 {LEADERBOARD_PATH} 不存在，請先執行回測。")
        return

    with open(LEADERBOARD_PATH, 'r', encoding='utf-8') as f:
        leaderboard = json.load(f)

    # 篩選勝率 >= 75% 的股票
    high_win = [x for x in leaderboard if x['winRate'] >= 75]
    print(f"[資訊] 共有 {len(high_win)} 檔股票勝率 >= 75%")

    # 尋找所有可用個股資料
    candidate_files = []
    for folder in [AUTO_EXPORT_DIR, STOCK_ORIGINAL_DIR]:
        if os.path.exists(folder):
            candidate_files.extend(glob.glob(os.path.join(folder, "*.csv")))
            candidate_files.extend(glob.glob(os.path.join(folder, "*.xlsx")))

    # 按代號映射最新檔案
    file_map = {}
    for f in candidate_files:
        filename = os.path.basename(f)
        parts = filename.split('_')
        if len(parts) >= 2 and parts[0].isdigit():
            code = parts[0]
            date_str = parts[-1].split('.')[0] if len(parts) >= 3 else ""
            if "AI" in filename or "V4" in filename or "V_Rebound" in filename or "V_Dip" in filename:
                continue
            if code not in file_map or date_str > file_map[code][1]:
                file_map[code] = (f, date_str)

    os.makedirs(SERVER_DATA_DIR, exist_ok=True)

    for item in high_win:
        code = item['code']
        name = item['name']
        strategy_type = item.get('type', 'long')
        strategy_full = item['strategy']
        win_rate = item['winRate']
        trades = item['trades']
        profit = item['profit']
        roi = item['roi']

        if code not in file_map:
            print(f"[警告] 找不到個股 {code} {name} 的資料檔案，跳過 Excel 匯出")
            continue

        file_path = file_map[code][0]
        filename = os.path.basename(file_path)

        # 讀取資料
        try:
            if file_path.endswith('.csv'):
                try:
                    df = pd.read_csv(file_path, encoding='cp950', header=0)
                except Exception:
                    df = pd.read_csv(file_path, encoding='utf-8', header=0)
            else:
                df = pd.read_excel(file_path, header=0)
        except Exception as ex:
            print(f"[警告] 無法讀取檔案 {filename}: {ex}")
            continue

        df.columns = [str(c).strip() for c in df.columns]
        df = df.tail(900).reset_index(drop=True)

        parts_strat = strategy_full.split(' & ')
        if len(parts_strat) != 2:
            print(f"[警告] 策略格式異常: {strategy_full}，跳過")
            continue
        ent_name = parts_strat[0].strip()
        ext_name = parts_strat[1].strip()

        is_short = (strategy_type == "short")
        # 計算進出場訊號
        ent_sig, ext_sig, tp, sl = get_strategy_signals(df, ent_name, ext_name, is_short=is_short)

        # 準備 Excel
        wb = Workbook()

        # Sheet 1: 參數與績效總覽
        ws_summary = wb.active
        ws_summary.title = "回測績效總覽"

        # 設計標題樣式與填充
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        title_font = Font(name="Noto Sans TC", size=11, bold=True, color="FFFFFF")
        normal_font = Font(name="Noto Sans TC", size=10)
        bold_font = Font(name="Noto Sans TC", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        ws_summary['A1'] = "高勝率量化分析回測參數與績效總覽"
        ws_summary['A1'].font = Font(name="Noto Sans TC", size=14, bold=True, color="1F497D")
        ws_summary.merge_cells('A1:E1')

        headers = ["屬性項目", "設定數值 / 分析指標", "說明", "交易公式變數", "參考儲存格"]
        for idx, h in enumerate(headers, 1):
            cell = ws_summary.cell(row=3, column=idx, value=h)
            cell.fill = header_fill
            cell.font = title_font
            cell.alignment = center_align

        # 寫入靜態參數
        pos_type_text = "空單" if is_short else "多單"
        params_rows = [
            ["個股代碼", code, "回測分析目標股號", "code", f"='{ws_summary.title}'!B4"],
            ["個股名稱", name, "回測分析目標股票名稱", "name", f"='{ws_summary.title}'!B5"],
            ["單筆下單股數", 1000, "每次交易之固定股數", "shares", f"='{ws_summary.title}'!B6"],
            ["策略類型", pos_type_text, "交易方向 (多單/空單)", "type", f"='{ws_summary.title}'!B7"],
            ["進場策略名稱", ent_name, "選定的進場訊號公式", "entry_strategy", f"='{ws_summary.title}'!B8"],
            ["出場策略名稱", ext_name, "選定的出場條件與停損利條件", "exit_strategy", f"='{ws_summary.title}'!B9"],
            ["強制停利點 (%)", tp if tp is not None else "無", "滿足百分比時即刻出場", "take_profit", f"='{ws_summary.title}'!B10"],
            ["強制停損點 (%)", sl if sl is not None else "無", "滿足百分比時即刻止損", "stop_loss", f"='{ws_summary.title}'!B11"],
            ["勝率 (Win Rate)", f"{win_rate:.1f}%", "高於零損益的交易比例", "win_rate", f"='{ws_summary.title}'!B12"],
            ["總交易次數", trades, "歷史回測總交易次數", "trades", f"='{ws_summary.title}'!B13"],
            ["整體 ROI (%)", f"{roi:.1f}%", "總損益 / 單筆最大交易金額", "roi", f"='{ws_summary.title}'!B14"],
            ["累積淨利 (TWD)", profit, "歷史累計淨損益金額", "net_profit", f"='{ws_summary.title}'!B15"]
        ]

        for r_idx, row in enumerate(params_rows, 4):
            for c_idx, val in enumerate(row, 1):
                cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
                cell.font = normal_font
                if c_idx == 1:
                    cell.font = bold_font

        # Sheet 2: 行情與模擬數據
        ws_data = wb.create_sheet(title="行情歷史與回測明細")

        # 寫入基本參數指標 (前幾列，可作為交易參數引用)
        ws_data['A1'] = "股票回測歷史數據明細表"
        ws_data['A1'].font = Font(name="Noto Sans TC", size=12, bold=True, color="1F497D")

        # B6 為股數
        ws_data['A6'] = "單筆股數:"
        ws_data['A6'].font = bold_font
        ws_data['B6'] = 1000
        ws_data['B6'].font = normal_font

        # 數據表欄位標頭
        num_cols = len(df.columns)
        orig_cols = list(df.columns)
        
        # 額外新增計算公式欄位 (R~W 欄)
        formula_headers = ["進場訊號", "出場訊號", "持倉狀態", "進場成本", "動作", "交易價格", "單筆損益", "累積損益"]
        
        # 合併欄位標頭
        all_headers = orig_cols + formula_headers
        for idx, h in enumerate(all_headers, 1):
            cell = ws_data.cell(row=8, column=idx, value=h)
            cell.fill = header_fill
            cell.font = title_font
            cell.alignment = center_align

        # 寫入 K 線行情數據
        for r_idx, row_values in enumerate(df.values, 9):
            for c_idx, val in enumerate(row_values, 1):
                cell = ws_data.cell(row=r_idx, column=c_idx, value=val)
                cell.font = normal_font
                # 第一欄日期格式整數
                if c_idx == 1:
                    try:
                        cell.value = int(float(val))
                    except:
                        pass
                    cell.alignment = center_align
                else:
                    try:
                        cell.value = float(val)
                    except:
                        pass

        # 取得新增公式的列字母
        col_P = get_column_letter(num_cols + 1) # 進場訊號
        col_Q = get_column_letter(num_cols + 2) # 出場訊號
        col_R = get_column_letter(num_cols + 3) # 持倉狀態
        col_S = get_column_letter(num_cols + 4) # 進場成本 (放空價)
        col_T = get_column_letter(num_cols + 5) # 動作
        col_U = get_column_letter(num_cols + 6) # 交易價格
        col_V = get_column_letter(num_cols + 7) # 單筆損益
        col_W = get_column_letter(num_cols + 8) # 累積損益

        # 寫入進出場訊號狀態與公式
        total_rows = ws_data.max_row
        for r in range(9, total_rows + 1):
            idx = r - 9
            # 寫入訊號 (由 Python 先行運算)
            ws_data[f'{col_P}{r}'] = bool(ent_sig[idx])
            ws_data[f'{col_Q}{r}'] = bool(ext_sig[idx])

            # 寫入交易模擬公式 (同前天開盤延遲一日買賣狀態機)
            pR = "FALSE" if r == 9 else f"{col_R}{r-1}"
            pP = "FALSE" if r == 9 else f"{col_P}{r-1}"
            pQ = "FALSE" if r == 9 else f"{col_Q}{r-1}"
            pS = "0" if r == 9 else f"{col_S}{r-1}"
            pW = "0" if r == 9 else f"{col_W}{r-1}"

            # 持倉狀態(R)
            ws_data[f'{col_R}{r}'] = f'=IF({pR}=FALSE, {pP}=TRUE, IF({pQ}=TRUE, FALSE, TRUE))'
            # 進場成本(S) (以開盤價 B 欄為交易基準)
            ws_data[f'{col_S}{r}'] = f'=IF(AND({pR}=FALSE, {col_R}{r}=TRUE), B{r}, IF(AND({pR}=TRUE, {col_R}{r}=TRUE), {pS}, 0))'
            
            if not is_short:
                # 多單動作(T)
                ws_data[f'{col_T}{r}'] = f'=IF(AND({pR}=FALSE, {col_R}{r}=TRUE), "買進", IF(AND({pR}=TRUE, {col_R}{r}=FALSE), "賣出", ""))'
                # 交易價格(U)
                ws_data[f'{col_U}{r}'] = f'=IF({col_T}{r}="買進", B{r}, IF({col_T}{r}="賣出", B{r}, 0))'
                # 多單單筆損益(V): (賣出價格 - 買進成本) * 股數
                ws_data[f'{col_V}{r}'] = f'=IF({col_T}{r}="賣出", ({col_U}{r} - {pS}) * $B$6, 0)'
            else:
                # 空單動作(T)
                ws_data[f'{col_T}{r}'] = f'=IF(AND({pR}=FALSE, {col_R}{r}=TRUE), "放空", IF(AND({pR}=TRUE, {col_R}{r}=FALSE), "回補", ""))'
                # 交易價格(U)
                ws_data[f'{col_U}{r}'] = f'=IF({col_T}{r}="放空", B{r}, IF({col_T}{r}="回補", B{r}, 0))'
                # 空單單筆損益(V): (放空成本 - 回補價格) * 股數
                ws_data[f'{col_V}{r}'] = f'=IF({col_T}{r}="回補", ({pS} - {col_U}{r}) * $B$6, 0)'

            # 累積損益(W)
            ws_data[f'{col_W}{r}'] = f'={pW} + {col_V}{r}'

            # 格式套用
            for col_letter in [col_R, col_S, col_T, col_U, col_V, col_W]:
                ws_data[f'{col_letter}{r}'].font = normal_font

        # 自動調整欄寬
        for ws in [ws_summary, ws_data]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # 儲存 Excel 到 2 個位置 (一個給使用者，一個給網頁伺服器)
        excel_name = f"{code}_{name}_AI_高勝率回測.xlsx"
        user_out_path = os.path.join(SCRIPT_DIR, excel_name)
        server_out_path = os.path.join(SERVER_DATA_DIR, excel_name)
        
        wb.save(user_out_path)
        shutil.copy2(user_out_path, server_out_path)
        print(f"  [成功] 產出並複製：{excel_name} (型態: {pos_type_text})")

    print("\n[完成] 所有勝率 75% 以上個股的 V8 回測 EXCEL 檔案已成功產出與複製！")

if __name__ == "__main__":
    main()
