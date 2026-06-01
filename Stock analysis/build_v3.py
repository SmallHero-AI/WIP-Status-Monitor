import os, shutil, pandas as pd, numpy as np, itertools
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font

src_html = r"E:\G-AI-1\Stock analysis\修正版_V2\技術分析測試_全個股_高勝率修正版.html"
out_dir = r"E:\G-AI-1\Stock analysis\修正版_V3"
input_dir = r"E:\G-AI-1\Stock analysis\Stock original"
os.makedirs(out_dir, exist_ok=True)

strategies = {
    '2330': {'name':'台積電','type':'V4'},
    '2360': {'name':'致茂','type':'V4'},
    '6205': {'name':'詮欣','type':'V4'},
    '6669': {'name':'緯穎','type':'V_Rebound'},
    '3443': {'name':'創意','type':'V_Rebound'},
    '3189': {'name':'景碩','type':'V_Rebound'},
    '3455': {'name':'由田','type':'V_Rebound'},
    '3535': {'name':'晶彩科','type':'V_Rebound'},
    '4908': {'name':'前鼎','type':'V_Rebound'},
    '6269': {'name':'台郡','type':'V_Rebound'},
    '4746': {'name':'台耀','type':'V_Dip_Buy'},
    '6261': {'name':'久元','type':'V_Dip_Buy'},
}

stock_data = {}
for file in os.listdir(input_dir):
    if file.endswith(".xlsx"):
        code = file.split('_')[0]
        if code in strategies:
            df = pd.read_excel(os.path.join(input_dir, file))
            df.columns = [str(c) for c in df.columns]
            for col in df.columns[1:]:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            stock_data[code] = df

def evaluate(df, params, strat):
    close, opn, high, low = df.iloc[:,4].values, df.iloc[:,1].values, df.iloc[:,2].values, df.iloc[:,3].values
    ma20, eom, eomsig, mfi, macd = df.iloc[:,7].values, df.iloc[:,10].values, df.iloc[:,11].values, df.iloc[:,12].values, df.iloc[:,14].values
    hold=False; bp=0; pnls=[]; entries=[]; mPrf=0
    for i in range(1, len(df)-1):
        if not hold:
            entry=False
            if strat=='V4':
                entry = close[i]>ma20[i] and eom[i]>(eomsig[i]+params.get('eom_off',0)) and macd[i]>params['macd'] and mfi[i]>params['mfi']
            elif strat=='V_Rebound':
                entry = close[i]<ma20[i] and mfi[i]<params['mfi_s'] and macd[i]>macd[i-1]
            elif strat=='V_Dip_Buy':
                entry = mfi[i]<params['mfi_s'] and eom[i]>eom[i-1] and macd[i]>0
            if entry:
                hold=True; bp=opn[i+1]; mPrf=0; entries.append(bp)
        else:
            pnlR=(opn[i+1]-bp)/bp
            if strat in ['V_Rebound','V_Dip_Buy']:
                if mfi[i]>params.get('mfi_b',80) or pnlR>=params['tp'] or pnlR<=-params['sl']:
                    hold=False; pnls.append((opn[i+1]-bp)*1000)
            else:
                if pnlR>=params['tp'] or pnlR<=-params['sl']:
                    hold=False; pnls.append((opn[i+1]-bp)*1000)
    if not pnls: return None
    cp=sum(pnls); mc=max(entries)*1000 if entries else 0
    wins=[p for p in pnls if p>0]
    return {**params, '累積淨損益總額':cp, '最大所需本金':mc,
            '真實投報率(ROI)':cp/mc if mc>0 else 0, '交易次數':len(pnls),
            '勝率':len(wins)/len(pnls), '單筆最大獲利':max(pnls), '單筆最大虧損':min(pnls),
            '平均每筆損益':cp/len(pnls),
            '期望報酬率(每筆)':(cp/len(pnls))/(mc/len(pnls)) if mc>0 else 0}

def gen_comment(rank, wr, roi):
    if rank==0: return "🏆 最高報酬：適合捕捉大趨勢波段。"
    if wr>=0.8: return "⭐ 極高勝率：穩定度極佳。"
    if roi>0.5: return "🚀 高資金效率：投報率優越。"
    if wr>=0.65: return "🛡️ 高勝率：穩定度佳。"
    return "💡 綜合優異：獲利與風險平衡。"

print("開始精細網格搜索...")
top3_db = {}

for code, info in strategies.items():
    if code not in stock_data: continue
    df = stock_data[code]; strat = info['type']
    results = []
    
    # 精細網格 (step 0.5% for tp/sl)
    if strat == 'V4':
        tp_g = np.arange(0.05, 0.25, 0.005)
        sl_g = np.arange(0.03, 0.12, 0.005)
        macd_g = np.arange(0, 6, 0.5)
        mfi_g = [25, 30, 35, 40]
        for tp,sl,md,mf in itertools.product(tp_g, sl_g, macd_g, mfi_g):
            r = evaluate(df, {'tp':tp,'sl':sl,'macd':md,'mfi':mf,'eom_off':0}, strat)
            if r: results.append(r)
    elif strat == 'V_Rebound':
        tp_g = np.arange(0.05, 0.25, 0.005)
        sl_g = np.arange(0.03, 0.12, 0.005)
        mfi_s_g = [20, 25, 30, 35]
        mfi_b_g = [70, 75, 80]
        for tp,sl,ms,mb in itertools.product(tp_g, sl_g, mfi_s_g, mfi_b_g):
            r = evaluate(df, {'tp':tp,'sl':sl,'mfi_s':ms,'mfi_b':mb}, strat)
            if r: results.append(r)
    elif strat == 'V_Dip_Buy':
        tp_g = np.arange(0.05, 0.20, 0.005)
        sl_g = np.arange(0.03, 0.10, 0.005)
        mfi_s_g = [25, 30, 35]
        for tp,sl,ms in itertools.product(tp_g, sl_g, mfi_s_g):
            r = evaluate(df, {'tp':tp,'sl':sl,'mfi_s':ms,'mfi_b':80}, strat)
            if r: results.append(r)
    
    if not results:
        print(f"[{code}] 無結果, 跳過"); continue
    
    res_df = pd.DataFrame(results).sort_values(by=['勝率','累積淨損益總額'], ascending=[False,False]).reset_index(drop=True)
    top3 = res_df.head(3)
    top3_db[code] = top3
    print(f"[{code}] 網格{len(results)}組, Top1 WR={top3.iloc[0]['勝率']*100:.1f}% PnL={top3.iloc[0]['累積淨損益總額']:,.0f}")
    
    # === Excel ===
    wb = Workbook()
    
    # Sheet 1: 參數最佳化結果 (all results sorted)
    ws_opt = wb.active; ws_opt.title = "參數最佳化結果"
    param_cols = [k for k in res_df.columns if k not in ['累積淨損益總額','最大所需本金','真實投報率(ROI)','交易次數','勝率','單筆最大獲利','單筆最大虧損','平均每筆損益','期望報酬率(每筆)']]
    stat_cols = ['累積淨損益總額','最大所需本金','真實投報率(ROI)','交易次數','勝率','單筆最大獲利','單筆最大虧損','平均每筆損益','期望報酬率(每筆)']
    all_cols = param_cols + stat_cols + ['🤖 AI 策略評估與註解']
    ws_opt.append(all_cols)
    
    # Write top results (max 30 rows for readability)
    top_n = res_df.head(30)
    for idx, row in top_n.iterrows():
        vals = [row.get(c, '') for c in param_cols + stat_cols]
        vals.append(gen_comment(idx, row['勝率'], row['真實投報率(ROI)']))
        ws_opt.append(vals)
        cr = ws_opt.max_row
        if idx < 3:  # Highlight top 3
            for col in range(1, len(all_cols)+1):
                ws_opt.cell(row=cr, column=col).fill = PatternFill("solid", fgColor="FEF08A")
    
    # Sheet 2-4: Top1/Top2/Top3 回測
    for ti in range(min(3, len(top3))):
        ws_data = wb.create_sheet(f"Top{ti+1}_回測")
        best_p = top3.iloc[ti]
        p_keys = param_cols
        
        ws_data['A1'] = f"【{strat} - Top{ti+1} 參數】"; ws_data['B1'] = "數值"
        for i, pk in enumerate(p_keys, 2):
            ws_data[f'A{i}'] = pk; ws_data[f'B{i}'] = best_p[pk]
        shares_row = len(p_keys) + 2
        ws_data[f'A{shares_row}'] = "交易張數 Shares"; ws_data[f'B{shares_row}'] = 1000
        
        headers = list(df.columns) + ["進場訊號(P)","出場訊號(Q)","持倉狀態(R)","進場成本(S)","動作(T)","交易價格(U)","單筆損益(V)","累積損益(W)"]
        for i, t in enumerate(headers, 1): ws_data.cell(row=8, column=i, value=t)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 9):
            for c_idx, v in enumerate(row, 1): ws_data.cell(row=r_idx, column=c_idx, value=v)
        
        for r in range(9, ws_data.max_row + 1):
            pR = "FALSE" if r==9 else f"R{r-1}"
            pP = "FALSE" if r==9 else f"P{r-1}"
            pS = "0" if r==9 else f"S{r-1}"
            pW = "0" if r==9 else f"W{r-1}"
            if r == 9:
                ws_data[f'P{r}']=False; ws_data[f'Q{r}']=False
            else:
                if strat=='V4':
                    ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}),FALSE,AND(E{r}>H{r},K{r}>(L{r}+$B$6),O{r}>$B$4,M{r}>$B$5,R{r}=FALSE))'
                    ws_data[f'Q{r}'] = f'=IF({pR}=TRUE,OR((B{r}-{pS})/{pS}>=$B$2,(B{r}-{pS})/{pS}<=-$B$3),FALSE)'
                elif strat=='V_Rebound':
                    ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}),FALSE,AND(E{r}<H{r},M{r}<$B$4,O{r}>O{r-1},R{r}=FALSE))'
                    ws_data[f'Q{r}'] = f'=IF({pR}=TRUE,OR((B{r}-{pS})/{pS}>=$B$2,(B{r}-{pS})/{pS}<=-$B$3,M{r-1}>$B$5),FALSE)'
                elif strat=='V_Dip_Buy':
                    ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}),FALSE,AND(M{r}<$B$4,K{r}>K{r-1},O{r}>0,R{r}=FALSE))'
                    ws_data[f'Q{r}'] = f'=IF({pR}=TRUE,OR((B{r}-{pS})/{pS}>=$B$2,(B{r}-{pS})/{pS}<=-$B$3,M{r-1}>$B$5),FALSE)'
            ws_data[f'R{r}'] = f'=IF({pR}=FALSE,{pP}=TRUE,IF(Q{r}=TRUE,FALSE,TRUE))'
            ws_data[f'S{r}'] = f'=IF(AND({pR}=FALSE,{pP}=TRUE),B{r},IF(AND({pR}=TRUE,Q{r}=FALSE),{pS},0))'
            ws_data[f'T{r}'] = f'=IF(AND({pR}=FALSE,{pP}=TRUE),"買進",IF(Q{r}=TRUE,"賣出",""))'
            ws_data[f'U{r}'] = f'=IF(T{r}="買進",B{r},IF(T{r}="賣出",B{r},0))'
            ws_data[f'V{r}'] = f'=IF(T{r}="賣出",(U{r}-{pS})*$B${shares_row},0)'
            ws_data[f'W{r}'] = f'={pW}+V{r}'
    
    wb.save(os.path.join(out_dir, f"{code}_{info['name']}_{strat}_高勝率回測.xlsx"))

# === Patch HTML: Top5 -> Top3 ===
with open(src_html, 'r', encoding='utf-8') as f: html = f.read()

# Replace info table content with top3 data
# First, change "Top 5" title to "Top 3"
html = html.replace('Top 5 最佳化組合', 'Top 3 最佳化組合')
html = html.replace('Top 5 最佳參數', 'Top 3 最佳參數')

# Remove Top4 and Top5 rows from each info table
import re
# Each Top 4/Top 5 row pattern: <tr><td>Top 4</td>.....</tr> or <tr><td>Top 5</td>.....</tr>
html = re.sub(r'<tr><td>Top 4</td>.*?</tr>', '', html)
html = re.sub(r'<tr><td>Top 5</td>.*?</tr>', '', html)

with open(os.path.join(out_dir, '技術分析測試_全個股_高勝率修正版.html'), 'w', encoding='utf-8') as f:
    f.write(html)

print(f"\n全部完成！輸出至: {out_dir}")
