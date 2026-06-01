import pandas as pd
import numpy as np
import itertools
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
import os

files = [
    "E:\\G-AI-1\\Stock analysis\\3443_創意_EOM_20260429.xlsx",
    "E:\\G-AI-1\\Stock analysis\\6269_台郡_EOM_20260429.xlsx"
]

def generate_comment(row, rank):
    if rank == 1:
        return "🏆 累積獲利最高：絕對金額最大，是捕捉大趨勢的最佳配置。"
    elif rank <= 5:
        if row['勝率'] > 0.6:
            return "⭐ 高勝率穩健型：勝率極佳，適合不想承受連續停損心理壓力的策略。"
        elif row['單筆最大虧損'] > -30000:
            return "🛡️ 風險控管優良：最大回檔低，單筆虧損控制得宜。"
        elif row['真實投報率(ROI)'] > 0.15:
            return "🚀 資金高效率：所需本金相對少，但能創造極高的投報率。"
        else:
            return "💡 綜合表現優異：在獲利與風險中取得良好平衡的參數組。"
    return ""

for input_file in files:
    base_name = os.path.basename(input_file).split('_EOM')[0]
    output_file = f"E:\\G-AI-1\\Stock analysis\\{base_name}_公式回測_可修改參數.xlsx"
    print(f"=============================================")
    print(f"正在處理: {base_name} ...")
    
    # 讀取資料
    df = pd.read_excel(input_file)
    open_col, close_col, ma20_col = df.columns[1], df.columns[4], df.columns[7]
    eom_col, eomsig_col = df.columns[10], df.columns[11]
    mfi_col, macd_col = df.columns[12], df.columns[14]
    
    for col in [open_col, close_col, ma20_col, eom_col, eomsig_col, mfi_col, macd_col]:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    open_arr = df[open_col].values
    close_arr = df[close_col].values
    ma20_arr = df[ma20_col].values
    eom_arr = df[eom_col].values
    eomsig_arr = df[eomsig_col].values
    macd_arr = df[macd_col].values
    mfi_arr = df[mfi_col].values
    len_df = len(df)
    shares = 1000

    # 建立 Workbook
    wb = Workbook()
    
    # ======== 1. 產生參數最佳化結果分頁 ========
    ws_opt = wb.active
    ws_opt.title = "參數最佳化結果"
    
    tps = [0.05, 0.07, 0.09, 0.12, 0.15]
    sls = [0.03, 0.046, 0.06, 0.08]
    macds = [-1.0, 0.0, 1.0]
    mfis = [30, 40, 50]
    
    results = []
    print("  計算最佳化參數組合...")
    for tp, sl, macd_th, mfi_th in itertools.product(tps, sls, macds, mfis):
        hold, buy_price = False, 0.0
        pnls, entry_prices = [], []
        
        for i in range(1, len_df - 1):
            if not hold:
                if close_arr[i] > ma20_arr[i] and eom_arr[i] > eomsig_arr[i] and macd_arr[i] > macd_th and mfi_arr[i] > mfi_th:
                    hold = True
                    buy_price = open_arr[i+1]
                    entry_prices.append(buy_price)
            else:
                n_open = open_arr[i+1]
                pnl_ratio = (n_open - buy_price) / buy_price if buy_price > 0 else 0
                if pnl_ratio >= tp or pnl_ratio <= -sl:
                    hold = False
                    sell_price = n_open
                    pnls.append((sell_price - buy_price) * shares)
                    buy_price = 0.0
        
        if len(pnls) > 0:
            tc = len(pnls)
            cp = sum(pnls)
            mc = max(entry_prices) * shares if entry_prices else 0
            results.append({
                'Take_Profit': tp, 'Stop_Loss': sl, 'MACD_Th': macd_th, 'MFI_Th': mfi_th,
                '累積淨損益總額': cp, '最大所需本金': mc,
                '真實投報率(ROI)': cp / mc if mc > 0 else 0,
                '交易次數': tc, '勝率': len([p for p in pnls if p > 0]) / tc,
                '單筆最大獲利': max(pnls), '單筆最大虧損': min(pnls),
                '平均每筆損益': cp / tc,
                '期望報酬率(每筆)': (cp / tc) / ((sum(entry_prices)/len(entry_prices))*shares) if entry_prices else 0
            })
            
    res_df = pd.DataFrame(results)
    res_df = res_df[res_df['交易次數'] >= 2].sort_values(by='累積淨損益總額', ascending=False).reset_index(drop=True)
    res_df['最佳參數註解'] = [generate_comment(row, i+1) for i, row in res_df.iterrows()]
    
    header_names = [
        '停利 %', '停損 %', 'MACD 閾值', 'MFI 閾值',
        '累積淨損益總額', '最大所需本金', '真實投報率(ROI)',
        '交易次數', '勝率', '單筆最大獲利', '單筆最大虧損',
        '平均每筆損益', '期望報酬率(每筆)', '🤖 AI 策略評估與註解'
    ]
    ws_opt.append(header_names)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E40AF")
    top5_fill = PatternFill("solid", fgColor="FEF08A")
    for cell in ws_opt[1]:
        cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="center")
        
    for r_idx, row in res_df.iterrows():
        ws_opt.append([row[col] for col in res_df.columns])
        cr = ws_opt.max_row
        ws_opt.cell(row=cr, column=1).number_format = '0.0%'
        ws_opt.cell(row=cr, column=2).number_format = '0.0%'
        for col in [5, 6, 10, 11, 12]: ws_opt.cell(row=cr, column=col).number_format = '#,##0'
        for col in [7, 9, 13]: ws_opt.cell(row=cr, column=col).number_format = '0.00%'
        
        if r_idx < 5:
            for col_num in range(1, 15):
                ws_opt.cell(row=cr, column=col_num).fill = top5_fill
                ws_opt.cell(row=cr, column=col_num).font = Font(bold=True)
                
    for col in ws_opt.columns:
        ml = max([len(str(cell.value)) for cell in col if cell.value] + [0])
        ws_opt.column_dimensions[col[0].column_letter].width = min(ml + 2, 40)
        
    # ======== 2. 產生動態公式回測分頁 ========
    ws_data = wb.create_sheet("Backtest Strategy")
    print("  建立動態公式與歷史資料...")
    
    # 如果最佳參數存在，取第一名，否則用預設值
    best_tp = res_df.iloc[0]['Take_Profit'] if len(res_df) > 0 else 0.09
    best_sl = res_df.iloc[0]['Stop_Loss'] if len(res_df) > 0 else 0.046
    best_macd = res_df.iloc[0]['MACD_Th'] if len(res_df) > 0 else 0.0
    best_mfi = res_df.iloc[0]['MFI_Th'] if len(res_df) > 0 else 40.0

    ws_data['A1'] = "【策略參數設定區】"; ws_data['B1'] = "數值 (可直接修改)"
    ws_data['A2'] = "停利 Take Profit (%)"; ws_data['B2'] = best_tp
    ws_data['A3'] = "停損 Stop Loss (%)"; ws_data['B3'] = best_sl
    ws_data['A4'] = "MACD 閾值"; ws_data['B4'] = best_macd
    ws_data['A5'] = "MFI 閾值"; ws_data['B5'] = best_mfi
    ws_data['A6'] = "交易張數 Shares"; ws_data['B6'] = 1000

    headers = list(df.columns) + ["進場訊號(P)", "出場訊號(Q)", "持倉狀態(R)", "進場成本(S)", "動作(T)", "交易價格(U)", "單筆損益(V)", "累積損益(W)"]
    for i, title in enumerate(headers, 1): ws_data.cell(row=8, column=i, value=title)
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 9):
        for c_idx, value in enumerate(row, 1): ws_data.cell(row=r_idx, column=c_idx, value=value)
        
    for r in range(9, ws_data.max_row + 1):
        prev_R, prev_P, prev_S, prev_W = ("FALSE", "FALSE", "0", "0") if r == 9 else (f"R{r-1}", f"P{r-1}", f"S{r-1}", f"W{r-1}")
        ws_data[f'P{r}'] = f'=IF(ISBLANK(E{r}), FALSE, AND(E{r}>H{r}, K{r}>L{r}, O{r}>$B$4, M{r}>$B$5, {prev_R}=FALSE))'
        ws_data[f'Q{r}'] = f'=IF({prev_R}=TRUE, OR((B{r}-{prev_S})/{prev_S} >= $B$2, (B{r}-{prev_S})/{prev_S} <= -$B$3), FALSE)'
        ws_data[f'R{r}'] = f'=IF({prev_R}=FALSE, {prev_P}=TRUE, IF(Q{r}=TRUE, FALSE, TRUE))'
        ws_data[f'S{r}'] = f'=IF(AND({prev_R}=FALSE, {prev_P}=TRUE), B{r}, IF(AND({prev_R}=TRUE, Q{r}=FALSE), {prev_S}, 0))'
        ws_data[f'T{r}'] = f'=IF(AND({prev_R}=FALSE, {prev_P}=TRUE), "買進", IF(Q{r}=TRUE, "賣出", ""))'
        ws_data[f'U{r}'] = f'=IF(T{r}="買進", B{r}, IF(T{r}="賣出", B{r}, 0))'
        ws_data[f'V{r}'] = f'=IF(T{r}="賣出", (U{r} - {prev_S}) * $B$6, 0)'
        ws_data[f'W{r}'] = f'={prev_W} + V{r}'

    for col in ['A', 'B']:
        for row in range(1, 7):
            ws_data[f'{col}{row}'].font = Font(bold=True)
            ws_data[f'{col}{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    wb.save(output_file)
    print(f"  > 輸出完成: {output_file}")

print("全部處理完畢！")
