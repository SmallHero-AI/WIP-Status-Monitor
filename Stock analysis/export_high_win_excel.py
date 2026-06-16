# -*- coding: utf-8 -*-
import os
import glob
import json
import shutil
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── 路徑設定 ──
SCRIPT_DIR = r"E:\G-AI-1\Stock analysis"
STOCK_ORIGINAL_DIR = os.path.join(SCRIPT_DIR, "Stock original")
AUTO_EXPORT_DIR = os.path.join(STOCK_ORIGINAL_DIR, "auto_export")
LEADERBOARD_PATH = os.path.join(SCRIPT_DIR, "修正版_V5_Server", "public", "leaderboard.json")
SERVER_DATA_DIR = os.path.join(SCRIPT_DIR, "修正版_V5_Server", "data")

def find_column(df, keywords):
    for col in df.columns:
        for kw in keywords:
            if kw.lower() in str(col).lower():
                return col
    return None

def clean_and_parse_series(df, col_name):
    if col_name is not None and col_name in df.columns:
        s = df[col_name].astype(str).str.replace(r'[▲▼↑↓△▽,%\s]', '', regex=True)
        return pd.to_numeric(s, errors='coerce').ffill().bfill().fillna(0).values
    return np.zeros(len(df))

def get_strategy_signals(df, ent_name, ext_name):
    """
    計算對應的進出場訊號 Boolean 陣列
    """
    n_rows = len(df)
    
    # 匹配技術指標欄位
    c_open = find_column(df, ['開盤價', '開盤', 'Open', 'open'])
    c_high = find_column(df, ['最高價', '最高', 'High', 'high'])
    c_low = find_column(df, ['最低價', '最低', 'Low', 'low'])
    c_close = find_column(df, ['收盤價', '收盤', 'Close', 'close'])
    c_ma5 = find_column(df, ['均價[5]', 'MA5', 'ma5'])
    c_ma20 = find_column(df, ['均價[20]', 'MA20', 'ma20'])
    c_ma60 = find_column(df, ['均價[60]', 'MA60', 'ma60'])
    c_ma120 = find_column(df, ['均價[120]', 'MA120', 'ma120'])
    c_eom = find_column(df, ['EOM[60]', 'EOM'])
    c_eom_sig = find_column(df, ['Signal[20]', 'EOM_Signal'])
    c_mfi = find_column(df, ['MFI[14]', 'MFI'])
    c_macd = find_column(df, ['MACD', 'MACD柱', 'MACD 柱', 'MACD 柱狀體'])
    c_k = find_column(df, ['%KS', 'K(9,3,3)', 'K值'])
    c_d = find_column(df, ['%DS', 'D(9,3,3)', 'D值'])
    c_rsi = find_column(df, ['RSI[14]', 'RSI'])
    c_bias5 = find_column(df, ['BIAS[5]', 'BIAS5'])
    c_bias20 = find_column(df, ['BIAS[20]', 'BIAS20'])
    c_cci = find_column(df, ['CCI[20]', 'CCI'])
    c_bb_upper = find_column(df, ['布林加通道 上漲', '布林加通道 上軌', '布林加通道 上', 'BB_Upper'])
    c_bb_lower = find_column(df, ['布林加通道 下跌', '布林加通道 下軌', '布林加通道 下', 'BB_Lower'])
    c_pivot = find_column(df, ['Pivot', 'pivot'])
    c_r1 = find_column(df, ['第1道壓力', 'R1'])
    c_s1 = find_column(df, ['第1道支撐', 'S1'])

    # 轉換陣列
    opens = clean_and_parse_series(df, c_open)
    highs = clean_and_parse_series(df, c_high)
    lows = clean_and_parse_series(df, c_low)
    closes = clean_and_parse_series(df, c_close)
    ma5 = clean_and_parse_series(df, c_ma5)
    ma20 = clean_and_parse_series(df, c_ma20)
    ma60 = clean_and_parse_series(df, c_ma60)
    ma120 = clean_and_parse_series(df, c_ma120)
    eom = clean_and_parse_series(df, c_eom)
    eom_sig = clean_and_parse_series(df, c_eom_sig)
    mfi = clean_and_parse_series(df, c_mfi)
    macd = clean_and_parse_series(df, c_macd)
    k = clean_and_parse_series(df, c_k)
    d = clean_and_parse_series(df, c_d)
    rsi = clean_and_parse_series(df, c_rsi)
    bias20 = clean_and_parse_series(df, c_bias20)
    cci = clean_and_parse_series(df, c_cci)
    bb_upper = clean_and_parse_series(df, c_bb_upper)
    bb_lower = clean_and_parse_series(df, c_bb_lower)
    pivot = clean_and_parse_series(df, c_pivot)
    r1 = clean_and_parse_series(df, c_r1)
    s1 = clean_and_parse_series(df, c_s1)

    prev_k = np.roll(k, 1); prev_k[0] = k[0]
    prev_d = np.roll(d, 1); prev_d[0] = d[0]
    prev_macd = np.roll(macd, 1); prev_macd[0] = macd[0]

    # 進場訊號定義
    f_kd_gold = (k > d) & (prev_k <= prev_d)
    f_rsi_oversold = rsi < 35
    f_rsi_bull = rsi > 50
    f_macd_grow = macd > prev_macd
    f_macd_positive = macd > 0
    f_bias_oversold = bias20 < -4
    f_cci_oversold = cci < -100
    f_eom_bullish = eom > eom_sig
    f_mfi_oversold = mfi < 30
    f_bb_oversold = (closes < bb_lower) & (bb_lower > 0)
    f_above_pivot = closes > pivot
    f_break_r1 = (closes > r1) & (r1 > 0)
    f_ma_bull = (ma5 > ma20) & (ma20 > ma60)
    f_above_ma20 = closes > ma20
    f_above_ma60 = closes > ma60
    f_above_ma120 = closes > ma120

    # 賣出基本訊號
    f_kd_dead = (k < d) & (prev_k >= prev_d)
    f_rsi_overbought = rsi > 65
    f_macd_shrink = macd < prev_macd
    f_below_ma20 = closes < ma20
    f_bb_overbought = (closes > bb_upper) & (bb_upper > 0)
    f_below_s1 = (closes < s1) & (s1 > 0)

    # 選擇進場訊號
    ent_sig = np.zeros(n_rows, dtype=bool)
    if ent_name == "強勢均線多頭+KD金叉": ent_sig = f_ma_bull & f_kd_gold & f_above_ma20
    elif ent_name == "中長期多頭+RSI低檔轉強": ent_sig = f_above_ma60 & f_above_ma120 & (rsi < 45) & f_macd_grow
    elif ent_name == "雙重均線突破+MACD翻紅": ent_sig = f_above_ma20 & f_above_ma60 & f_macd_positive & f_macd_grow
    elif ent_name == "EOM動能突破+均線多頭": ent_sig = f_eom_bullish & f_ma_bull & f_above_ma20
    elif ent_name == "極度超跌共振 (BIAS+RSI+MFI)": ent_sig = f_bias_oversold & f_rsi_oversold & f_mfi_oversold
    elif ent_name == "布林通道下軌+KD低檔金叉": ent_sig = f_bb_oversold & f_kd_gold
    elif ent_name == "CCI超賣+MFI超賣+KD金叉": ent_sig = f_cci_oversold & f_mfi_oversold & f_kd_gold
    elif ent_name == "雙重超賣 (RSI+CCI)+MACD轉強": ent_sig = f_rsi_oversold & f_cci_oversold & f_macd_grow
    elif ent_name == "樞軸點(S1)支撐+KD金叉": ent_sig = (closes > s1) & (lows <= s1) & f_kd_gold & (s1 > 0)
    elif ent_name == "突破阻力(R1)+MACD多頭": ent_sig = f_break_r1 & f_macd_positive & f_above_ma20
    elif ent_name == "突破樞軸(Pivot)+RSI轉強": ent_sig = f_above_pivot & f_rsi_bull & f_macd_grow
    elif ent_name == "S1支撐不破+MACD紅柱增長": ent_sig = (closes > s1) & f_macd_grow & (s1 > 0)
    elif ent_name == "多頭拉回：MA60之上+KD金叉+RSI<45": ent_sig = f_above_ma60 & f_kd_gold & (rsi < 45)
    elif ent_name == "布林中軌支撐+MACD紅柱": ent_sig = (closes > bb_lower) & (closes < bb_upper) & f_above_ma20 & f_macd_grow
    elif ent_name == "動能共振：EOM突破+MACD>0+RSI>50": ent_sig = f_eom_bullish & f_macd_positive & f_rsi_bull
    elif ent_name == "雙保險超跌：BIAS超賣+布林下軌觸及": ent_sig = f_bias_oversold & f_bb_oversold
    elif ent_name == "主力資金流入：MFI超賣+EOM突破": ent_sig = f_mfi_oversold & f_eom_bullish
    elif ent_name == "壓力突破：價格>R1+EOM強勢+MACD>0": ent_sig = f_break_r1 & f_eom_bullish & f_macd_positive
    elif ent_name == "長期均線支撐：站上MA120+KD金叉+CCI超賣": ent_sig = f_above_ma120 & f_kd_gold & f_cci_oversold
    elif ent_name == "CCI超賣反彈+MACD紅柱增長": ent_sig = f_cci_oversold & f_macd_grow & f_above_ma20

    # 選擇出場條件
    tp, sl = None, None
    ext_sig = np.zeros(n_rows, dtype=bool)
    if ext_name == "6%停利 / 6%停損 (穩健勝率)":
        tp, sl = 0.06, 0.06
    elif ext_name == "8%停利 / 8%停損 (均衡配置)":
        tp, sl = 0.08, 0.08
    elif ext_name == "12%停利 / 6%停損 (高盈虧比)":
        tp, sl = 0.12, 0.06
    elif ext_name == "KD死叉離場或6%停損":
        sl, ext_sig = 0.06, f_kd_dead
    elif ext_name == "RSI超買離場或6%停損":
        sl, ext_sig = 0.06, f_rsi_overbought
    elif ext_name == "MACD紅柱縮短離場或6%停損":
        sl, ext_sig = 0.06, f_macd_shrink
    elif ext_name == "收盤跌破MA20或6%停損":
        sl, ext_sig = 0.06, f_below_ma20
    elif ext_name == "觸碰布林上軌停利或6%停損":
        sl, ext_sig = 0.06, f_bb_overbought
    elif ext_name == "跌破樞軸支撐(S1)或8%停利":
        tp, ext_sig = 0.08, f_below_s1

    return ent_sig, ext_sig, tp, sl

def main():
    print("=" * 60)
    print("  AI 高勝率回測 Excel 匯出工具啟動...")
    print("=" * 60)

    if not os.path.exists(LEADERBOARD_PATH):
        print(f"[錯誤] 排行榜檔案 {LEADERBOARD_PATH} 不存在，請先執行回測。")
        return

    with open(LEADERBOARD_PATH, 'r', encoding='utf-8') as f:
        leaderboard = json.load(f)

    # 篩選勝率 >= 70% 的股票
    high_win = [x for x in leaderboard if x['winRate'] >= 70]
    print(f"[資訊] 共有 {len(high_win)} 檔股票勝率 >= 70%")

    # 尋找所有可用個股資料
    candidate_files = []
    for folder in [AUTO_EXPORT_DIR, STOCK_ORIGINAL_DIR]:
        if os.path.exists(folder):
            candidate_files.extend(glob.glob(os.path.join(folder, "*.csv")))
            candidate_files.extend(glob.glob(os.path.join(folder, "*.xlsx")))

    # 按代號映射最新檔案
    file_map = {}
    for f in candidate_files:
        filename = os.path.basename(f)
        parts = filename.split('_')
        if len(parts) >= 2 and parts[0].isdigit():
            code = parts[0]
            date_str = parts[-1].split('.')[0] if len(parts) >= 3 else ""
            if code not in file_map or date_str > file_map[code][1]:
                file_map[code] = (f, date_str)

    os.makedirs(SERVER_DATA_DIR, exist_ok=True)

    for item in high_win:
        code = item['code']
        name = item['name']
        strategy_full = item['strategy']
        win_rate = item['winRate']
        trades = item['trades']
        profit = item['profit']
        roi = item['roi']

        if code not in file_map:
            print(f"[警告] 找不到個股 {code} {name} 的資料檔案，跳過 Excel 匯出")
            continue

        file_path = file_map[code][0]
        filename = os.path.basename(file_path)

        # 讀取資料
        try:
            if file_path.endswith('.csv'):
                try:
                    df = pd.read_csv(file_path, encoding='cp950', header=0)
                except Exception:
                    df = pd.read_csv(file_path, encoding='utf-8', header=0)
            else:
                df = pd.read_excel(file_path, header=0)
        except Exception as ex:
            print(f"[警告] 無法讀取檔案 {filename}: {ex}")
            continue

        df.columns = [str(c).strip() for c in df.columns]
        df = df.tail(900).reset_index(drop=True)

        # 解析組合策略名稱
        # strategy_full 例如: "樞軸點(S1)支撐+KD金叉 & 收盤跌破MA20或6%停損"
        parts_strat = strategy_full.split(' & ')
        if len(parts_strat) != 2:
            print(f"[警告] 策略格式異常: {strategy_full}，跳過")
            continue
        ent_name = parts_strat[0].strip()
        ext_name = parts_strat[1].strip()

        # 計算進出場訊號
        ent_sig, ext_sig, tp, sl = get_strategy_signals(df, ent_name, ext_name)

        # 準備 Excel
        wb = Workbook()

        # Sheet 1: 參數與績效總覽
        ws_summary = wb.active
        ws_summary.title = "回測績效總覽"
        
        ws_summary.append(["AI 交叉組合策略回測報告", ""])
        ws_summary.append(["個股代碼", code])
        ws_summary.append(["個股名稱", name])
        ws_summary.append(["最佳組合策略", strategy_full])
        ws_summary.append(["進場訊號定義", ent_name])
        ws_summary.append(["出場條件定義", ext_name])
        ws_summary.append(["固定停利 (%)", f"{tp*100}%" if tp else "無"])
        ws_summary.append(["固定停損 (%)", f"{sl*100}%" if sl else "無"])
        ws_summary.append(["回測天數", f"過去 {len(df)} 天"])
        ws_summary.append(["總交易次數", f"{trades} 次"])
        ws_summary.append(["回測勝率 (%)", f"{win_rate:.2f}%"])
        ws_summary.append(["累積淨損益 (元)", f"${profit:,.0f} 元"])
        ws_summary.append(["真實投資報酬率 (ROI)", f"{roi:.2f}%"])
        ws_summary.append(["🤖 AI 策略評估與建議", "🏆 獲利能力與勝率極其卓越，適合中長期或高頻波段操作。"])

        # 樣式調整
        title_font = Font(name="微軟正黑體", size=14, bold=True, color="1E40AF")
        header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
        normal_font = Font(name="微軟正黑體", size=10)
        bold_font = Font(name="微軟正黑體", size=10, bold=True)
        header_fill = PatternFill("solid", fgColor="1E40AF")
        accent_fill = PatternFill("solid", fgColor="FEF08A")

        ws_summary.cell(row=1, column=1).font = title_font
        for r in range(2, 15):
            ws_summary.cell(row=r, column=1).font = bold_font
            ws_summary.cell(row=r, column=2).font = normal_font
            if r in [10, 11, 12, 13]:
                ws_summary.cell(row=r, column=1).fill = accent_fill
                ws_summary.cell(row=r, column=2).fill = accent_fill
                ws_summary.cell(row=r, column=2).font = bold_font

        # Sheet 2: 回測明細
        ws_data = wb.create_sheet("Top1_回測")

        # 寫入參數設定在 A1:B6
        ws_data['A1'] = "【AI 最佳化策略參數】"
        ws_data['A1'].font = bold_font
        ws_data['A2'] = "進場訊號"
        ws_data['B2'] = ent_name
        ws_data['A3'] = "出場條件"
        ws_data['B3'] = ext_name
        ws_data['A4'] = "停利門檻 (tp)"
        ws_data['B4'] = tp if tp else 0
        ws_data['A5'] = "停損門檻 (sl)"
        ws_data['B5'] = sl if sl else 0
        ws_data['A6'] = "交易張數 Shares"
        ws_data['B6'] = 1000

        for r in range(2, 7):
            ws_data[f'A{r}'].font = bold_font
            ws_data[f'B{r}'].font = normal_font

        # 寫入歷史資料欄位與公式
        num_cols = len(df.columns)
        headers = list(df.columns) + [
            "進場訊號(P)", "出場訊號(Q)", "持倉狀態(R)", "進場成本(S)", 
            "動作(T)", "交易價格(U)", "單筆損益(V)", "累積損益(W)"
        ]

        # 寫入欄位名稱
        for i, t in enumerate(headers, 1):
            cell = ws_data.cell(row=8, column=i, value=t)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 寫入歷史價格與指標資料
        for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 9):
            for c_idx, val in enumerate(row_data, 1):
                # 清理與數值轉換
                cell_val = val
                if c_idx > 0: # 對價格與指標進行浮點數轉換
                    try:
                        if isinstance(val, str):
                            val_clean = val.replace('%', '').replace(',', '').strip()
                            cell_val = float(val_clean)
                        else:
                            cell_val = float(val)
                    except ValueError:
                        cell_val = val
                ws_data.cell(row=r_idx, column=c_idx, value=cell_val)

        # 取得公式所屬的欄位字母
        col_P = get_column_letter(num_cols + 1)
        col_Q = get_column_letter(num_cols + 2)
        col_R = get_column_letter(num_cols + 3)
        col_S = get_column_letter(num_cols + 4)
        col_T = get_column_letter(num_cols + 5)
        col_U = get_column_letter(num_cols + 6)
        col_V = get_column_letter(num_cols + 7)
        col_W = get_column_letter(num_cols + 8)

        # 寫入進出場訊號狀態與公式
        total_rows = ws_data.max_row
        for r in range(9, total_rows + 1):
            idx = r - 9
            # 寫入訊號 (由 Python 先行運算)
            ws_data[f'{col_P}{r}'] = bool(ent_sig[idx])
            ws_data[f'{col_Q}{r}'] = bool(ext_sig[idx])

            # 寫入交易模擬公式 (同前天開盤延遲一日買賣狀態機)
            pR = "FALSE" if r == 9 else f"{col_R}{r-1}"
            pP = "FALSE" if r == 9 else f"{col_P}{r-1}"
            pQ = "FALSE" if r == 9 else f"{col_Q}{r-1}"
            pS = "0" if r == 9 else f"{col_S}{r-1}"
            pW = "0" if r == 9 else f"{col_W}{r-1}"

            # 持倉狀態(R) AJ
            ws_data[f'{col_R}{r}'] = f'=IF({pR}=FALSE, {pP}=TRUE, IF({pQ}=TRUE, FALSE, TRUE))'
            # 進場成本(S) AK (以開盤價 B 欄為交易基準)
            ws_data[f'{col_S}{r}'] = f'=IF(AND({pR}=FALSE, {col_R}{r}=TRUE), B{r}, IF(AND({pR}=TRUE, {col_R}{r}=TRUE), {pS}, 0))'
            # 動作(T) AL
            ws_data[f'{col_T}{r}'] = f'=IF(AND({pR}=FALSE, {col_R}{r}=TRUE), "買進", IF(AND({pR}=TRUE, {col_R}{r}=FALSE), "賣出", ""))'
            # 交易價格(U) AM
            ws_data[f'{col_U}{r}'] = f'=IF({col_T}{r}="買進", B{r}, IF({col_T}{r}="賣出", B{r}, 0))'
            # 單筆損益(V) AN
            ws_data[f'{col_V}{r}'] = f'=IF({col_T}{r}="賣出", ({col_U}{r} - {pS}) * $B$6, 0)'
            # 累積損益(W) AO
            ws_data[f'{col_W}{r}'] = f'={pW} + {col_V}{r}'

            # 格式套用
            for col_letter in [col_R, col_S, col_T, col_U, col_V, col_W]:
                ws_data[f'{col_letter}{r}'].font = normal_font

        # 自動調整欄寬
        for ws in [ws_summary, ws_data]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # 儲存 Excel 到 2 個位置 (一個給使用者，一個給網頁伺服器)
        excel_name = f"{code}_{name}_AI_高勝率回測.xlsx"
        user_out_path = os.path.join(SCRIPT_DIR, excel_name)
        server_out_path = os.path.join(SERVER_DATA_DIR, excel_name)
        
        wb.save(user_out_path)
        shutil.copy2(user_out_path, server_out_path)
        print(f"  [成功] 產出並複製：{excel_name}")

    print("\n[完成] 所有勝率 70% 以上個股的回測 EXCEL 檔案已成功產出與複製！")

if __name__ == "__main__":
    main()
