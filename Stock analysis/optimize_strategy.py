import pandas as pd
import numpy as np
import itertools
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

input_file = "E:\\G-AI-1\\Stock analysis\\2330_台積電_EOM_20260430.xlsx"
target_excel = "E:\\G-AI-1\\Stock analysis\\2330_台積電_公式回測_可修改參數.xlsx"

print("讀取原始資料...")
df = pd.read_excel(input_file)

# 為了避免中文編碼錯誤，直接用 iloc 對應
open_col = df.columns[1]
close_col = df.columns[4]
ma20_col = df.columns[7]
eom_col = df.columns[10]
eomsig_col = df.columns[11]
mfi_col = df.columns[12]
macd_col = df.columns[14]

for col in [open_col, close_col, ma20_col, eom_col, eomsig_col, mfi_col, macd_col]:
    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

open_arr = df[open_col].values
close_arr = df[close_col].values
ma20_arr = df[ma20_col].values
eom_arr = df[eom_col].values
eomsig_arr = df[eomsig_col].values
macd_arr = df[macd_col].values
mfi_arr = df[mfi_col].values
len_df = len(df)

shares = 1000

# 參數網格
tps = [0.05, 0.07, 0.09, 0.12, 0.15]
sls = [0.03, 0.046, 0.06, 0.08]
macds = [-1.0, 0.0, 1.0]
mfis = [30, 40, 50]

print("開始計算最佳化參數組合...")
results = []

for tp, sl, macd_th, mfi_th in itertools.product(tps, sls, macds, mfis):
    hold = False
    buy_price = 0.0
    pnls = []
    entry_prices = []
    
    for i in range(1, len_df - 1):
        if not hold:
            if close_arr[i] > ma20_arr[i] and eom_arr[i] > eomsig_arr[i] and macd_arr[i] > macd_th and mfi_arr[i] > mfi_th:
                hold = True
                buy_price = open_arr[i+1]
                entry_prices.append(buy_price)
        else:
            n_open = open_arr[i+1]
            pnl_ratio = (n_open - buy_price) / buy_price if buy_price > 0 else 0
            if pnl_ratio >= tp or pnl_ratio <= -sl:
                hold = False
                sell_price = n_open
                pnls.append((sell_price - buy_price) * shares)
                buy_price = 0.0

    if len(pnls) > 0:
        trade_count = len(pnls)
        cum_pnl = sum(pnls)
        max_cap = max(entry_prices) * shares if entry_prices else 0
        roi = cum_pnl / max_cap if max_cap > 0 else 0
        wins = [p for p in pnls if p > 0]
        win_rate = len(wins) / trade_count if trade_count > 0 else 0
        max_profit = max(pnls)
        max_loss = min(pnls)
        avg_pnl = cum_pnl / trade_count
        avg_entry = (sum(entry_prices) / len(entry_prices)) * shares if entry_prices else 0
        expected_return = avg_pnl / avg_entry if avg_entry > 0 else 0
        
        results.append({
            'Take_Profit': tp,
            'Stop_Loss': sl,
            'MACD_Th': macd_th,
            'MFI_Th': mfi_th,
            '累積淨損益總額': cum_pnl,
            '最大所需本金': max_cap,
            '真實投報率(ROI)': roi,
            '交易次數': trade_count,
            '勝率': win_rate,
            '單筆最大獲利': max_profit,
            '單筆最大虧損': max_loss,
            '平均每筆損益': avg_pnl,
            '期望報酬率(每筆)': expected_return
        })

res_df = pd.DataFrame(results)

# 篩選至少有 2 次交易的組合以避免極端值
res_df = res_df[res_df['交易次數'] >= 2]

# 綜合評分排序 (主要看累積損益)
res_df = res_df.sort_values(by='累積淨損益總額', ascending=False).reset_index(drop=True)

# 產生註解邏輯
def generate_comment(row, rank):
    if rank == 1:
        return "🏆 累積獲利最高：絕對金額最大，是捕捉大趨勢的最佳配置。"
    elif rank <= 5:
        if row['勝率'] > 0.6:
            return "⭐ 高勝率穩健型：勝率極佳，適合不想承受連續停損心理壓力的策略。"
        elif row['單筆最大虧損'] > -30000: # 虧損較小
            return "🛡️ 風險控管優良：最大回檔低，單筆虧損控制得宜。"
        elif row['真實投報率(ROI)'] > 0.15:
            return "🚀 資金高效率：所需本金相對少，但能創造極高的投報率。"
        else:
            return "💡 綜合表現優異：在獲利與風險中取得良好平衡的參數組。"
    return ""

res_df['最佳參數註解'] = [generate_comment(row, i+1) for i, row in res_df.iterrows()]

# 將結果寫入 Excel
print("正在寫入 Excel 新分頁...")
try:
    wb = load_workbook(target_excel)
except Exception:
    print("找不到目標 Excel 檔案，將建立新檔案。")
    from openpyxl import Workbook
    wb = Workbook()

if "參數最佳化結果" in wb.sheetnames:
    del wb["參數最佳化結果"]

ws = wb.create_sheet("參數最佳化結果", 0) # 放到第一個分頁

# 寫入標題
columns_order = [
    'Take_Profit', 'Stop_Loss', 'MACD_Th', 'MFI_Th', 
    '累積淨損益總額', '最大所需本金', '真實投報率(ROI)', 
    '交易次數', '勝率', '單筆最大獲利', '單筆最大虧損', 
    '平均每筆損益', '期望報酬率(每筆)', '最佳參數註解'
]

# 中文顯示標題
header_names = [
    '停利 %', '停損 %', 'MACD 閾值', 'MFI 閾值',
    '累積淨損益總額', '最大所需本金', '真實投報率(ROI)',
    '交易次數', '勝率', '單筆最大獲利', '單筆最大虧損',
    '平均每筆損益', '期望報酬率(每筆)', '🤖 AI 策略評估與註解'
]

ws.append(header_names)

# 樣式設定
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1E40AF")
top5_fill = PatternFill("solid", fgColor="FEF08A") # 黃色標記前五名

for col_num, cell in enumerate(ws[1], 1):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# 寫入資料
for r_idx, row in res_df.iterrows():
    row_data = [row[col] for col in columns_order]
    ws.append(row_data)
    
    current_row = ws.max_row
    
    # 數值格式化
    ws.cell(row=current_row, column=1).number_format = '0.0%'
    ws.cell(row=current_row, column=2).number_format = '0.0%'
    ws.cell(row=current_row, column=5).number_format = '#,##0'
    ws.cell(row=current_row, column=6).number_format = '#,##0'
    ws.cell(row=current_row, column=7).number_format = '0.00%'
    ws.cell(row=current_row, column=9).number_format = '0.00%'
    ws.cell(row=current_row, column=10).number_format = '#,##0'
    ws.cell(row=current_row, column=11).number_format = '#,##0'
    ws.cell(row=current_row, column=12).number_format = '#,##0'
    ws.cell(row=current_row, column=13).number_format = '0.00%'

    # 前五名黃色標記
    if r_idx < 5:
        for col_num in range(1, 15):
            ws.cell(row=current_row, column=col_num).fill = top5_fill
            ws.cell(row=current_row, column=col_num).font = Font(bold=True)

# 自動調整欄寬
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    if adjusted_width > 40: adjusted_width = 40
    ws.column_dimensions[col_letter].width = adjusted_width

wb.save(target_excel)
print("參數最佳化完成！")
